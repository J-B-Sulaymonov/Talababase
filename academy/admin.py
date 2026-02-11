import json
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime, date

from django.contrib import admin
from django.urls import reverse
from django.db.models import (
    Sum, F, OuterRef, Subquery, Q, Value, Case, When,
    DecimalField, Count, Prefetch
)
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.urls import path, reverse
from django.utils.html import format_html
from django.utils.translation import gettext_lazy as _
from django.shortcuts import render

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .models import (
    WeekDay, Level, Teacher, Group,
    Student, Enrollment, StudentPayment, TeacherSalary
)


# =============================================================================
# 📅 HAFTA KUNLARI
# =============================================================================
@admin.register(WeekDay)
class WeekDayAdmin(admin.ModelAdmin):
    list_display = ('name', 'order')
    list_display_links = ('name',)
    ordering = ('order',)
    search_fields = ('name',)


# =============================================================================
# 📊 DARAJALAR
# =============================================================================
@admin.register(Level)
class LevelAdmin(admin.ModelAdmin):
    list_display = ('name', 'id')
    search_fields = ('name',)


# =============================================================================
# 👨‍🏫 O'QITUVCHILAR
# =============================================================================
# =============================================================================
# 👨‍🏫 O'QITUVCHILAR
# =============================================================================

class TeacherSalaryInline(admin.StackedInline):
    model = TeacherSalary
    extra = 0
    fields = ('amount', 'payment_date', 'month_for', 'type', 'payment_method', 'comment')

@admin.register(Teacher)
class TeacherAdmin(admin.ModelAdmin):
    list_display = ('full_name_display', 'phone_number', 'is_active')
    list_filter = ('is_active',)
    search_fields = ('first_name', 'last_name', 'phone_number')
    list_per_page = 50
    inlines = [TeacherSalaryInline]

    @admin.display(description="F.I.SH", ordering='first_name')
    def full_name_display(self, obj):
        return f"{obj.first_name} {obj.last_name or ''}"


# =============================================================================
# 👥 GURUHLAR
# =============================================================================
@admin.register(Group)
class GroupAdmin(admin.ModelAdmin):
    change_list_template = "admin/academy/group/change_list.html"

    list_display = (
        'name', 'level_badge', 'teacher_display', 'formatted_price',
        'formatted_teacher_price', 'get_monthly_revenue',
        'lessons_per_month', 'display_payment_per_lesson', 'display_teacher_share',
        'get_days_display', 'lesson_time', 'student_count_display',
        'status_badge', 'view_group_link',
    )
    list_filter = ('is_active', 'level', 'teacher')
    search_fields = ('name', 'teacher__first_name', 'teacher__last_name')
    filter_horizontal = ('days',)
    list_per_page = 50
    show_full_result_count = False
    actions = None

    class Media:
        css = {
            'all': ('admin/css/custom_scroll.css',)
        }

    def get_queryset(self, request):
        qs = super().get_queryset(request).select_related('level', 'teacher').prefetch_related('days')
        qs = qs.annotate(
            student_count=Count('enrollments', filter=Q(enrollments__is_active=True))
        )

        # JS filterlar uchun comma-separated qiymatlarni qo'llab-quvvatlash
        name_filter = getattr(request, '_custom_name_filter', '')
        if name_filter:
            name_ids = [v for v in name_filter.split(',') if v]
            if name_ids:
                qs = qs.filter(id__in=name_ids)

        level_filter = request.GET.get('level', '')
        if level_filter and ',' in level_filter:
            level_ids = [v for v in level_filter.split(',') if v]
            if level_ids:
                qs = qs.filter(level__id__in=level_ids)

        teacher_filter = request.GET.get('teacher', '')
        if teacher_filter and ',' in teacher_filter:
            teacher_ids = [v for v in teacher_filter.split(',') if v]
            if teacher_ids:
                qs = qs.filter(teacher__id__in=teacher_ids)

        is_active_filter = getattr(request, '_custom_is_active_filter', '')
        if is_active_filter:
            if is_active_filter in ('true', 'false'):
                qs = qs.filter(is_active=(is_active_filter == 'true'))
            elif ',' in is_active_filter:
                # Agar ikkala qiymat tanlangan bo'lsa — filter qo'llanmaydi
                pass

        return qs

    # ---------------------------------------------------------
    # CUSTOM URLS
    # ---------------------------------------------------------
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('export-excel/', self.admin_site.admin_view(self.export_groups_excel),
                 name='export_groups_excel'),
        ]
        return custom_urls + urls

    # ---------------------------------------------------------
    # STATISTIKA PANELI + FILTER DATA
    # ---------------------------------------------------------
    def changelist_view(self, request, extra_context=None):
        # Custom JS filter params — Django admin tanimaydi
        # request attribute sifatida saqlab, GET dan tozalaymiz
        request._custom_name_filter = request.GET.get('name', '')
        request._custom_is_active_filter = request.GET.get('is_active', '')
        
        need_copy = False
        if 'name' in request.GET or request._custom_is_active_filter in ('true', 'false'):
            if not need_copy:
                request.GET = request.GET.copy()
                need_copy = True
            if 'name' in request.GET:
                request.GET.pop('name')
            if request._custom_is_active_filter in ('true', 'false'):
                request.GET.pop('is_active')

        # Filter data for JS filters
        groups_data = list(Group.objects.filter(is_active=True).values('id', 'name').order_by('name'))
        levels_data = list(Level.objects.values('id', 'name').order_by('name'))
        teachers = Teacher.objects.filter(is_active=True).only('id', 'first_name', 'last_name')
        teachers_data = [{'id': t.id, 'name': f"{t.first_name} {t.last_name or ''}".strip()} for t in teachers]
        teachers_data.sort(key=lambda x: x['name'])
        status_data = [
            {'id': 'true', 'name': 'Faol'},
            {'id': 'false', 'name': 'Nofaol'},
        ]

        extra_context = extra_context or {}
        extra_context['groups_json'] = json.dumps(groups_data)
        extra_context['levels_json'] = json.dumps(levels_data)
        extra_context['teachers_json'] = json.dumps(teachers_data)
        extra_context['status_json'] = json.dumps(status_data)

        response = super().changelist_view(request, extra_context)

        if not hasattr(response, 'context_data') or 'cl' not in response.context_data:
            return response

        cl = response.context_data['cl']
        qs = cl.queryset

        from decimal import Decimal

        total_students = 0
        monthly_revenue = Decimal(0)
        teacher_expense = Decimal(0)
        active_groups = 0

        for obj in qs:
            count = getattr(obj, 'student_count', 0)
            total_students += count
            if obj.is_active:
                active_groups += 1
                monthly_revenue += (obj.price or Decimal(0)) * count
                teacher_expense += obj.teacher_price or Decimal(0)

        net_profit = monthly_revenue - teacher_expense

        response.context_data['footer_stats'] = {
            'active_groups': active_groups,
            'total_students': total_students,
            'monthly_revenue': monthly_revenue,
            'teacher_expense': teacher_expense,
            'net_profit': net_profit,
        }

        return response

    # ---------------------------------------------------------
    # EXCEL EXPORT
    # ---------------------------------------------------------
    def export_groups_excel(self, request):
        """Guruhlar ro'yxatini Excelga export qilish"""
        qs = self.get_queryset(request)

        selected_fields = request.POST.getlist('selected_fields')
        if not selected_fields:
            selected_fields = [
                'name', 'level', 'teacher', 'price', 'teacher_price',
                'monthly_revenue', 'lessons_per_month', 'days', 'lesson_time',
                'student_count', 'is_active',
            ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Guruhlar Export"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        money_format = '#,##0'

        field_titles = {
            'name': 'Guruh kodi',
            'level': 'Daraja',
            'teacher': "O'qituvchi",
            'price': 'Narxi (oylik)',
            'teacher_price': "O'qituvchi narxi",
            'monthly_revenue': 'Oylik daromad',
            'lessons_per_month': 'Darslar soni',
            'days': 'Dars kunlari',
            'lesson_time': 'Dars vaqti',
            'student_count': "O'quvchilar",
            'is_active': 'Holati',
        }

        money_fields = ['price', 'teacher_price', 'monthly_revenue']

        for col_num, field in enumerate(selected_fields, 1):
            column_letter = get_column_letter(col_num)
            cell = ws.cell(row=1, column=col_num)
            cell.value = field_titles.get(field, field)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            ws.column_dimensions[column_letter].width = 20

        row_num = 2
        for obj in qs:
            count = getattr(obj, 'student_count', 0)
            for col_num, field in enumerate(selected_fields, 1):
                cell = ws.cell(row=row_num, column=col_num)
                val = None

                if field == 'name':
                    val = obj.name
                elif field == 'level':
                    val = obj.level.name if obj.level else '-'
                elif field == 'teacher':
                    val = f"{obj.teacher.first_name} {obj.teacher.last_name or ''}" if obj.teacher else '-'
                elif field == 'price':
                    val = int(obj.price) if obj.price else 0
                elif field == 'teacher_price':
                    val = int(obj.teacher_price) if obj.teacher_price else 0
                elif field == 'monthly_revenue':
                    val = int((obj.price or 0) * count)
                elif field == 'lessons_per_month':
                    val = obj.lessons_per_month
                elif field == 'days':
                    val = ", ".join([d.name for d in obj.days.all()])
                elif field == 'lesson_time':
                    start = obj.lesson_time_start.strftime('%H:%M') if obj.lesson_time_start else ''
                    end = obj.lesson_time_end.strftime('%H:%M') if obj.lesson_time_end else ''
                    val = f"{start} - {end}" if start and end else '-'
                elif field == 'student_count':
                    val = count
                elif field == 'is_active':
                    val = 'Faol' if obj.is_active else 'Nofaol'

                if val is None:
                    val = ""

                if field in money_fields:
                    try:
                        cell.value = int(float(val)) if val else 0
                        cell.number_format = money_format
                    except (ValueError, TypeError):
                        cell.value = val
                else:
                    cell.value = val if not isinstance(val, str) else str(val)

                cell.border = thin_border
                cell.alignment = center_align

            row_num += 1

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        response['Content-Disposition'] = f'attachment; filename=Academy_Groups_{timestamp}.xlsx'
        wb.save(response)
        return response

    # ---------------------------------------------------------
    # DISPLAY METODLARI
    # ---------------------------------------------------------

    @admin.display(description="Daraja", ordering='level')
    def level_badge(self, obj):
        if obj.level:
            return format_html(
                '<span class="group-badge" style="background:#e7f5ff; color:#1c7ed6; border-color:#a5d8ff;">{}</span>',
                obj.level.name
            )
        return "-"

    @admin.display(description="O'qituvchi", ordering='teacher')
    def teacher_display(self, obj):
        if obj.teacher:
            name = f"{obj.teacher.first_name} {obj.teacher.last_name or ''}"
            return format_html(
                '<span class="group-badge" style="background:#fff4e6; color:#e8590c; border-color:#ffd8a8;">{}</span>',
                name
            )
        return format_html('<span style="color:#adb5bd;">—</span>')

    @admin.display(description="Narxi (oylik)")
    def formatted_price(self, obj):
        if obj.price:
            return format_html(
                '<span style="font-weight:600; color:#2b8a3e;">{}</span>',
                f"{obj.price:,.0f}".replace(",", " ")
            )
        return "-"

    @admin.display(description="O'qituvchi narxi")
    def formatted_teacher_price(self, obj):
        if obj.teacher_price:
            return format_html(
                '<span style="font-weight:600; color:#e64980;">{}</span>',
                f"{obj.teacher_price:,.0f}".replace(",", " ")
            )
        return "-"

    @admin.display(description="Oylik daromad", ordering='student_count')
    def get_monthly_revenue(self, obj):
        count = getattr(obj, 'student_count', 0)
        if count > 0 and obj.price:
            revenue = obj.price * count
            return format_html(
                '<span style="font-weight:700; color:#0ca678;">{}</span>',
                f"{revenue:,.0f}".replace(",", " ")
            )
        return format_html('<span style="color:#adb5bd;">0</span>')

    @admin.display(description="Dars kunlari")
    def get_days_display(self, obj):
        days = obj.days.all()
        if days:
            return ", ".join([d.name for d in days])
        return "-"

    @admin.display(description="Dars vaqti")
    def lesson_time(self, obj):
        start = obj.lesson_time_start.strftime('%H:%M') if obj.lesson_time_start else ''
        end = obj.lesson_time_end.strftime('%H:%M') if obj.lesson_time_end else ''
        if start and end:
            return format_html(
                '<span style="font-weight:500; color:#495057;">{} - {}</span>',
                start, end
            )
        return "-"

    @admin.display(description="O'quvchilar", ordering='student_count')
    def student_count_display(self, obj):
        count = getattr(obj, 'student_count', 0)
        if count > 0:
            return format_html(
                '<span class="status-badge badge-info">{} ta</span>',
                count
            )
        return format_html('<span style="color:#adb5bd;">0</span>')

    @admin.display(description="Holati")
    def status_badge(self, obj):
        if obj.is_active:
            return format_html(
                '<span class="status-badge badge-success">✓ Faol</span>'
            )
        return format_html(
            '<span class="status-badge badge-danger">✗ Nofaol</span>'
        )

    @admin.display(description="Ko'rish")
    def view_group_link(self, obj):
        url = reverse('admin:academy_student_changelist') + f"?group_filter={obj.pk}"
        return format_html(
            '<a href="{}" title="Guruh o\'quvchilarini ko\'rish" style="color: #adb5bd; font-size: 1.2rem; transition: 0.2s;">'
            '<i style="color:blue" class="fas fa-eye"></i>'
            '</a>',
            url
        )

    # === MOLIYAVIY KO'RSATKICHLAR ===

    @admin.display(description="Dars uchun to'lov")
    def display_payment_per_lesson(self, obj):
        """Har bir dars uchun to'lov"""
        payment = obj.payment_per_lesson()
        if payment > 0:
            return format_html(
                '<span style="font-weight:600; color:#7950f2;">{}</span>',
                f"{payment:,.0f}".replace(",", " ")
            )
        return format_html('<span style="color:#adb5bd;">—</span>')

    @admin.display(description="Ulushi %")
    def display_teacher_share(self, obj):
        """O'rtacha to'lanadigan ulush (foizda)"""
        share = obj.teacher_share_percentage()
        if share > 0:
            # Rang: yashil (past) -> qizil (yuqori)
            if share < 30:
                color = "#2b8a3e"  # Yashil
            elif share < 50:
                color = "#e8590c"  # To'q sariq
            else:
                color = "#c92a2a"  # Qizil
            return format_html(
                '<span style="font-weight:700; color:{};">{}</span>',
                color, f"{share:.1f}%"
            )
        return format_html('<span style="color:#adb5bd;">—</span>')


# =============================================================================
# 💰 O'QITUVCHI MAOSHLARI
# =============================================================================
@admin.register(TeacherSalary)
class TeacherSalaryAdmin(admin.ModelAdmin):
    list_display = ('teacher', 'formatted_amount', 'payment_date', 'month_for_display', 'type', 'payment_method')
    list_filter = ('type', 'payment_method', 'teacher', 'payment_date')
    search_fields = ('teacher__first_name', 'teacher__last_name')
    date_hierarchy = 'payment_date'
    list_per_page = 50

    @admin.display(description="Summa")
    def formatted_amount(self, obj):
        if obj.amount:
            return format_html(
                '<span style="font-weight:700; color:#e64980;">{}</span>',
                f"{obj.amount:,.0f}".replace(",", " ")
            )
        return "-"

    @admin.display(description="Qaysi oy uchun")
    def month_for_display(self, obj):
        if obj.month_for:
            return obj.month_for.strftime('%B %Y')
        return "-"


# =============================================================================
# 💳 O'QUVCHI TO'LOVLARI
# =============================================================================
@admin.register(StudentPayment)
class StudentPaymentAdmin(admin.ModelAdmin):
    list_display = ('student', 'group', 'formatted_amount', 'payment_date', 'month_for_display', 'payment_type')
    list_filter = ('payment_type', 'group', 'payment_date')
    search_fields = ('student__full_name', 'group__name')
    date_hierarchy = 'payment_date'
    autocomplete_fields = ['student', 'group']
    list_per_page = 50

    @admin.display(description="Summa")
    def formatted_amount(self, obj):
        if obj.amount:
            return format_html(
                '<span style="font-weight:700; color:#2b8a3e;">{}</span>',
                f"{obj.amount:,.0f}".replace(",", " ")
            )
        return "-"

    @admin.display(description="Qaysi oy uchun")
    def month_for_display(self, obj):
        if obj.month_for:
            return obj.month_for.strftime('%B %Y')
        return "-"


# =============================================================================
# 📋 KURSGA QABUL (Enrollment) Admin
# =============================================================================
@admin.register(Enrollment)
class EnrollmentAdmin(admin.ModelAdmin):
    list_display = (
        'student', 'group', 'enrolled_date', 'left_date',
        'get_months_display', 'get_leave_reason_display_col', 'is_active'
    )
    list_filter = ('is_active', 'group', 'leave_reason')
    search_fields = ('student__full_name', 'group__name')
    autocomplete_fields = ['student', 'group']
    list_per_page = 50

    @admin.display(description="Oy")
    def get_months_display(self, obj):
        months = obj.months_enrolled()
        return format_html(
            '<span style="font-weight:600; color:#1c7ed6;">{} oy</span>', months
        )

    @admin.display(description="Sababi")
    def get_leave_reason_display_col(self, obj):
        if obj.leave_reason:
            return obj.get_leave_reason_display()
        return "-"


# =============================================================================
# 🔍 O'QUVCHI FILTERLARI
# =============================================================================

class StatusFilter(admin.SimpleListFilter):
    """Faol / Nofaol o'quvchilar filteri (Ko'plik tanlash imkoniyati bilan)"""
    title = "Status"
    parameter_name = 'status'

    def lookups(self, request, model_admin):
        return (
            ('active', "Faol"),
            ('inactive', "Nofaol"),
            ('all', "Barcha o'quvchilar"),
        )

    def choices(self, cl):
        # Admin panelning o'ng tarafidagi standart filter uchun
        for lookup, title in self.lookup_choices:
            yield {
                'selected': self.value() == str(lookup) or (self.value() is None and str(lookup) == 'active'),
                'query_string': cl.get_query_string({self.parameter_name: lookup}, []),
                'display': title,
            }

    def queryset(self, request, queryset):
        value = self.value()
        if not value:
            # Default holatda faqat aktivlar
            return queryset.filter(status='active')
        
        values = value.split(',')
        if 'all' in values:
            return queryset
            
        # Agar faqat 'active' yoki 'inactive' yoki ikkalasi bo'lsa
        return queryset.filter(status__in=values)


class AcademyGroupFilter(admin.SimpleListFilter):
    """Guruh bo'yicha filter (Enrollment orqali)"""
    title = "Guruh"
    parameter_name = 'group_filter'

    def lookups(self, request, model_admin):
        return [(g.id, g.name) for g in Group.objects.filter(is_active=True)]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(enrollments__group__id__in=self.value().split(','))
        return queryset


class AcademyLevelFilter(admin.SimpleListFilter):
    """Daraja bo'yicha filter"""
    title = "Daraja"
    parameter_name = 'level_filter'

    def lookups(self, request, model_admin):
        return [(l.id, l.name) for l in Level.objects.all()]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(enrollments__group__level__id__in=self.value().split(','))
        return queryset


class AcademyTeacherFilter(admin.SimpleListFilter):
    """O'qituvchi bo'yicha filter"""
    title = "O'qituvchi"
    parameter_name = 'teacher_filter'

    def lookups(self, request, model_admin):
        return [(t.id, f"{t.first_name} {t.last_name}") for t in Teacher.objects.filter(is_active=True)]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(enrollments__group__teacher__id__in=self.value().split(','))
        return queryset



class EnrollmentInline(admin.StackedInline):
    """O'quvchining kursga qabul qilinishi / chiqarilishi"""
    model = Enrollment
    extra = 1
    fields = ('group', 'enrolled_date', 'left_date', 'leave_reason', 'leave_comment', 'is_active')
    autocomplete_fields = ['group']
    verbose_name = "Kursga qabul"
    verbose_name_plural = "Kursga qabul / chiqarish"


class StudentPaymentInline(admin.StackedInline):
    """O'quvchi to'lovlarini uning sahifasida ko'rsatish"""
    model = StudentPayment
    extra = 0
    fields = ('group', 'amount', 'payment_date', 'month_for', 'payment_type', 'comment')
    autocomplete_fields = ['group']
    verbose_name = "To'lov"
    verbose_name_plural = "To'lovlar"


@admin.register(Student)
class AcademyStudentAdmin(admin.ModelAdmin):
    change_list_template = "admin/academy/student/change_list.html"

    # ---------------------------------------------------------
    # 1. DISPLAY VA FILTERLASH
    # ---------------------------------------------------------
    list_display = (
        'full_name',
        'phone_number',
        'get_created_at_display',
        'get_groups_display',
        'get_level_display',
        'get_teacher_display',
        'status',
        # --- TO'LOV USTUNLARI ---
        'get_monthly_payment',
        'get_months_display',
        'get_expected_total_display',
        'get_total_payment',
        'get_payment_debt',
        'get_paid_percent',
    )

    list_filter = (
        AcademyGroupFilter,
        AcademyLevelFilter,
        AcademyTeacherFilter,
        StatusFilter,
    )

    search_fields = (
        'full_name',
        'phone_number',
        'parent_phone',
        'enrollments__group__name',
    )

    inlines = [EnrollmentInline, StudentPaymentInline]

    fieldsets = (
        ("Asosiy ma'lumotlar", {
            'fields': ('full_name', 'phone_number', 'parent_phone', 'status')
        }),
    )

    list_per_page = 50
    show_full_result_count = False
    actions = None

    class Media:
        css = {
            'all': ('admin/css/custom_scroll.css',)
        }

    # ---------------------------------------------------------
    # 2. QUERYSET — ENROLLMENT ASOSIDA
    # ---------------------------------------------------------
    def get_queryset(self, request):
        queryset = super().get_queryset(request)

        # Prefetch active enrollments with related group data
        queryset = queryset.prefetch_related(
            Prefetch(
                'enrollments',
                queryset=Enrollment.objects.filter(is_active=True).select_related(
                    'group', 'group__level', 'group__teacher'
                ),
                to_attr='_active_enrollments'
            ),
            Prefetch(
                'enrollments',
                queryset=Enrollment.objects.select_related('group'),
                to_attr='_all_enrollments'
            ),
        )

        # Oylik to'lov — faol guruhlar narxi yig'indisi
        monthly_subquery = Enrollment.objects.filter(
            student=OuterRef('pk'),
            is_active=True
        ).values('student').annotate(
            total=Sum('group__price')
        ).values('total')

        # Jami to'langan
        payment_subquery = StudentPayment.objects.filter(
            student=OuterRef('pk')
        ).values('student').annotate(
            total=Sum('amount')
        ).values('total')

        queryset = queryset.annotate(
            monthly_payment=Coalesce(Subquery(monthly_subquery), Value(Decimal(0))),
            total_paid=Coalesce(Subquery(payment_subquery), Value(Decimal(0))),
        )

        return queryset

    # ---------------------------------------------------------
    # 3. YORDAMCHI HISOBLASH METODLARI
    # ---------------------------------------------------------
    def _get_expected_total(self, obj):
        """Barcha enrollmentlar bo'yicha kutilgan summani hisoblash"""
        enrollments = getattr(obj, '_all_enrollments', None)
        if enrollments is None:
            enrollments = obj.enrollments.select_related('group').all()
        total = Decimal(0)
        for enr in enrollments:
            total += enr.months_enrolled() * enr.group.price
        return total

    def _get_max_months(self, obj):
        """Eng uzoq o'qigan oyi"""
        enrollments = getattr(obj, '_active_enrollments', None)
        if enrollments is None:
            enrollments = obj.enrollments.filter(is_active=True)
        if not enrollments:
            return 0
        return max(enr.months_enrolled() for enr in enrollments)

    # ---------------------------------------------------------
    # 4. STATISTIKA PANELI
    # ---------------------------------------------------------
    def changelist_view(self, request, extra_context=None):
        groups_data = list(Group.objects.filter(is_active=True).values('id', 'name').order_by('name'))
        levels_data = list(Level.objects.values('id', 'name').order_by('name'))
        
        teachers = Teacher.objects.filter(is_active=True).only('id', 'first_name', 'last_name')
        teachers_data = [{'id': t.id, 'name': f"{t.first_name} {t.last_name or ''}".strip()} for t in teachers]
        teachers_data.sort(key=lambda x: x['name'])
        
        status_data = [
            {'id': 'active', 'name': 'Faol'},
            {'id': 'inactive', 'name': 'Nofaol'},
            {'id': 'all', 'name': 'Barcha o\'quvchilar'}
        ]

        extra_context = extra_context or {}
        extra_context['groups_json'] = json.dumps(groups_data)
        extra_context['levels_json'] = json.dumps(levels_data)
        extra_context['teachers_json'] = json.dumps(teachers_data)
        extra_context['status_json'] = json.dumps(status_data)

        response = super().changelist_view(request, extra_context)

        if not hasattr(response, 'context_data') or 'cl' not in response.context_data:
            return response

        cl = response.context_data['cl']
        qs = cl.queryset

        # Jami oylik to'lovlar va jami to'langan
        metrics = qs.aggregate(
            jami_oylik=Sum('monthly_payment'),
            jami_tolov=Sum('total_paid'),
        )

        val_monthly = metrics['jami_oylik'] or 0
        val_paid = metrics['jami_tolov'] or 0

        # Jami kutilgan va qarz — Python da hisoblash
        total_expected = Decimal(0)
        total_debt = Decimal(0)
        for obj in qs:
            expected = self._get_expected_total(obj)
            paid = obj.total_paid or Decimal(0)
            debt = max(expected - paid, Decimal(0))
            total_expected += expected
            total_debt += debt

        response.context_data['footer_stats'] = {
            'monthly': val_monthly,
            'paid': val_paid,
            'expected': total_expected,
            'debt': total_debt,
        }

        return response

    # ---------------------------------------------------------
    # 5. DISPLAY METODLARI
    # ---------------------------------------------------------

    @admin.display(description="Guruhlari")
    def get_groups_display(self, obj):
        enrollments = getattr(obj, '_active_enrollments', [])
        if enrollments:
            badges = ''.join([
                f'<span class="group-badge">{enr.group.name}</span> '
                for enr in enrollments
            ])
            return format_html(badges)
        return "-"

    @admin.display(description="Daraja")
    def get_level_display(self, obj):
        enrollments = getattr(obj, '_active_enrollments', [])
        if enrollments:
            levels = set(enr.group.level.name for enr in enrollments if enr.group.level)
            badges = ''.join([
                f'<span class="group-badge" style="background:#e7f5ff; color:#1c7ed6; border-color:#a5d8ff;">{lvl}</span> '
                for lvl in levels
            ])
            return format_html(badges)
        return "-"

    @admin.display(description="O'qituvchi")
    def get_teacher_display(self, obj):
        enrollments = getattr(obj, '_active_enrollments', [])
        if enrollments:
            teachers = set()
            for enr in enrollments:
                if enr.group.teacher:
                    teachers.add(f"{enr.group.teacher.first_name} {enr.group.teacher.last_name or ''}")
            if teachers:
                badges = ''.join([
                    f'<span class="group-badge" style="background:#fff4e6; color:#e8590c; border-color:#ffd8a8;">{t}</span> '
                    for t in teachers
                ])
                return format_html(badges)
        return "-"

    @admin.display(description="Oylik to'lov", ordering='monthly_payment')
    def get_monthly_payment(self, obj):
        val = getattr(obj, 'monthly_payment', Decimal(0))
        if val and val > 0:
            return format_html(
                '<span style="font-weight:600; color:#2b8a3e;">{}</span>',
                f"{val:,.0f}".replace(",", " ")
            )
        return "-"

    @admin.display(description="Jami summa")
    def get_expected_total_display(self, obj):
        total = self._get_expected_total(obj)
        if total > 0:
            return format_html(
                '<span style="font-weight:600; color:#343a40;">{}</span>',
                f"{total:,.0f}".replace(",", " ")
            )
        return "-"

    @admin.display(description="LTV", ordering='total_paid')
    def get_total_payment(self, obj):
        val = getattr(obj, 'total_paid', Decimal(0))
        if val and val > 0:
            return f"{val:,.0f}".replace(",", " ")
        return "-"

    @admin.display(description="Foizi")
    def get_paid_percent(self, obj):
        expected = self._get_expected_total(obj)
        paid = getattr(obj, 'total_paid', Decimal(0)) or Decimal(0)

        if expected <= 0:
            return format_html('<span class="status-badge badge-secondary">-</span>')

        percent = (paid / expected * 100).quantize(Decimal('1.'), rounding=ROUND_HALF_UP)

        if percent >= 100:
            css_class = 'badge-success'
            icon = '✓'
        elif percent >= 50:
            css_class = 'badge-info'
            icon = ''
        elif percent > 0:
            css_class = 'badge-warning'
            icon = ''
        else:
            css_class = 'badge-danger'
            icon = '!'

        return format_html(
            '<div class="status-badge {}">{} {}%</div>',
            css_class, icon, percent
        )

    @admin.display(description="Qarzi")
    def get_payment_debt(self, obj):
        expected = self._get_expected_total(obj)
        paid = getattr(obj, 'total_paid', Decimal(0)) or Decimal(0)

        if expected <= 0:
            return "-"

        debt = expected - paid
        if debt <= 0:
            return format_html('<span style="color: #20c997; font-weight: bold; font-size: 16px;">✓</span>')

        formatted = f"{debt:,.0f}".replace(",", " ")
        return format_html(
            '<span class="status-badge badge-danger">{}</span>',
            formatted
        )

    def dehydrate_qabul_order_date(self, student):
        order = student.enrollments.order_by('enrolled_date').first()
        if order and order.enrolled_date:
            return order.enrolled_date.strftime('%d.%m.%Y')
        return ""

    @admin.display(description="Qabul sanasi", ordering='created_at')
    def get_created_at_display(self, obj):
        if obj.created_at:
            return obj.created_at.strftime("%d.%m.%Y")
        return "-"

    @admin.display(description="Oy")
    def get_months_display(self, obj):
        months = self._get_max_months(obj)
        if months > 0:
            return format_html(
                '<span style="font-weight:600; color:#1c7ed6;">{} oy</span>',
                months
            )
        return "-"

    # ---------------------------------------------------------
    # 6. QO'SHIMCHA URL VA VIEWLAR
    # ---------------------------------------------------------
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('export-excel/', self.admin_site.admin_view(self.export_excel_view),
                 name='export_students_excel'),
        ]
        return custom_urls + urls

    @admin.display(description="Ko'rish")
    def view_student_link(self, obj):
        url = reverse('admin:academy_student_change', args=[obj.pk])
        return format_html(
            '<a href="{}" title="To\'liq ma\'lumot" style="color: #adb5bd; font-size: 1.2rem; transition: 0.2s;">'
            '<i style="color:blue" class="fas fa-eye"></i>'
            '</a>',
            url
        )

    # ---------------------------------------------------------
    # 7. EXCEL EXPORT
    # ---------------------------------------------------------
    def export_excel_view(self, request):
        """O'quvchilar ro'yxatini Excelga export qilish"""

        try:
            cl = self.get_changelist_instance(request)
            queryset = cl.get_queryset(request)
        except AttributeError:
            queryset = self.get_queryset(request)

        selected_fields = request.POST.getlist('selected_fields')
        if not selected_fields:
            selected_fields = [
                'full_name', 'phone_number', 'parent_phone',
                'groups', 'level', 'teacher',
                'monthly_payment', 'total_paid', 'debt', 'percent', 'months',
            ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "O'quvchilar Export"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        money_format = '#,##0'

        field_titles = {
            'full_name': 'F.I.SH.',
            'phone_number': 'Telefon',
            'parent_phone': 'Ota-ona telefoni',
            'groups': 'Guruhlari',
            'level': 'Daraja',
            'teacher': "O'qituvchi",
            'status': 'Status',
            'monthly_payment': 'Oylik to\'lov',
            'total_paid': "To'langan summa",
            'debt': 'Qarzdorlik',
            'percent': 'Foiz (%)',
            'months': 'Oy',
            'enrolled_date': 'Qabul sanasi',
        }

        money_fields = ['monthly_payment', 'total_paid', 'debt']

        for col_num, field in enumerate(selected_fields, 1):
            column_letter = get_column_letter(col_num)
            cell = ws.cell(row=1, column=col_num)
            cell.value = field_titles.get(field, field)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

            if field == 'full_name':
                ws.column_dimensions[column_letter].width = 35
            elif field in ('groups', 'teacher'):
                ws.column_dimensions[column_letter].width = 30
            else:
                ws.column_dimensions[column_letter].width = 18

        row_num = 2
        for obj in queryset:
            enrollments = getattr(obj, '_active_enrollments', None)
            if enrollments is None:
                enrollments = list(obj.enrollments.filter(is_active=True).select_related(
                    'group', 'group__level', 'group__teacher'
                ))

            all_enrollments = getattr(obj, '_all_enrollments', None)
            if all_enrollments is None:
                all_enrollments = list(obj.enrollments.select_related('group').all())

            expected = sum(enr.months_enrolled() * enr.group.price for enr in all_enrollments)
            paid = getattr(obj, 'total_paid', Decimal(0)) or Decimal(0)
            debt_val = max(expected - paid, Decimal(0))
            pct = int(paid / expected * 100) if expected > 0 else 0
            max_months = max((enr.months_enrolled() for enr in enrollments), default=0)

            for col_num, field in enumerate(selected_fields, 1):
                cell = ws.cell(row=row_num, column=col_num)
                val = None

                if field == 'groups':
                    val = ", ".join([enr.group.name for enr in enrollments])
                elif field == 'level':
                    levels = set(enr.group.level.name for enr in enrollments if enr.group.level)
                    val = ", ".join(levels)
                elif field == 'teacher':
                    teachers = set()
                    for enr in enrollments:
                        if enr.group.teacher:
                            teachers.add(f"{enr.group.teacher.first_name} {enr.group.teacher.last_name or ''}")
                    val = ", ".join(teachers)
                elif field == 'monthly_payment':
                    val = getattr(obj, 'monthly_payment', 0)
                elif field == 'total_paid':
                    val = paid
                elif field == 'debt':
                    val = debt_val
                elif field == 'percent':
                    val = pct
                elif field == 'months':
                    val = max_months
                elif field == 'enrolled_date':
                    if enrollments:
                        earliest = min(enr.enrolled_date for enr in enrollments)
                        val = earliest.strftime('%d.%m.%Y')
                    else:
                        val = "-"
                elif field == 'status':
                    val = obj.get_status_display()
                elif hasattr(obj, field):
                    val = getattr(obj, field)
                    if isinstance(val, (datetime, date)):
                        val = val.strftime('%d.%m.%Y')
                    elif callable(val):
                        val = val()

                if val is None:
                    val = ""

                if field in money_fields:
                    try:
                        cell.value = int(float(val)) if val else 0
                        cell.number_format = money_format
                    except (ValueError, TypeError):
                        cell.value = val
                elif field in ('percent', 'months'):
                    cell.value = val
                    cell.alignment = center_align
                else:
                    cell.value = str(val)

                cell.border = thin_border

                if field not in ['full_name', 'groups', 'teacher'] and field not in money_fields:
                    cell.alignment = center_align
                elif field in ['full_name']:
                    cell.alignment = left_align

            row_num += 1

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        response['Content-Disposition'] = f'attachment; filename=Academy_Students_{timestamp}.xlsx'
        wb.save(response)
        return response

    def get_changelist_instance(self, request):
        """ChangeList instance olish (filter ishlashi uchun)"""
        list_display = self.get_list_display(request)
        list_display_links = self.get_list_display_links(request, list_display)
        list_filter = self.get_list_filter(request)
        search_fields = self.get_search_fields(request)
        list_select_related = self.get_list_select_related(request)

        try:
            actions = self.get_actions(request)
            if actions:
                list_display = ['action_checkbox'] + list(list_display)
        except (AttributeError, KeyError):
            pass

        ChangeListClass = self.get_changelist(request)

        return ChangeListClass(
            request,
            self.model,
            list_display,
            list_display_links,
            list_filter,
            self.date_hierarchy,
            search_fields,
            list_select_related,
            self.list_per_page,
            self.list_max_show_all,
            self.list_editable,
            self,
            self.sortable_by,
            self.search_help_text,
        )


def academy_general_view(request):
    models_links = [
        {"title": "O'qituvchilar", "subtitle": "Teacher", "url": reverse('admin:academy_teacher_changelist'), "icon": "fas fa-chalkboard-teacher"},
        {"title": "Darajalar", "subtitle": "Level", "url": reverse('admin:academy_level_changelist'), "icon": "fas fa-layer-group"},
        {"title": "Hafta kunlari", "subtitle": "WeekDay", "url": reverse('admin:academy_weekday_changelist'), "icon": "fas fa-calendar-day"},
        {"title": "Kursga qabul", "subtitle": "Enrollment", "url": reverse('admin:academy_enrollment_changelist'), "icon": "fas fa-user-plus"},
        {"title": "O'quvchi to'lovlari", "subtitle": "StudentPayment", "url": reverse('admin:academy_studentpayment_changelist'), "icon": "fas fa-money-bill-wave"},
        {"title": "O'qituvchi maoshlari", "subtitle": "TeacherSalary", "url": reverse('admin:academy_teachersalary_changelist'), "icon": "fas fa-hand-holding-usd"},
    ]
    
    # Global admin kontekstini olish
    context = admin.site.each_context(request)
    context.update({
        'title': "Academy Sozlamalari",
        'models_links': models_links,
    })
    return render(request, 'admin/academy/general.html', context)

original_get_urls = admin.site.get_urls

def get_urls():
    custom_urls = [
        path('academy/general/', admin.site.admin_view(academy_general_view), name='academy_general'),
    ]
    return custom_urls + original_get_urls()

admin.site.get_urls = get_urls
