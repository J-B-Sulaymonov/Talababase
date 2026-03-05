from .base import *

@admin.register(Position)
class PositionAdmin(admin.ModelAdmin):
    list_display = ('name',)
    search_fields = ('name',)

class DepartmentMultiFilter(admin.SimpleListFilter):
    title = "Bo'lim / Kafedra"
    parameter_name = 'department__id'  # JS dagi nom bilan bir xil

    def lookups(self, request, model_admin):
        # Admin panelda ko'rsatish uchun (agar kerak bo'lsa)
        return [(d.id, d.name) for d in Department.objects.all()]

    def queryset(self, request, queryset):
        value = self.value()
        if not value:
            return queryset

        # "1,2,5" ko'rinishidagi stringni ro'yxatga aylantiramiz
        try:
            ids = value.split(',')
            # Bo'sh yoki noto'g'ri qiymatlarni olib tashlaymiz
            clean_ids = [x for x in ids if x.isdigit()]
            if clean_ids:
                return queryset.filter(department__id__in=clean_ids)
        except ValueError:
            pass
        return queryset

# 2. STATUS UCHUN MULTI-FILTER (Active, Pending va h.k. bir vaqtda tanlash uchun)
class StatusMultiFilter(admin.SimpleListFilter):
    title = "Holati"
    parameter_name = 'status'

    def lookups(self, request, model_admin):
        return [
            ('active', 'Faol'),
            ('pending', 'Kutilmoqda'),
            ('dismissed', "Bo'shatilgan"),
        ]

    def queryset(self, request, queryset):
        value = self.value()
        if not value:
            return queryset

        # "active,pending" -> ['active', 'pending']
        statuses = value.split(',')
        return queryset.filter(status__in=statuses)
class RoleListFilter(admin.SimpleListFilter):
    title = "Roli"
    parameter_name = 'role_filter'  # HTML/JS dagi nom bilan bir xil bo'lishi SHART

    def lookups(self, request, model_admin):
        # Admin panel o'ng tomonida chiqadigan filterlar (agar kerak bo'lsa)
        return [
            ('1', "O'qituvchi"),
            ('0', "Xodim"),
        ]

    def queryset(self, request, queryset):
        # URL dan kelgan qiymatni olamiz
        value = self.value()

        if not value:
            return queryset

        # JS dan '1,0' shaklida kelishi mumkin, shuning uchun split qilamiz
        roles = value.split(',')

        if '1' in roles and '0' in roles:
            return queryset
        elif '1' in roles:
            return queryset.filter(is_teacher=True)
        elif '0' in roles:
            return queryset.filter(Q(is_teacher=False) | Q(is_teacher__isnull=True))

        return queryset


class ScientificDegreeFilter(admin.SimpleListFilter):
    title = "Ilmiy daraja"
    parameter_name = 'scientific_degree'

    def lookups(self, request, model_admin):
        return Employee.DEGREE_CHOICES

    def queryset(self, request, queryset):
        value = self.value()
        if not value: return queryset
        degrees = value.split(',')
        return queryset.filter(scientific_degree__in=degrees)


class ScientificTitleFilter(admin.SimpleListFilter):
    title = "Ilmiy unvon"
    parameter_name = 'scientific_title'

    def lookups(self, request, model_admin):
        return Employee.TITLE_CHOICES

    def queryset(self, request, queryset):
        value = self.value()
        if not value: return queryset
        titles = value.split(',')
        return queryset.filter(scientific_title__in=titles)


# --- EMPLOYEE ADMIN (TO'LIQ) ---

@admin.register(Employee)
class EmployeeAdmin(admin.ModelAdmin):
    change_list_template = "admin/kadrlar/employee_change_list.html"

    # 1. RO'YXAT KO'RINISHI (List Display)
    list_display = ('get_full_name', 'department', 'get_positions_display',
                    'get_scientific_degree_display', 'get_scientific_title_display',
                    'status_badge', 'is_teacher_icon')

    # 2. FILTERLAR
    list_filter = (
        StatusMultiFilter,
        'gender',
        DepartmentMultiFilter,
        'positions',  # Lavozim bo'yicha filter (M2M)
        ScientificDegreeFilter,  # Yangilangan filter
        ScientificTitleFilter,  # Yangilangan filter
        'approved',
        RoleListFilter
    )

    search_fields = ('first_name', 'last_name', 'pid', 'passport_info')

    # ManyToMany maydonlar uchun qulay tanlash oynasi
    filter_horizontal = ('positions',)

    # 3. FORMA KO'RINISHI (Fieldsets)
    fieldsets = (
        ('Shaxsiy maʼlumot', {
            'fields': ('first_name', 'last_name', 'middle_name', 'gender', 'passport_info', 'pid', 'birth_date',
                       'photo')
        }),
        ('Ish joyi va Ilmiy Salohiyat', {
            'fields': ('department', 'positions', 'scientific_degree', 'scientific_title', 'is_teacher', 'order')
        }),
        ('HR Tasdiq', {
            'fields': ('hired_at', 'status', 'approved', 'archived')
        }),
    )

    # --- URLS VA CUSTOM VIEWS ---
    def get_urls(self):
        urls = super().get_urls()

        # MUHIM: Model nomini aniqlaymiz ('employee' yoki 'archivedemployee')
        # Bu URL nomlari to'qnashuvini oldini oladi
        model_name = self.model._meta.model_name

        custom_urls = [
            # 'name' endi dinamik bo'ldi: kadrlar_employee_export yoki kadrlar_archivedemployee_export
            path('export/', self.admin_site.admin_view(self.export_employees_view),
                 name=f'kadrlar_{model_name}_export'),
            path('<path:object_id>/card/', self.admin_site.admin_view(self.card_view), name='employee_card_print'),
        ]
        return custom_urls + urls

    def export_employees_view(self, request):
        """
        TUZATILGAN VERSION: ChangeList argumentlari to'g'irlandi.
        """
        print("\n" + "=" * 50)
        print(f">>> EXPORT SO'ROVI: {request.path}")

        if request.method != 'POST':
            return HttpResponse("Faqat POST so'rov qabul qilinadi", status=405)

        # 1. Tanlangan ustunlar
        selected_fields = request.POST.getlist('selected_fields')
        if not selected_fields:
            selected_fields = [
                'first_name', 'last_name', 'passport_info',
                'department', 'positions', 'status',
            ]

        # 2. QUERYSETNI FILTRLASH (ChangeList)
        # Javascript orqali kelgan filtrlarni tekshiramiz
        print(f">>> URL Parametrlari (GET): {request.GET}")

        from django.contrib.admin.views.main import ChangeList

        queryset = None

        try:
            # ChangeList uchun barcha kerakli argumentlarni yig'amiz
            list_display = self.get_list_display(request)
            list_display_links = self.get_list_display_links(request, list_display)
            list_filter = self.get_list_filter(request)
            search_fields = self.get_search_fields(request)
            list_select_related = self.get_list_select_related(request)

            # --- TUZATISH SHU YERDA ---
            # sortable_by va search_help_text argumentlarini xavfsiz olish
            sortable_by = getattr(self, 'sortable_by', ())
            search_help_text = getattr(self, 'search_help_text', None)

            cl = ChangeList(
                request,
                self.model,
                list_display,
                list_display_links,
                list_filter,
                self.date_hierarchy,
                search_fields,
                list_select_related,
                self.list_per_page,
                self.list_max_show_all,
                self.list_editable,
                self,
                sortable_by,  # <--- YANGI QO'SHILDI
                search_help_text  # <--- YANGI QO'SHILDI
            )

            # Filtrlangan ma'lumotlar
            queryset = cl.get_queryset(request)
            print(f">>> ChangeList muvaffaqiyatli ishladi. Natija: {queryset.count()} ta")

        except Exception as e:
            print(f">>> XATOLIK (ChangeList): {e}")
            # Xatolik bo'lsa, asosiy querysetni olamiz
            queryset = self.get_queryset(request)
            print(f">>> Fallback ishlatildi. Natija: {queryset.count()} ta")

        # 3. EXCEL YARATISH (OpenPyXL)
        import openpyxl
        import datetime
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Xodimlar"

        # Dizayn
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        # Field nomlari (Mapping)
        field_map = {
            'first_name': 'Ism', 'last_name': 'Familiya', 'middle_name': 'Otasining ismi',
            'passport_info': 'Pasport', 'pid': 'JSHSHIR', 'birth_date': 'Tug‘ilgan sana',
            'gender': 'Jinsi', 'department': "Bo‘lim / Kafedra", 'positions': 'Lavozimi',
            'status': 'Holati', 'hired_at': 'Ishga kirgan sana',
            'scientific_degree': 'Ilmiy daraja', 'scientific_title': 'Ilmiy unvon',
            'is_teacher': 'Roli', 'approved': 'HR Tasdiq'
        }

        headers = ["№"] + [field_map.get(f, f) for f in selected_fields]

        # Sarlavha yozish
        for col, title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=title)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

        ws.column_dimensions['A'].width = 5

        # Ma'lumotlarni to'ldirish
        if queryset and queryset.exists():
            for r, obj in enumerate(queryset, start=2):
                ws.cell(row=r, column=1, value=r - 1).border = border
                ws.cell(row=r, column=1).alignment = center

                for c, field in enumerate(selected_fields, start=2):
                    val = "-"
                    try:
                        if field == 'positions':
                            val = ", ".join([p.name for p in obj.positions.all()])
                        elif field == 'department':
                            val = obj.department.name if obj.department else "-"
                        elif field == 'status':
                            val = obj.get_status_display()
                        elif field == 'gender':
                            val = obj.get_gender_display()
                        elif field == 'scientific_degree':
                            val = obj.get_scientific_degree_display()
                        elif field == 'scientific_title':
                            val = obj.get_scientific_title_display()
                        elif field == 'is_teacher':
                            val = "O'qituvchi" if obj.is_teacher else "Xodim"
                        elif field == 'approved':
                            val = "Ha" if obj.approved else "Yo'q"
                        else:
                            raw = getattr(obj, field, "-")
                            val = str(raw) if raw is not None else ""
                    except Exception:
                        val = ""

                    cell = ws.cell(row=r, column=c, value=val)
                    cell.border = border
                    cell.alignment = left
        else:
            # Bo'sh bo'lsa
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
            cell = ws.cell(row=2, column=1, value="Ma'lumot topilmadi")
            cell.alignment = center
            cell.font = Font(italic=True, color="FF0000")

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"Export_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        print("=" * 50 + "\n")
        return response

    def get_field_verbose_name(self, field_name):
        try:
            return self.model._meta.get_field(field_name).verbose_name
        except:
            return field_name

    # --- METODLAR ---

    # 1. Lavozimlar (M2M)
    def get_positions_display(self, obj):
        return ", ".join([p.name for p in obj.positions.all()])

    get_positions_display.short_description = "Lavozimlar"

    # 2. Ilmiy Daraja (Wrapper)
    @admin.display(description="Ilmiy Daraja", ordering='scientific_degree')
    def get_scientific_degree_display(self, obj):
        return obj.get_scientific_degree_display()

    # 3. Ilmiy Unvon (Wrapper)
    @admin.display(description="Ilmiy Unvon", ordering='scientific_title')
    def get_scientific_title_display(self, obj):
        return obj.get_scientific_title_display()

    # 4. Boshqa ustunlar
    def get_full_name(self, obj):
        return f"{obj.last_name} {obj.first_name}"

    get_full_name.short_description = "F.I.SH"

    def passport_info(self, obj):
        return obj.passport_info if obj.passport_info else "-"

    passport_info.short_description = "Passport"

    def status_badge(self, obj):
        colors = {'active': '#10b981', 'pending': '#f59e0b', 'dismissed': '#ef4444'}
        color = colors.get(obj.status, '#64748b')
        return format_html(
            '<span style="color: white; background-color: {}; padding: 3px 10px; border-radius: 10px; font-size: 11px; font-weight: bold;">{}</span>',
            color, obj.get_status_display())

    status_badge.short_description = "HR Holati"

    def is_teacher_icon(self, obj):
        if obj.is_teacher:
            # Teacher profiliga o'tish uchun link
            url = reverse('admin:kadrlar_teacher_changelist') + f"?employee__id={obj.id}"
            return format_html(
                '<a href="{}" style="color: #3b82f6; font-weight:bold; text-decoration:none;"><i class="fas fa-chalkboard-teacher"></i> O\'qituvchi</a>',
                url)
        return format_html(
            '<span style="color: #64748b; font-weight:500;"><i class="fas fa-user-tie"></i> Xodim</span>')

    is_teacher_icon.short_description = "Roli"
    is_teacher_icon.allow_tags = True
    is_teacher_icon.admin_order_field = 'is_teacher'

    def view_card_icon(self, obj):
        url = reverse('admin:employee_card_print', args=[obj.pk])
        return format_html(
            '''<a href="{}" title="Profilni ko'rish"
                  style="background-color: #3b82f6; color: white; padding: 5px 10px; border-radius: 6px; text-decoration: none; font-weight: bold; display: inline-flex; align-items: center; gap: 5px;">
                <i class="fas fa-id-card"></i> Karta
               </a>''', url
        )

    view_card_icon.short_description = "Karta"
    view_card_icon.allow_tags = True

    # --- VIEW MOSLASHUVCHANLIGI (RUXSATLAR) ---
    def get_list_display(self, request):
        columns = list(super().get_list_display(request))
        if is_hr_admin(request.user):
            columns.append('view_card_icon')
        return columns

    def get_inlines(self, request, obj):
        if is_hr_admin(request.user):
            return [DocumentInline, OrderInline]
        return [DocumentInline]

    def get_readonly_fields(self, request, obj=None):
        if is_hr_admin(request.user):
            return ('created_by',)
        readonly = ['hired_at', 'status', 'approved', 'archived', 'created_by', 'order']
        if obj:
            readonly.append('department')
            if obj.approved:
                all_fields = [f.name for f in self.model._meta.fields]
                return all_fields
        return readonly

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if not is_hr_admin(request.user):
            dept = Department.objects.filter(head_manager=request.user).first()
            if 'department' in form.base_fields:
                form.base_fields['department'].disabled = True
                form.base_fields['department'].required = False
                if dept:
                    form.base_fields['department'].initial = dept.id
                    form.base_fields[
                        'department'].help_text = f"Siz faqat o'zingizning bo'limingiz ({dept.name}) ga xodim qo'sha olasiz."
        return form

    # --- KARTA KO'RISH VIEW ---
    def card_view(self, request, object_id):
        employee = get_object_or_404(Employee, pk=object_id)
        teacher_profile = getattr(employee, 'teacher_profile', None)

        raw_results = QuizResult.objects.filter(employee=employee).order_by('-created_at')
        enhanced_results = []
        for res in raw_results:
            data = res.struct if isinstance(res.struct, dict) else {}
            overall_text = data.get('overall_conclusion')
            if not overall_text or overall_text == "Natija izohi mavjud emas.":
                infos = QuizScoringInfo.objects.filter(quiz=res.quiz)
                for info in infos:
                    if info.min_score <= res.total_score <= info.max_score:
                        overall_text = info.conclusion
                        break
                if not overall_text:
                    overall_text = "Natija izohi mavjud emas."
            scoring_scale = QuizScoringInfo.objects.filter(quiz=res.quiz).order_by('min_score')
            enhanced_results.append({
                'obj': res,
                'struct': data,
                'calculated_conclusion': overall_text,
                'scale': scoring_scale
            })

        context = {
            **self.admin_site.each_context(request),
            'opts': self.model._meta,
            'employee': employee,
            'teacher': teacher_profile,
            'quiz_results': enhanced_results,
            'title': f"Profil: {employee.last_name} {employee.first_name}",
        }
        return TemplateResponse(request, 'admin/kadrlar/employee_card.html', context)

    # --- QUERYSET ---
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if is_hr_admin(request.user) or is_edu_admin(request.user):
            return qs.filter(archived=False)
        return qs.filter(department__head_manager=request.user, archived=False)

    # --- STATISTIKA (ChangeList View) ---
    def changelist_view(self, request, extra_context=None):
        # =========================================================
        # 1. DEFAULT FILTER (Agar filtr tanlanmagan bo'lsa -> Faol)
        # =========================================================
        if request.method == 'GET' and not request.GET:
            # Faqat 'Employee' modeli uchun ishlaydi (ArchivedEmployee uchun emas)
            if self.model.__name__ == 'Employee':
                # URLga ?status=active qo'shib yuboramiz
                return redirect(f"{request.path}?status=active")

        # =========================================================
        # 2. ASOSIY KOD (Eski kod o'z holicha qoladi)
        # =========================================================
        extra_context = extra_context or {}
        extra_context['is_kadr_member'] = is_hr_admin(request.user)

        response = super().changelist_view(request, extra_context=extra_context)

        if hasattr(response, 'context_data'):
            try:
                base_qs = self.get_queryset(request)
                # Faqat faol va arxivlanmagan xodimlarni olamiz
                active_qs = base_qs.filter(status='active', archived=False)

                # --- STATISTIKA KODI (O'zgarmaydi) ---
                def get_gender_stats(queryset):
                    total = queryset.count()
                    male = queryset.filter(gender='male').count()
                    female = queryset.filter(gender='female').count()
                    male_pct = round((male / total) * 100) if total > 0 else 0
                    female_pct = round((female / total) * 100) if total > 0 else 0
                    return {
                        'count': total,
                        'male': male,
                        'female': female,
                        'male_pct': male_pct,
                        'female_pct': female_pct
                    }

                stats = {
                    'all': get_gender_stats(active_qs),
                    'teachers': get_gender_stats(active_qs.filter(is_teacher=True)),
                    'staff': get_gender_stats(active_qs.filter(Q(is_teacher=False) | Q(is_teacher__isnull=True))),
                }

                # --- ILMIY SALOHIYAT ---
                science_qs = active_qs

                scientific_stats = {
                    'degree': {
                        'phd': science_qs.filter(scientific_degree='phd').count(),
                        'dsc': science_qs.filter(scientific_degree='dsc').count(),
                    },
                    'title': {
                        'docent': science_qs.filter(scientific_title='docent').count(),
                        'professor': science_qs.filter(scientific_title='professor').count(),
                        'academic': science_qs.filter(scientific_title='academic').count(),
                    }
                }

                departments = Department.objects.all().values('id', 'name')
                response.context_data['departments_json'] = json.dumps(list(departments))
                response.context_data['stats'] = stats
                response.context_data['sc_stats'] = scientific_stats

            except (KeyError, AttributeError):
                pass
        return response

    # --- SAVE METODLARI ---
    def save_model(self, request, obj, form, change):
        if not is_hr_admin(request.user):
            dept = Department.objects.filter(head_manager=request.user).first()
            if dept:
                obj.department = dept
            if not change:
                obj.status = 'pending'
                obj.approved = False
                obj.created_by = request.user
        super().save_model(request, obj, form, change)

    def save_formset(self, request, form, formset, change):
        instances = formset.save(commit=False)
        for instance in instances:
            if isinstance(instance, Order):
                if not is_hr_admin(request.user): continue
                instance.created_by = request.user
                emp = instance.employee
                if instance.order_type == 'hire':
                    emp.status = 'active'
                    emp.approved = True
                    emp.archived = False
                    emp.hired_at = instance.date if not emp.hired_at else emp.hired_at
                elif instance.order_type == 'dismiss':
                    emp.status = 'dismissed'
                    emp.archived = True
                emp.save()
            instance.save()
        formset.save_m2m()  # M2M (Positions) saqlanishi uchun muhim!

    # --- PERMISSIONS ---
    def has_add_permission(self, request):
        if is_hr_admin(request.user): return True
        if Department.objects.filter(head_manager=request.user).exists(): return True
        return False

    def has_change_permission(self, request, obj=None):
        if is_hr_admin(request.user): return True
        if obj and obj.department and obj.department.head_manager == request.user: return True
        return False

    def has_delete_permission(self, request, obj=None):
        return is_hr_admin(request.user)


@admin.register(ArchivedEmployee)
class ArchivedEmployeeAdmin(EmployeeAdmin):
    """
    Arxivdagi (bo'shatilgan) xodimlar uchun maxsus Admin klass.
    """
    # 1. Maxsus shablonni ulaymiz (faqat kerakli statistika chiqishi uchun)
    change_list_template = "admin/kadrlar/archive_change_list.html"

    # 2. Ro'yxatda ko'rinadigan ustunlar
    list_display = ('get_full_name', 'pid', 'department', 'hired_at', 'dismissed_date_col', 'restore_button')
    list_filter = ('department', 'gender')
    search_fields = ('first_name', 'last_name', 'pid')
    list_editable = ()
    # 3. Arxivdagi ma'lumotlarni tahrirlashni cheklash (Read-only)
    readonly_fields = [f.name for f in Employee._meta.fields]

    def get_queryset(self, request):
        """Faqat arxivlangan (archived=True) xodimlarni qaytaradi"""
        return self.model.objects.filter(archived=True)

    # --- RUXSATLAR (PERMISSIONS) ---
    def has_module_permission(self, request):
        """Faqat HR va Superuser ko'ra oladi"""
        return is_hr_admin(request.user)

    def has_view_permission(self, request, obj=None):
        return is_hr_admin(request.user)

    def has_change_permission(self, request, obj=None):
        return is_hr_admin(request.user)

    def has_add_permission(self, request):
        """Arxivga qo'lda odam qo'shib bo'lmaydi, faqat buyruq orqali tushadi"""
        return False

    def has_delete_permission(self, request, obj=None):
        return is_hr_admin(request.user)

    # --- QO'SHIMCHA USTUNLAR ---
    def dismissed_date_col(self, obj):
        """Oxirgi 'dismiss' buyrug'i sanasini chiqarish"""
        dismiss_order = obj.orders.filter(order_type='dismiss').order_by('-date').first()
        return dismiss_order.date if dismiss_order else "-"

    dismissed_date_col.short_description = "Bo'shatilgan sana"

    def restore_button(self, obj):
        return format_html('<span style="color: #f59e0b; font-weight:bold;">Arxivlangan</span>')

    restore_button.short_description = "Holati"

    # --- HARAKATLAR (ACTIONS) ---
    actions = ['restore_employees']

    def restore_employees(self, request, queryset):
        """Tanlanganlarni arxivdan chiqarib, Active holatiga qaytarish"""
        updated_count = queryset.update(archived=False, status='active')
        self.message_user(request, f"{updated_count} nafar xodim arxivdan chiqarildi va 'Faol' holatiga o'tkazildi.")

    restore_employees.short_description = "♻️ Tanlanganlarni arxivdan qaytarish (Active)"

    # --- STATISTIKA (ENG MUHIM QISMI) ---
    def changelist_view(self, request, extra_context=None):
        """
        Bu yerda biz 'Faol' va 'Kutilmoqda' statistikasini olib tashlaymiz.
        Faqat Jami va Jins bo'yicha hisob-kitob qilamiz.
        """
        response = super().changelist_view(request, extra_context=extra_context)

        # Agar sahifa muvaffaqiyatli yuklansa va kontekst bo'lsa
        if hasattr(response, 'context_data'):
            # Hozirgi filterdagi querysetni olamiz
            try:
                cl = response.context_data['cl']
                qs = cl.queryset

                total = qs.count()
                male = qs.filter(gender='male').count()
                female = qs.filter(gender='female').count()

                # STATISTIKA LUG'ATINI QAYTA YOZAMIZ
                # 'active' va 'vacation' kalitlari yo'q!
                stats = {
                    'total': total,
                    'male': male,
                    'male_percent': round((male / total) * 100) if total else 0,
                    'female': female,
                    'female_percent': round((female / total) * 100) if total else 0,
                }
                response.context_data['stats'] = stats
            except (KeyError, AttributeError):
                pass

        return response
