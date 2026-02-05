from django.utils import timezone
from datetime import timedelta
from mptt.admin import DraggableMPTTAdmin
from .models import SimpleStructure
from django.http import HttpResponse, JsonResponse
from django.utils.safestring import mark_safe
import json
from django.db import models
from django.utils.html import format_html
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import path, reverse
from django.utils.http import urlencode
from django.db.models import Count, Q
from django.contrib.auth import get_user_model
from django import forms
from django.template.response import TemplateResponse
from django.contrib import admin


try:
    from nested_admin import NestedModelAdmin, NestedStackedInline, NestedTabularInline
except ImportError:
    # Fallback agar kutubxona bo'lmasa
    NestedModelAdmin = admin.ModelAdmin
    NestedStackedInline = admin.StackedInline
    NestedTabularInline = admin.TabularInline

from .models import (
    Department, Employee, Document, Order,
    Teacher, TeacherAvailability, Weekday, TimeSlot, Quiz, QuizQuestion, QuizAnswer, QuizResultKey, QuizPermission,
    QuizResult, QuizScoringRule, QuizScoringInfo, ArchivedEmployee, Position, OrganizationStructure
)

User = get_user_model()


# =========================================================
# 1. RUXSATLARNI ANIQLASH (HELPER FUNCTIONS)
# =========================================================

def is_hr_admin(user):
    """Foydalanuvchi Kadrlar bo'limi yoki Superuser ekanligini tekshiradi."""
    return user.is_superuser or user.groups.filter(name='Kadrlar').exists()


def is_edu_admin(user):
    """Foydalanuvchi O'quv bo'limi yoki Superuser ekanligini tekshiradi."""
    return user.is_superuser or user.groups.filter(name='OquvBolimi').exists()


# =========================================================
# 2. DASHBOARD VIEW (STATISTIKA)
# =========================================================
def kadrlar_structure_view(request):
    """
    YANGI: Faqat Struktura modellarini ko'rsatuvchi ichki dashboard.
    """
    user = request.user
    # Ruxsatni tekshirish (Faqat HR yoki Superuser)
    if not is_hr_admin(user):
        return render(request, 'admin/login.html') # Yoki 403 error

    models_links = []

    models_links.append({
        "title": "Xodimlar Strukturasi",
        "subtitle": "Tugunlar, bo'limlar va xodimlarni bog'lash",
        "url": reverse('admin:kadrlar_simplestructure_changelist'),
        "icon": "fas fa-sitemap",
    })
    models_links.append({
        "title": "Qo'lda structura yasash",
        "subtitle": "Interaktiv tuzilma yaratish va tahrirlash",
        "url": reverse('admin:kadrlar_organizationstructure_changelist'),
        "icon": "fas fa-project-diagram",
    })

    context = {
        'title': "Tashkiliy Tuzilma Sozlamalari",
        'models_links': models_links,
        'stats': {}, # Bu yerda statistika shart emas
        **admin.site.each_context(request),
    }
    # Xuddi general.html shablonidan foydalanaveramiz, chunki tuzilishi bir xil
    return render(request, 'admin/kadrlar/general.html', context)

def kadrlar_general_view(request):
    user = request.user

    def get_qs(model):
        if is_hr_admin(user) or is_edu_admin(user):
            return model.objects.all()
        if model == Teacher:
            return model.objects.filter(employee__department__head_manager=user)
        return model.objects.none()

    teacher_qs = get_qs(Teacher).filter(employee__archived=False)

    # Statistika yig'ish
    stats = {
        'teacher_approved': teacher_qs.filter(schedule_approved=True).count(),
        'teacher_total': teacher_qs.count(),
        'order_total': Order.objects.count() if is_hr_admin(user) else 0,

        # --- QUIZ STATISTIKASI ---
        'quiz_active': Quiz.objects.filter(is_active=True).count(),
        'quiz_total': Quiz.objects.count(),
        'permissions_active': QuizPermission.objects.filter(is_active=True).count(),
        'results_total': QuizResult.objects.count(),
    }

    models_links = []

    # 1. O'qituvchilar
    models_links.append({
        "title": "O'qituvchilar",
        "subtitle": f"Tasdiqlangan: {stats['teacher_approved']} / {stats['teacher_total']}",
        "url": reverse('admin:kadrlar_teacher_changelist'),
        "icon": "fas fa-chalkboard-teacher",
    })
    # 2. HR Admin uchun bo'limlar
    if is_hr_admin(user):
        # --- YANGI: BO'LIMLAR VA KAFEDRALAR ---
        dept_count = Department.objects.count()
        models_links.append({
            "title": "Bo'limlar va Kafedralar",
            "subtitle": f"Jami: {dept_count} ta",
            "url": reverse('admin:kadrlar_department_changelist'),
            "icon": "fas fa-sitemap",  # Estetik ikonka
        })
        models_links.append({
            "title": "Struktura",
            "subtitle": "Diagramma va Tuzilma sozlamalari",
            "url": reverse('admin:kadrlar_structure_subview'),  # Pastda URL nomini belgilaymiz
            "icon": "fas fa-network-wired",
        })
        # ---------------------------------------
        # --- YANGI: LAVOZIMLAR ---
        pos_count = Position.objects.count()
        models_links.append({
            "title": "Lavozimlar (Shtat)",
            "subtitle": f"Jami: {pos_count} ta",
            "url": reverse('admin:kadrlar_position_changelist'),
            "icon": "fas fa-user-tag",  # Ikonka: User Tag
        })
        models_links.append({
            "title": "Buyruqlar",
            "subtitle": f"Jami: {stats['order_total']}",
            "url": reverse('admin:kadrlar_order_changelist'),
            "icon": "fas fa-file-signature",
        })

        doc_count = Document.objects.count()
        models_links.append({
            "title": "Hujjatlar",
            "subtitle": f"Jami: {doc_count} ta fayl",
            "url": reverse('admin:kadrlar_document_changelist'),
            "icon": "fas fa-folder-open",
        })

        archived_count = Employee.objects.filter(archived=True).count()
        models_links.append({
            "title": "Arxiv (Bo'shaganlar)",
            "subtitle": f"Jami: {archived_count} nafar",
            "url": reverse('admin:kadrlar_archivedemployee_changelist'),
            "icon": "fas fa-archive",
        })

    # --- 3. QUIZ BO'LIMLARI ---
    if is_hr_admin(user):
        # A) Quizlar (Testlar)
        models_links.append({
            "title": "Testlar (Quiz)",
            "subtitle": f"Faol: {stats['quiz_active']} / Jami: {stats['quiz_total']}",
            "url": reverse('admin:kadrlar_quiz_changelist'),
            "icon": "fas fa-clipboard-list",
        })

        # B) Testga Ruxsatlar
        models_links.append({
            "title": "Testga ruxsatlar",
            "subtitle": f"Ochiq ruxsatlar: {stats['permissions_active']}",
            "url": reverse('admin:kadrlar_quizpermission_changelist'),
            "icon": "fas fa-unlock-alt",
        })

        # C) Test Natijalari
        models_links.append({
            "title": "Test Natijalari",
            "subtitle": f"Topshirilgan: {stats['results_total']}",
            "url": reverse('admin:kadrlar_quizresult_changelist'),
            "icon": "fas fa-chart-pie",
        })


    context = {
        'title': "Kadrlar Bo'limi Boshqaruv Paneli",
        'models_links': models_links,
        'stats': stats,
        **admin.site.each_context(request),
    }
    return render(request, 'admin/kadrlar/general.html', context)


original_get_urls = admin.site.get_urls


def get_urls():
    custom_urls = [path('kadrlar/general/', admin.site.admin_view(kadrlar_general_view), name='kadrlar_general'),
                   path('kadrlar/structure-settings/', admin.site.admin_view(kadrlar_structure_view),
                        name='kadrlar_structure_subview'),
                   ]
    return custom_urls + original_get_urls()


admin.site.get_urls = get_urls


# =========================================================
# 3. YORDAMCHI INLINELAR
# =========================================================

# O'ZGARISH: NestedStackedInline -> admin.StackedInline
class DocumentInline(admin.StackedInline):
    model = Document
    extra = 0
    fields = ('doc_type', 'file', 'number')

    def has_change_permission(self, request, obj=None):
        if obj and obj.approved and not is_hr_admin(request.user):
            return False
        return True

    def has_add_permission(self, request, obj=None):
        if obj and obj.approved and not is_hr_admin(request.user):
            return False
        return True

    def has_delete_permission(self, request, obj=None):
        if obj and obj.approved and not is_hr_admin(request.user):
            return False
        return True


# O'ZGARISH: NestedStackedInline -> admin.StackedInline
class OrderInline(admin.StackedInline):
    model = Order
    extra = 0
    fields = ('number', 'order_type', 'date', 'document')
    verbose_name = "Buyruq"
    verbose_name_plural = "Buyruqlar"

    def has_add_permission(self, request, obj):
        return is_hr_admin(request.user)

    def has_change_permission(self, request, obj=None):
        return is_hr_admin(request.user)

    def has_delete_permission(self, request, obj=None):
        return is_hr_admin(request.user)


class TeacherAvailabilityInline(admin.StackedInline):
    model = TeacherAvailability
    extra = 0
    min_num = 1
    max_num = 7
    verbose_name = "Bo'sh vaqt"
    verbose_name_plural = "O'qituvchining bo'sh vaqtlari"

    formfield_overrides = {
        models.ManyToManyField: {'widget': forms.CheckboxSelectMultiple},
    }

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "weekday":
            kwargs["empty_label"] = "Kunni tanlang"
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    # --- RUXSATLAR ---
    def get_readonly_fields(self, request, obj=None):
        # HR uchun hamma narsa readonly
        if is_hr_admin(request.user):
            return ['weekday', 'timeslots']
        return []

    def has_add_permission(self, request, obj=None):
        if is_hr_admin(request.user):
            return False
        if obj and obj.schedule_approved and not is_edu_admin(request.user):
            return False
        return True

    def has_change_permission(self, request, obj=None):
        if is_hr_admin(request.user):
            return True  # Ko'rish uchun ruxsat, lekin readonly bo'ladi
        if obj and obj.schedule_approved and not is_edu_admin(request.user):
            return False
        return True

    def has_delete_permission(self, request, obj=None):
        if is_hr_admin(request.user):
            return False
        if obj and obj.schedule_approved and not is_edu_admin(request.user):
            return False
        return True


# =========================================================
# 4. KAFEDRA / BO'LIM ADMINI
# =========================================================

@admin.register(Department)
class DepartmentAdmin(admin.ModelAdmin):
    change_list_template = "admin/kadrlar/department_change_list.html"
    list_display = ('colored_name', 'head_manager_col', 'styled_employee_count','order')
    list_display_links = ('colored_name',)
    search_fields = ('name',)
    list_editable = ('order',)
    ordering = ('order',)

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if is_hr_admin(request.user) or is_edu_admin(request.user):
            return qs
        return qs.filter(head_manager=request.user)

    def has_change_permission(self, request, obj=None):
        if is_hr_admin(request.user):
            return True
        if obj and obj.head_manager == request.user:
            return True
        return False

    def get_readonly_fields(self, request, obj=None):
        if not is_hr_admin(request.user):
            return ('name', 'slug', 'head_manager', 'created_by')
        return ()

    def colored_name(self, obj):
        return format_html(
            '''<div style="display: flex; align-items: center;">
                <span style="display: inline-flex; align-items: center; justify-content: center; 
                    width: 35px; height: 35px; background: #eef2f7; border-radius: 8px; 
                    margin-right: 12px; color: #3b82f6;">
                    <i class="fas fa-building"></i>
                </span>
                <span style="font-size: 15px; font-weight: 600; color: #334155;">{}</span>
            </div>''', obj.name
        )

    colored_name.short_description = "Kafedra / Bo'lim"

    def head_manager_col(self, obj):
        if obj.head_manager:
            return format_html('<i class="fas fa-user-tie" style="color:#64748b"></i> {}',
                               obj.head_manager.get_full_name() or obj.head_manager.username)
        return format_html('<span style="color:#ef4444">Tayinlanmagan</span>')

    head_manager_col.short_description = "Rahbar"

    def styled_employee_count(self, obj):
        count = obj.employees.filter(archived=False).count()
        if count == 0:
            url = reverse("admin:kadrlar_employee_changelist")
            query = urlencode({'department__id': obj.id})
            style = "background-color: #fff1f2; color: #e11d48; padding: 6px 14px; border-radius: 20px; font-size: 12px; font-weight: bold; box-shadow: 0 4px 6px -1px rgba(16, 185, 129, 0.4); text-decoration: none;"
            return format_html('<a href="{}?{}" style="{}">{} <i class="fas fa-arrow-right"></i></a>', url, query,
                               style,
                               f"Xodimlar yo'q")
        url = reverse("admin:kadrlar_employee_changelist")
        query = urlencode({'department__id': obj.id})
        style = "background: linear-gradient(135deg, #10b981 0%, #059669 100%); color: white; padding: 6px 14px; border-radius: 20px; font-size: 12px; font-weight: bold; box-shadow: 0 4px 6px -1px rgba(16, 185, 129, 0.4); text-decoration: none;"
        return format_html('<a href="{}?{}" style="{}">{} <i class="fas fa-arrow-right"></i></a>', url, query, style,
                           f"{count} nafar")

    styled_employee_count.short_description = "Xodimlar Soni"


    def changelist_view(self, request, extra_context=None):
        response = super().changelist_view(request, extra_context=extra_context)
        if hasattr(response, 'context_data'):
            try:
                cl = response.context_data['cl']
                qs = cl.queryset

                # Arxivlanmagan barcha xodimlarni olamiz (ushbu filtrlangan kafedralar bo'yicha)
                all_employees = Employee.objects.filter(department__in=qs, archived=False)

                stats = {
                    'total_depts': qs.count(),

                    # Jami xodimlar (arxivsiz)
                    'total_employees': all_employees.count(),

                    # Faol xodimlar (Active statusda va arxivsiz)
                    'active_employees': all_employees.filter(status='active').count(),

                    # Kutilayotganlar
                    'pending_employees': all_employees.filter(status='pending').count(),
                }
                response.context_data['stats'] = stats
            except (KeyError, AttributeError):
                pass
        return response


# =========================================================
# 5. XODIMLAR ADMIN (HR) + ID CARD VIEW
# =========================================================

@admin.register(Position)
class PositionAdmin(admin.ModelAdmin):
    list_display = ('name',)
    search_fields = ('name',)

class DepartmentMultiFilter(admin.SimpleListFilter):
    title = "Bo'lim / Kafedra"
    parameter_name = 'department__id'  # JS dagi nom bilan bir xil

    def lookups(self, request, model_admin):
        # Admin panelda ko'rsatish uchun (agar kerak bo'lsa)
        return [(d.id, d.name) for d in Department.objects.all()]

    def queryset(self, request, queryset):
        value = self.value()
        if not value:
            return queryset

        # "1,2,5" ko'rinishidagi stringni ro'yxatga aylantiramiz
        try:
            ids = value.split(',')
            # Bo'sh yoki noto'g'ri qiymatlarni olib tashlaymiz
            clean_ids = [x for x in ids if x.isdigit()]
            if clean_ids:
                return queryset.filter(department__id__in=clean_ids)
        except ValueError:
            pass
        return queryset

# 2. STATUS UCHUN MULTI-FILTER (Active, Pending va h.k. bir vaqtda tanlash uchun)
class StatusMultiFilter(admin.SimpleListFilter):
    title = "Holati"
    parameter_name = 'status'

    def lookups(self, request, model_admin):
        return [
            ('active', 'Faol'),
            ('pending', 'Kutilmoqda'),
            ('dismissed', "Bo'shatilgan"),
        ]

    def queryset(self, request, queryset):
        value = self.value()
        if not value:
            return queryset

        # "active,pending" -> ['active', 'pending']
        statuses = value.split(',')
        return queryset.filter(status__in=statuses)
class RoleListFilter(admin.SimpleListFilter):
    title = "Roli"
    parameter_name = 'role_filter'  # HTML/JS dagi nom bilan bir xil bo'lishi SHART

    def lookups(self, request, model_admin):
        # Admin panel o'ng tomonida chiqadigan filterlar (agar kerak bo'lsa)
        return [
            ('1', "O'qituvchi"),
            ('0', "Xodim"),
        ]

    def queryset(self, request, queryset):
        # URL dan kelgan qiymatni olamiz
        value = self.value()

        if not value:
            return queryset

        # JS dan '1,0' shaklida kelishi mumkin, shuning uchun split qilamiz
        roles = value.split(',')

        if '1' in roles and '0' in roles:
            return queryset
        elif '1' in roles:
            return queryset.filter(is_teacher=True)
        elif '0' in roles:
            return queryset.filter(Q(is_teacher=False) | Q(is_teacher__isnull=True))

        return queryset


class ScientificDegreeFilter(admin.SimpleListFilter):
    title = "Ilmiy daraja"
    parameter_name = 'scientific_degree'

    def lookups(self, request, model_admin):
        return Employee.DEGREE_CHOICES

    def queryset(self, request, queryset):
        value = self.value()
        if not value: return queryset
        degrees = value.split(',')
        return queryset.filter(scientific_degree__in=degrees)


class ScientificTitleFilter(admin.SimpleListFilter):
    title = "Ilmiy unvon"
    parameter_name = 'scientific_title'

    def lookups(self, request, model_admin):
        return Employee.TITLE_CHOICES

    def queryset(self, request, queryset):
        value = self.value()
        if not value: return queryset
        titles = value.split(',')
        return queryset.filter(scientific_title__in=titles)


# --- EMPLOYEE ADMIN (TO'LIQ) ---

@admin.register(Employee)
class EmployeeAdmin(admin.ModelAdmin):
    change_list_template = "admin/kadrlar/employee_change_list.html"

    # 1. RO'YXAT KO'RINISHI (List Display)
    list_display = ('get_full_name', 'department', 'get_positions_display',
                    'get_scientific_degree_display', 'get_scientific_title_display',
                    'status_badge', 'is_teacher_icon')

    # 2. FILTERLAR
    list_filter = (
        StatusMultiFilter,
        'gender',
        DepartmentMultiFilter,
        'positions',  # Lavozim bo'yicha filter (M2M)
        ScientificDegreeFilter,  # Yangilangan filter
        ScientificTitleFilter,  # Yangilangan filter
        'approved',
        RoleListFilter
    )

    search_fields = ('first_name', 'last_name', 'pid', 'passport_info')

    # ManyToMany maydonlar uchun qulay tanlash oynasi
    filter_horizontal = ('positions',)

    # 3. FORMA KO'RINISHI (Fieldsets)
    fieldsets = (
        ('Shaxsiy maʼlumot', {
            'fields': ('first_name', 'last_name', 'middle_name', 'gender', 'passport_info', 'pid', 'birth_date',
                       'photo')
        }),
        ('Ish joyi va Ilmiy Salohiyat', {
            'fields': ('department', 'positions', 'scientific_degree', 'scientific_title', 'is_teacher', 'order')
        }),
        ('HR Tasdiq', {
            'fields': ('hired_at', 'status', 'approved', 'archived')
        }),
    )

    # --- URLS VA CUSTOM VIEWS ---
    def get_urls(self):
        urls = super().get_urls()

        # MUHIM: Model nomini aniqlaymiz ('employee' yoki 'archivedemployee')
        # Bu URL nomlari to'qnashuvini oldini oladi
        model_name = self.model._meta.model_name

        custom_urls = [
            # 'name' endi dinamik bo'ldi: kadrlar_employee_export yoki kadrlar_archivedemployee_export
            path('export/', self.admin_site.admin_view(self.export_employees_view),
                 name=f'kadrlar_{model_name}_export'),
            path('<path:object_id>/card/', self.admin_site.admin_view(self.card_view), name='employee_card_print'),
        ]
        return custom_urls + urls

    def export_employees_view(self, request):
        """
        TUZATILGAN VERSION: ChangeList argumentlari to'g'irlandi.
        """
        print("\n" + "=" * 50)
        print(f">>> EXPORT SO'ROVI: {request.path}")

        if request.method != 'POST':
            return HttpResponse("Faqat POST so'rov qabul qilinadi", status=405)

        # 1. Tanlangan ustunlar
        selected_fields = request.POST.getlist('selected_fields')
        if not selected_fields:
            selected_fields = [
                'first_name', 'last_name', 'passport_info',
                'department', 'positions', 'status',
            ]

        # 2. QUERYSETNI FILTRLASH (ChangeList)
        # Javascript orqali kelgan filtrlarni tekshiramiz
        print(f">>> URL Parametrlari (GET): {request.GET}")

        from django.contrib.admin.views.main import ChangeList

        queryset = None

        try:
            # ChangeList uchun barcha kerakli argumentlarni yig'amiz
            list_display = self.get_list_display(request)
            list_display_links = self.get_list_display_links(request, list_display)
            list_filter = self.get_list_filter(request)
            search_fields = self.get_search_fields(request)
            list_select_related = self.get_list_select_related(request)

            # --- TUZATISH SHU YERDA ---
            # sortable_by va search_help_text argumentlarini xavfsiz olish
            sortable_by = getattr(self, 'sortable_by', ())
            search_help_text = getattr(self, 'search_help_text', None)

            cl = ChangeList(
                request,
                self.model,
                list_display,
                list_display_links,
                list_filter,
                self.date_hierarchy,
                search_fields,
                list_select_related,
                self.list_per_page,
                self.list_max_show_all,
                self.list_editable,
                self,
                sortable_by,  # <--- YANGI QO'SHILDI
                search_help_text  # <--- YANGI QO'SHILDI
            )

            # Filtrlangan ma'lumotlar
            queryset = cl.get_queryset(request)
            print(f">>> ChangeList muvaffaqiyatli ishladi. Natija: {queryset.count()} ta")

        except Exception as e:
            print(f">>> XATOLIK (ChangeList): {e}")
            # Xatolik bo'lsa, asosiy querysetni olamiz
            queryset = self.get_queryset(request)
            print(f">>> Fallback ishlatildi. Natija: {queryset.count()} ta")

        # 3. EXCEL YARATISH (OpenPyXL)
        import openpyxl
        import datetime
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Xodimlar"

        # Dizayn
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        # Field nomlari (Mapping)
        field_map = {
            'first_name': 'Ism', 'last_name': 'Familiya', 'middle_name': 'Otasining ismi',
            'passport_info': 'Pasport', 'pid': 'JSHSHIR', 'birth_date': 'Tug‘ilgan sana',
            'gender': 'Jinsi', 'department': "Bo‘lim / Kafedra", 'positions': 'Lavozimi',
            'status': 'Holati', 'hired_at': 'Ishga kirgan sana',
            'scientific_degree': 'Ilmiy daraja', 'scientific_title': 'Ilmiy unvon',
            'is_teacher': 'Roli', 'approved': 'HR Tasdiq'
        }

        headers = ["№"] + [field_map.get(f, f) for f in selected_fields]

        # Sarlavha yozish
        for col, title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=title)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

        ws.column_dimensions['A'].width = 5

        # Ma'lumotlarni to'ldirish
        if queryset and queryset.exists():
            for r, obj in enumerate(queryset, start=2):
                ws.cell(row=r, column=1, value=r - 1).border = border
                ws.cell(row=r, column=1).alignment = center

                for c, field in enumerate(selected_fields, start=2):
                    val = "-"
                    try:
                        if field == 'positions':
                            val = ", ".join([p.name for p in obj.positions.all()])
                        elif field == 'department':
                            val = obj.department.name if obj.department else "-"
                        elif field == 'status':
                            val = obj.get_status_display()
                        elif field == 'gender':
                            val = obj.get_gender_display()
                        elif field == 'scientific_degree':
                            val = obj.get_scientific_degree_display()
                        elif field == 'scientific_title':
                            val = obj.get_scientific_title_display()
                        elif field == 'is_teacher':
                            val = "O'qituvchi" if obj.is_teacher else "Xodim"
                        elif field == 'approved':
                            val = "Ha" if obj.approved else "Yo'q"
                        else:
                            raw = getattr(obj, field, "-")
                            val = str(raw) if raw is not None else ""
                    except Exception:
                        val = ""

                    cell = ws.cell(row=r, column=c, value=val)
                    cell.border = border
                    cell.alignment = left
        else:
            # Bo'sh bo'lsa
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
            cell = ws.cell(row=2, column=1, value="Ma'lumot topilmadi")
            cell.alignment = center
            cell.font = Font(italic=True, color="FF0000")

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"Export_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        print("=" * 50 + "\n")
        return response

    def get_field_verbose_name(self, field_name):
        try:
            return self.model._meta.get_field(field_name).verbose_name
        except:
            return field_name

    # --- METODLAR ---

    # 1. Lavozimlar (M2M)
    def get_positions_display(self, obj):
        return ", ".join([p.name for p in obj.positions.all()])

    get_positions_display.short_description = "Lavozimlar"

    # 2. Ilmiy Daraja (Wrapper)
    @admin.display(description="Ilmiy Daraja", ordering='scientific_degree')
    def get_scientific_degree_display(self, obj):
        return obj.get_scientific_degree_display()

    # 3. Ilmiy Unvon (Wrapper)
    @admin.display(description="Ilmiy Unvon", ordering='scientific_title')
    def get_scientific_title_display(self, obj):
        return obj.get_scientific_title_display()

    # 4. Boshqa ustunlar
    def get_full_name(self, obj):
        return f"{obj.last_name} {obj.first_name}"

    get_full_name.short_description = "F.I.SH"

    def passport_info(self, obj):
        return obj.passport_info if obj.passport_info else "-"

    passport_info.short_description = "Passport"

    def status_badge(self, obj):
        colors = {'active': '#10b981', 'pending': '#f59e0b', 'dismissed': '#ef4444'}
        color = colors.get(obj.status, '#64748b')
        return format_html(
            '<span style="color: white; background-color: {}; padding: 3px 10px; border-radius: 10px; font-size: 11px; font-weight: bold;">{}</span>',
            color, obj.get_status_display())

    status_badge.short_description = "HR Holati"

    def is_teacher_icon(self, obj):
        if obj.is_teacher:
            # Teacher profiliga o'tish uchun link
            url = reverse('admin:kadrlar_teacher_changelist') + f"?employee__id={obj.id}"
            return format_html(
                '<a href="{}" style="color: #3b82f6; font-weight:bold; text-decoration:none;"><i class="fas fa-chalkboard-teacher"></i> O\'qituvchi</a>',
                url)
        return format_html(
            '<span style="color: #64748b; font-weight:500;"><i class="fas fa-user-tie"></i> Xodim</span>')

    is_teacher_icon.short_description = "Roli"
    is_teacher_icon.allow_tags = True
    is_teacher_icon.admin_order_field = 'is_teacher'

    def view_card_icon(self, obj):
        url = reverse('admin:employee_card_print', args=[obj.pk])
        return format_html(
            '''<a href="{}" title="Profilni ko'rish"
                  style="background-color: #3b82f6; color: white; padding: 5px 10px; border-radius: 6px; text-decoration: none; font-weight: bold; display: inline-flex; align-items: center; gap: 5px;">
                <i class="fas fa-id-card"></i> Karta
               </a>''', url
        )

    view_card_icon.short_description = "Karta"
    view_card_icon.allow_tags = True

    # --- VIEW MOSLASHUVCHANLIGI (RUXSATLAR) ---
    def get_list_display(self, request):
        columns = list(super().get_list_display(request))
        if is_hr_admin(request.user):
            columns.append('view_card_icon')
        return columns

    def get_inlines(self, request, obj):
        if is_hr_admin(request.user):
            return [DocumentInline, OrderInline]
        return [DocumentInline]

    def get_readonly_fields(self, request, obj=None):
        if is_hr_admin(request.user):
            return ('created_by',)
        readonly = ['hired_at', 'status', 'approved', 'archived', 'created_by', 'order']
        if obj:
            readonly.append('department')
            if obj.approved:
                all_fields = [f.name for f in self.model._meta.fields]
                return all_fields
        return readonly

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if not is_hr_admin(request.user):
            dept = Department.objects.filter(head_manager=request.user).first()
            if 'department' in form.base_fields:
                form.base_fields['department'].disabled = True
                form.base_fields['department'].required = False
                if dept:
                    form.base_fields['department'].initial = dept.id
                    form.base_fields[
                        'department'].help_text = f"Siz faqat o'zingizning bo'limingiz ({dept.name}) ga xodim qo'sha olasiz."
        return form

    # --- KARTA KO'RISH VIEW ---
    def card_view(self, request, object_id):
        employee = get_object_or_404(Employee, pk=object_id)
        teacher_profile = getattr(employee, 'teacher_profile', None)

        raw_results = QuizResult.objects.filter(employee=employee).order_by('-created_at')
        enhanced_results = []
        for res in raw_results:
            data = res.struct if isinstance(res.struct, dict) else {}
            overall_text = data.get('overall_conclusion')
            if not overall_text or overall_text == "Natija izohi mavjud emas.":
                infos = QuizScoringInfo.objects.filter(quiz=res.quiz)
                for info in infos:
                    if info.min_score <= res.total_score <= info.max_score:
                        overall_text = info.conclusion
                        break
                if not overall_text:
                    overall_text = "Natija izohi mavjud emas."
            scoring_scale = QuizScoringInfo.objects.filter(quiz=res.quiz).order_by('min_score')
            enhanced_results.append({
                'obj': res,
                'struct': data,
                'calculated_conclusion': overall_text,
                'scale': scoring_scale
            })

        context = {
            **self.admin_site.each_context(request),
            'opts': self.model._meta,
            'employee': employee,
            'teacher': teacher_profile,
            'quiz_results': enhanced_results,
            'title': f"Profil: {employee.last_name} {employee.first_name}",
        }
        return TemplateResponse(request, 'admin/kadrlar/employee_card.html', context)

    # --- QUERYSET ---
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if is_hr_admin(request.user) or is_edu_admin(request.user):
            return qs.filter(archived=False)
        return qs.filter(department__head_manager=request.user, archived=False)

    # --- STATISTIKA (ChangeList View) ---
    def changelist_view(self, request, extra_context=None):
        # =========================================================
        # 1. DEFAULT FILTER (Agar filtr tanlanmagan bo'lsa -> Faol)
        # =========================================================
        if request.method == 'GET' and not request.GET:
            # Faqat 'Employee' modeli uchun ishlaydi (ArchivedEmployee uchun emas)
            if self.model.__name__ == 'Employee':
                # URLga ?status=active qo'shib yuboramiz
                return redirect(f"{request.path}?status=active")

        # =========================================================
        # 2. ASOSIY KOD (Eski kod o'z holicha qoladi)
        # =========================================================
        extra_context = extra_context or {}
        extra_context['is_kadr_member'] = is_hr_admin(request.user)

        response = super().changelist_view(request, extra_context=extra_context)

        if hasattr(response, 'context_data'):
            try:
                base_qs = self.get_queryset(request)
                # Faqat faol va arxivlanmagan xodimlarni olamiz
                active_qs = base_qs.filter(status='active', archived=False)

                # --- STATISTIKA KODI (O'zgarmaydi) ---
                def get_gender_stats(queryset):
                    total = queryset.count()
                    male = queryset.filter(gender='male').count()
                    female = queryset.filter(gender='female').count()
                    male_pct = round((male / total) * 100) if total > 0 else 0
                    female_pct = round((female / total) * 100) if total > 0 else 0
                    return {
                        'count': total,
                        'male': male,
                        'female': female,
                        'male_pct': male_pct,
                        'female_pct': female_pct
                    }

                stats = {
                    'all': get_gender_stats(active_qs),
                    'teachers': get_gender_stats(active_qs.filter(is_teacher=True)),
                    'staff': get_gender_stats(active_qs.filter(Q(is_teacher=False) | Q(is_teacher__isnull=True))),
                }

                # --- ILMIY SALOHIYAT ---
                science_qs = active_qs

                scientific_stats = {
                    'degree': {
                        'phd': science_qs.filter(scientific_degree='phd').count(),
                        'dsc': science_qs.filter(scientific_degree='dsc').count(),
                    },
                    'title': {
                        'docent': science_qs.filter(scientific_title='docent').count(),
                        'professor': science_qs.filter(scientific_title='professor').count(),
                        'academic': science_qs.filter(scientific_title='academic').count(),
                    }
                }

                departments = Department.objects.all().values('id', 'name')
                response.context_data['departments_json'] = json.dumps(list(departments))
                response.context_data['stats'] = stats
                response.context_data['sc_stats'] = scientific_stats

            except (KeyError, AttributeError):
                pass
        return response

    # --- SAVE METODLARI ---
    def save_model(self, request, obj, form, change):
        if not is_hr_admin(request.user):
            dept = Department.objects.filter(head_manager=request.user).first()
            if dept:
                obj.department = dept
            if not change:
                obj.status = 'pending'
                obj.approved = False
                obj.created_by = request.user
        super().save_model(request, obj, form, change)

    def save_formset(self, request, form, formset, change):
        instances = formset.save(commit=False)
        for instance in instances:
            if isinstance(instance, Order):
                if not is_hr_admin(request.user): continue
                instance.created_by = request.user
                emp = instance.employee
                if instance.order_type == 'hire':
                    emp.status = 'active'
                    emp.approved = True
                    emp.archived = False
                    emp.hired_at = instance.date if not emp.hired_at else emp.hired_at
                elif instance.order_type == 'dismiss':
                    emp.status = 'dismissed'
                    emp.archived = True
                emp.save()
            instance.save()
        formset.save_m2m()  # M2M (Positions) saqlanishi uchun muhim!

    # --- PERMISSIONS ---
    def has_add_permission(self, request):
        if is_hr_admin(request.user): return True
        if Department.objects.filter(head_manager=request.user).exists(): return True
        return False

    def has_change_permission(self, request, obj=None):
        if is_hr_admin(request.user): return True
        if obj and obj.department and obj.department.head_manager == request.user: return True
        return False

    def has_delete_permission(self, request, obj=None):
        return is_hr_admin(request.user)


@admin.register(ArchivedEmployee)
class ArchivedEmployeeAdmin(EmployeeAdmin):
    """
    Arxivdagi (bo'shatilgan) xodimlar uchun maxsus Admin klass.
    """
    # 1. Maxsus shablonni ulaymiz (faqat kerakli statistika chiqishi uchun)
    change_list_template = "admin/kadrlar/archive_change_list.html"

    # 2. Ro'yxatda ko'rinadigan ustunlar
    list_display = ('get_full_name', 'pid', 'department', 'hired_at', 'dismissed_date_col', 'restore_button')
    list_filter = ('department', 'gender')
    search_fields = ('first_name', 'last_name', 'pid')
    list_editable = ()
    # 3. Arxivdagi ma'lumotlarni tahrirlashni cheklash (Read-only)
    readonly_fields = [f.name for f in Employee._meta.fields]

    def get_queryset(self, request):
        """Faqat arxivlangan (archived=True) xodimlarni qaytaradi"""
        return self.model.objects.filter(archived=True)

    # --- RUXSATLAR (PERMISSIONS) ---
    def has_module_permission(self, request):
        """Faqat HR va Superuser ko'ra oladi"""
        return is_hr_admin(request.user)

    def has_view_permission(self, request, obj=None):
        return is_hr_admin(request.user)

    def has_change_permission(self, request, obj=None):
        return is_hr_admin(request.user)

    def has_add_permission(self, request):
        """Arxivga qo'lda odam qo'shib bo'lmaydi, faqat buyruq orqali tushadi"""
        return False

    def has_delete_permission(self, request, obj=None):
        return is_hr_admin(request.user)

    # --- QO'SHIMCHA USTUNLAR ---
    def dismissed_date_col(self, obj):
        """Oxirgi 'dismiss' buyrug'i sanasini chiqarish"""
        dismiss_order = obj.orders.filter(order_type='dismiss').order_by('-date').first()
        return dismiss_order.date if dismiss_order else "-"

    dismissed_date_col.short_description = "Bo'shatilgan sana"

    def restore_button(self, obj):
        return format_html('<span style="color: #f59e0b; font-weight:bold;">Arxivlangan</span>')

    restore_button.short_description = "Holati"

    # --- HARAKATLAR (ACTIONS) ---
    actions = ['restore_employees']

    def restore_employees(self, request, queryset):
        """Tanlanganlarni arxivdan chiqarib, Active holatiga qaytarish"""
        updated_count = queryset.update(archived=False, status='active')
        self.message_user(request, f"{updated_count} nafar xodim arxivdan chiqarildi va 'Faol' holatiga o'tkazildi.")

    restore_employees.short_description = "♻️ Tanlanganlarni arxivdan qaytarish (Active)"

    # --- STATISTIKA (ENG MUHIM QISMI) ---
    def changelist_view(self, request, extra_context=None):
        """
        Bu yerda biz 'Faol' va 'Kutilmoqda' statistikasini olib tashlaymiz.
        Faqat Jami va Jins bo'yicha hisob-kitob qilamiz.
        """
        response = super().changelist_view(request, extra_context=extra_context)

        # Agar sahifa muvaffaqiyatli yuklansa va kontekst bo'lsa
        if hasattr(response, 'context_data'):
            # Hozirgi filterdagi querysetni olamiz
            try:
                cl = response.context_data['cl']
                qs = cl.queryset

                total = qs.count()
                male = qs.filter(gender='male').count()
                female = qs.filter(gender='female').count()

                # STATISTIKA LUG'ATINI QAYTA YOZAMIZ
                # 'active' va 'vacation' kalitlari yo'q!
                stats = {
                    'total': total,
                    'male': male,
                    'male_percent': round((male / total) * 100) if total else 0,
                    'female': female,
                    'female_percent': round((female / total) * 100) if total else 0,
                }
                response.context_data['stats'] = stats
            except (KeyError, AttributeError):
                pass

        return response
# =========================================================
# 6. O'QITUVCHI ADMIN (O'QUV BO'LIMI)
# =========================================================

@admin.register(Teacher)
class TeacherAdmin(admin.ModelAdmin):
    # 1. RO'YXAT KO'RINISHI
    list_display = ('get_full_name', 'department_col', 'work_type_display',

                    'schedule_status_col')

    list_filter = ('schedule_approved', 'work_type_permanent', 'work_type_hourly',
                   'employee__department')

    autocomplete_fields = ['employee']
    filter_horizontal = ('subjects',)
    inlines = [TeacherAvailabilityInline]
    search_fields = ['employee__first_name', 'employee__last_name']
    # 2. FORMA KO'RINISHI
    fieldsets = (
        ("Xodim", {'fields': ('employee',)}),


        # ------------------------------

        ("Ishlash turi", {
            'fields': ('work_type_permanent', 'work_type_hourly'),
            'description': "O'qituvchi bir vaqtning o'zida ham doimiy, ham soatbay ishlashi mumkin."
        }),
        ("Yuklama",
         {'fields': ('subjects',)}),
        ("Tasdiqlash", {'fields': ('schedule_approved',), 'classes': ('collapse',)}),
    )

    # 3. QUERYSET (Kafedra mudiri faqat o'z xodimlarini ko'radi)
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        qs = qs.filter(employee__archived=False)

        if is_hr_admin(request.user) or is_edu_admin(request.user):
            return qs

        return qs.filter(employee__department__head_manager=request.user)

    # 4. TAHRIRLASHNI CHEKLASH (READONLY FIELDS)
    def get_readonly_fields(self, request, obj=None):
        # --- KADRLAR BO'LIMI (HR) ---
        if is_hr_admin(request.user):
            # HR hamma narsani ko'radi, lekin faqat SHTAT, DARJA va UNVONni o'zgartira oladi.
            # Boshqa narsalar (Fanlar, yuklama) HR uchun yopiq bo'ladi.

            readonly_cols = [f.name for f in self.model._meta.fields]

            # Tahrirlashga ruxsat berilgan maydonlar:
            editable_fields = [
                'work_type_permanent',
                'work_type_hourly',
            ]

            for field in editable_fields:
                if field in readonly_cols:
                    readonly_cols.remove(field)

            # Subjects M2M bo'lgani uchun alohida qo'shib qo'yamiz (HR o'zgartirmasligi uchun)
            return readonly_cols + ['subjects']

        # --- O'QUV BO'LIMI VA KAFEDRA ---
        # Ular shtat va ilmiy darajani faqat ko'radi, o'zgartira olmaydi.
        readonly = [
            'work_type_permanent',
            'work_type_hourly',

        ]

        edu_admin = is_edu_admin(request.user)

        if not edu_admin:
            readonly.append('schedule_approved')

        if obj:
            readonly.append('employee')
            # Agar tasdiqlangan bo'lsa, O'quv bo'limidan boshqa hech kim o'zgartira olmaydi
            if obj.schedule_approved and not edu_admin:
                readonly.extend([
                    'subjects',

                ])
        return readonly

    # 5. FILTR (Department bo'yicha)
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "employee" and not (is_hr_admin(request.user) or is_edu_admin(request.user)):
            kwargs["queryset"] = Employee.objects.filter(
                department__head_manager=request.user,
                is_teacher=True
            )
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    # 6. RUXSATLAR (Permissions)
    def has_add_permission(self, request):
        # HR Teacher profilini qo'sholmaydi (Faqat Kafedra yoki O'quv bo'limi)
        if is_hr_admin(request.user):
            return False
        return True

    def has_change_permission(self, request, obj=None):
        # HR ga o'zgartirish ruxsati KERAK (Ilmiy daraja va shtat uchun)
        if is_hr_admin(request.user):
            return True

        if is_edu_admin(request.user):
            return True
        if obj and obj.employee.department.head_manager == request.user:
            return True
        if obj is None: return True
        return False

    def has_delete_permission(self, request, obj=None):
        if is_hr_admin(request.user):
            return True
        if obj and obj.schedule_approved and not is_edu_admin(request.user):
            return False

        return True

    # 7. DIZAYN METODLARI
    def get_full_name(self, obj):
        return str(obj.employee)

    get_full_name.short_description = "O'qituvchi"

    def department_col(self, obj):
        return obj.employee.department

    department_col.short_description = "Kafedra / Bo'lim"

    def work_type_display(self, obj):
        tags = []
        if getattr(obj, 'work_type_permanent', False):
            tags.append(
                '<span style="background: #e0f2fe; color: #0369a1; padding: 2px 6px; border-radius: 4px; margin-right: 4px; font-size: 11px;">Doimiy</span>')
        if getattr(obj, 'work_type_hourly', False):
            tags.append(
                '<span style="background: #fef3c7; color: #b45309; padding: 2px 6px; border-radius: 4px; font-size: 11px;">Soatbay</span>')

        if not tags:
            return "-"
        return format_html("".join(tags))

    work_type_display.short_description = "Ishlash turi"

    def schedule_status_col(self, obj):
        return format_html(
            '<span style="background: {}; color: {}; padding: 4px 8px; border-radius: 4px; font-weight: bold;">{}</span>',
            '#dcfce7' if obj.schedule_approved else '#fee2e2',
            '#166534' if obj.schedule_approved else '#991b1b',
            '✅ Tasdiqlangan' if obj.schedule_approved else '⏳ Kutilmoqda')

    schedule_status_col.short_description = "O'quv Bo'limi"


# =========================================================
# 7. QOLGAN MODELLAR ADMINI
# =========================================================

@admin.register(Order)
class OrderAdmin(admin.ModelAdmin):
    list_display = ('number', 'employee', 'order_type', 'date')
    list_filter = ('order_type', 'date')
    search_fields = ('number', 'employee__first_name', 'employee__last_name', 'employee__pid')
    date_hierarchy = 'date'

    def has_module_permission(self, request):
        return is_hr_admin(request.user)

    def has_add_permission(self, request):
        return is_hr_admin(request.user)

    def has_change_permission(self, request, obj=None):
        return is_hr_admin(request.user)

    def has_delete_permission(self, request, obj=None):
        return is_hr_admin(request.user)


@admin.register(Document)
class DocumentAdmin(admin.ModelAdmin):
    list_display = ('doc_type', 'employee', 'number', 'uploaded_at')
    list_filter = ('doc_type',)
    search_fields = ('number', 'employee__first_name')

    def has_module_permission(self, request):
        return is_hr_admin(request.user)


@admin.register(TeacherAvailability)
class TeacherAvailabilityAdmin(admin.ModelAdmin):
    list_display = ('teacher', 'weekday')
    list_filter = ('weekday',)

    def has_module_permission(self, request):
        return is_edu_admin(request.user)


@admin.register(Weekday)
class WeekdayAdmin(admin.ModelAdmin):
    list_display = ('name', 'order')
    ordering = ('order',)

    def has_module_permission(self, request):
        return is_edu_admin(request.user) or is_hr_admin(request.user)


@admin.register(TimeSlot)
class TimeSlotAdmin(admin.ModelAdmin):
    list_display = ('index', 'start_time', 'end_time', 'is_active')
    ordering = ('index',)

    def has_module_permission(self, request):
        return is_edu_admin(request.user) or is_hr_admin(request.user)


class QuizResultKeyInline(NestedTabularInline):
    model = QuizResultKey
    extra = 0
    fields = ('code', 'description',)
    verbose_name = "Natija Kaliti (Masalan: A, B, C, D)"
    verbose_name_plural = "Test Kalitlari va Tavsiflari"
    classes = ['wide'] # Kengroq ko'rinish uchun

class QuizAnswerInline(NestedTabularInline):
    model = QuizAnswer
    extra = 0
    min_num = 1
    fields = ('text', 'symbol', 'score')
    verbose_name = "Javob varianti"
    verbose_name_plural = "Javob variantlari"
    # classes = ['collapse']  # Buni olib tashlab turing, ochilib turgani ma'qul

# 2. Savollar (O'rta qism - Savol va ichida Javoblar)
class QuizQuestionInline(NestedStackedInline):
    model = QuizQuestion
    extra = 0
    fields = ('text', 'order')
    inlines = [QuizAnswerInline] # <--- MANA SHU JAVOBLARNI CHIQARADI
    verbose_name = "Savol"
    verbose_name_plural = "Savollar"
class QuizScoringRuleInline(NestedTabularInline):
    model = QuizScoringRule
    extra = 0
    fields = ('category_name', 'related_questions', 'min_score', 'max_score', 'conclusion')
    verbose_name = "Natija talqini"
    verbose_name_plural = "Natija talqinlari (Min-Max ballar)"
    classes = ['wide']

class QuizScoringInfoInline(NestedTabularInline):
    model = QuizScoringInfo
    extra = 0
    fields = ('min_score', 'max_score', 'conclusion')
    verbose_name = "Natija bali izohi"
    verbose_name_plural = "Natija bali (oraliq ballari)"
    classes = ['wide']

# 3. Asosiy Quiz Admin
@admin.register(Quiz)
class QuizAdmin(NestedModelAdmin):
    list_display = ('title', 'question_count', 'is_active', 'created_at','id')
    inlines = [QuizQuestionInline,QuizResultKeyInline,QuizScoringRuleInline,QuizScoringInfoInline]
    search_fields = ('title',)

    def question_count(self, obj):
        return obj.questions.count()
    question_count.short_description = "Savollar soni"

    class Media:
        css = {
            'all': ('admin/css/admin_quiz.css',)
        }


@admin.register(QuizPermission)
class QuizPermissionAdmin(admin.ModelAdmin):
    list_display = ('employee', 'quiz', 'is_active', 'created_at')
    list_filter = ('quiz', 'is_active', 'employee__department')
    search_fields = ('employee__first_name', 'employee__last_name', 'employee__pid')
    autocomplete_fields = ['employee', 'quiz']
    actions = ['activate_permissions', 'deactivate_permissions']

    def activate_permissions(self, request, queryset):
        queryset.update(is_active=True)

    activate_permissions.short_description = "Tanlanganlarga qayta topshirishga RUXSAT berish"

    def deactivate_permissions(self, request, queryset):
        queryset.update(is_active=False)

    deactivate_permissions.short_description = "Ruxsatni YOPISH"


@admin.register(QuizResult)
class QuizResultAdmin(admin.ModelAdmin):
    list_display = ('employee', 'quiz', 'total_score', 'created_at')
    list_filter = ('quiz', 'created_at', 'employee__department')
    readonly_fields = ('formatted_struct',)
    exclude = ('struct',)  # Xom JSON ni yashiramiz

    def formatted_struct(self, obj):
        """JSON ma'lumotni chiroyli jadval shaklida chiqarish"""
        if not obj.struct:
            return "-"

        # 1. Ma'lumot turini tekshirish va to'g'irlash
        # Agar baza allaqachon DICT yoki LIST qaytargan bo'lsa, o'zini olamiz.
        if isinstance(obj.struct, (dict, list)):
            data = obj.struct
        else:
            # Agar string bo'lsa, json.loads qilamiz
            try:
                data = json.loads(obj.struct)
            except:
                return "Ma'lumot formati noto'g'ri"

        # 2. Javoblar ro'yxatini ajratib olish
        # Biz yangi formatda { "answers": [...], "analysis": [...] } qildik.
        # Shuning uchun 'answers' kalitini qidiramiz.
        answers_list = []

        if isinstance(data, dict):
            # Yangi format
            answers_list = data.get('answers', [])
        elif isinstance(data, list):
            # Eski format (agar eski testlar bo'lsa)
            answers_list = data

        # 3. Jadval chizish
        html = '<table style="width:100%; border-collapse: collapse; border: 1px solid #ddd;">'
        html += '<thead style="background:#f8f9fa;"><tr>' \
                '<th style="padding:10px; border:1px solid #ddd; text-align:left;">Savol</th>' \
                '<th style="padding:10px; border:1px solid #ddd; text-align:left;">Tanlangan Javob</th>' \
                '<th style="padding:10px; border:1px solid #ddd; text-align:center;">Ball</th></tr></thead><tbody>'

        for item in answers_list:
            # item dict ekanligiga ishonch hosil qilamiz
            if isinstance(item, dict):
                html += f"<tr>" \
                        f"<td style='padding:8px; border:1px solid #ddd;'>{item.get('question', '-')}</td>" \
                        f"<td style='padding:8px; border:1px solid #ddd;'>{item.get('selected', '-')}</td>" \
                        f"<td style='padding:8px; border:1px solid #ddd; text-align:center;'>{item.get('score', 0)}</td>" \
                        f"</tr>"

        html += '</tbody></table>'

        # Qo'shimcha: Agar Tahlil (Analysis) qismi bo'lsa, uni ham pastda ko'rsatish mumkin
        if isinstance(data, dict) and data.get('analysis'):
            html += '<h4 style="margin-top:20px;">Tahlil natijalari:</h4>'
            html += '<table style="width:100%; border-collapse: collapse; border: 1px solid #ddd;"><thead><tr style="background:#eef2ff;"><th>Kategoriya</th><th>Ball</th><th>Xulosa</th></tr></thead><tbody>'
            for anal in data['analysis']:
                html += f"<tr><td style='border:1px solid #ddd; padding:5px;'>{anal.get('category')}</td>" \
                        f"<td style='border:1px solid #ddd; padding:5px;'><b>{anal.get('score')}</b></td>" \
                        f"<td style='border:1px solid #ddd; padding:5px;'>{anal.get('conclusion')}</td></tr>"
            html += '</tbody></table>'

        return mark_safe(html)

    formatted_struct.short_description = "Batafsil Natijalar"

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False


original_each_context = admin.site.each_context


def get_new_context(request):
    """
    Tug'ilgan kunlarni faqat 'Kadrlar' guruhi a'zolariga ko'rsatish.
    Superuserga ko'rsatilmaydi.
    """
    context = original_each_context(request)
    user = request.user

    # 1. RUXSATNI ANIQLASH
    is_kadr_notification_viewer = False

    if user.is_authenticated:
        # O'ZGARISH SHU YERDA:
        # Biz user.is_superuser ni TEKSHIRMAYMIZ.
        # Faqatgina "Kadrlar" guruhida bor bo'lsa True bo'ladi.
        if user.groups.filter(name='Kadrlar').exists():
            is_kadr_notification_viewer = True

    # Context o'zgaruvchisini yangilaymiz
    context['is_kadr_member'] = is_kadr_notification_viewer

    # 2. TUG'ILGAN KUNLAR LOGIKASI (Faqat ruxsati borlarga hisoblanadi)
    if is_kadr_notification_viewer:
        today = timezone.now().date()
        tomorrow = today + timedelta(days=1)

        # Faol xodimlarni olamiz
        active_employees = Employee.objects.filter(status='active', archived=False).exclude(birth_date__isnull=True)

        birthdays_today = []
        birthdays_tomorrow = []

        for emp in active_employees:
            try:
                bday_this_year = emp.birth_date.replace(year=today.year)
            except ValueError:
                # 29-fevral muammosi
                bday_this_year = emp.birth_date.replace(year=today.year, day=28) + timedelta(days=1)

            if bday_this_year == today:
                birthdays_today.append(emp)
            elif bday_this_year == tomorrow:
                birthdays_tomorrow.append(emp)

        context['notify_birthdays_count'] = len(birthdays_today) + len(birthdays_tomorrow)
        context['notify_birthdays_today'] = birthdays_today
        context['notify_birthdays_tomorrow'] = birthdays_tomorrow
    else:
        # Superuser va boshqalar uchun bo'sh
        context['notify_birthdays_count'] = 0
        context['notify_birthdays_today'] = []
        context['notify_birthdays_tomorrow'] = []

    return context


# Funksiyani qayta ulaymiz
admin.site.each_context = get_new_context


@admin.register(OrganizationStructure)
class OrganizationStructureAdmin(admin.ModelAdmin):
    # Biz yaratgan maxsus shablonni ulaymiz
    change_form_template = "admin/kadrlar/org_structure_change_form.html"

    list_display = ('title', 'is_active', 'updated_at')

    # Kadrlar admini va superuser ko'ra oladi
    def has_module_permission(self, request):
        return is_hr_admin(request.user)

    def has_add_permission(self, request):
        return is_hr_admin(request.user)

    def has_change_permission(self, request, obj=None):
        return is_hr_admin(request.user)

    def has_delete_permission(self, request, obj=None):
        return is_hr_admin(request.user)


@admin.register(SimpleStructure)
class SimpleStructureAdmin(DraggableMPTTAdmin):
    # 1. Ro'yxatda nimalar ko'rinsin?
    list_display = (
        'tree_actions',
        'indented_title',
        'layout_display',
        'node_type_display',
        'mapping_info',
        'employee_count_display',
        'order'
    )
    list_display_links = ('indented_title',)

    # 2. Qidiruv va Autocomplete (Katta bazalar uchun qulay)
    search_fields = ('name',)
    autocomplete_fields = ['department', 'employee']

    # 3. Forma ko'rinishi (Guruhlarga bo'lingan)
    fieldsets = (
        ('Tugun Ma\'lumotlari', {
            'fields': ('name', 'parent', 'order')
        }),
        ('Dizayn va Joylashuv (Muhim)', {
            'fields': ('children_layout', 'node_type'),
        }),
        ('Kimni biriktiramiz? (Faqat bittasini tanlang)', {
            'fields': ('employee', 'department'),
        }),
    )

    # --- LIST DISPLAY METODLARI (Ro'yxatni chiroyli qilish uchun) ---

    def layout_display(self, obj):
        """Bolalar qanday joylashishini rangli qilib ko'rsatish"""
        if obj.children_layout == 'vertical':
            return format_html('<span style="color:#d97706; font-weight:bold;">⬇ Vertikal (Ustma-ust)</span>')
        return format_html('<span style="color:#059669;">➡ Gorizontal</span>')

    layout_display.short_description = "Joylashuv"

    def node_type_display(self, obj):
        """Tugun turini rangli bejiklar bilan ko'rsatish"""
        if obj.node_type == 'staff_left':
            return format_html(
                '<span style="background-color:#fee2e2; color:#991b1b; padding:3px 8px; border-radius:12px; font-size:11px; font-weight:bold; border:1px solid #fecaca;">⬅ Chap (Shtat)</span>'
            )
        elif obj.node_type == 'staff_right':
            return format_html(
                '<span style="background-color:#e0f2fe; color:#075985; padding:3px 8px; border-radius:12px; font-size:11px; font-weight:bold; border:1px solid #bae6fd;">➡ O\'ng (Shtat)</span>'
            )
        return format_html('<span style="color:#64748b;">Oddiy</span>')


    node_type_display.short_description = "Turi (Pozitsiya)"

    def mapping_info(self, instance):
        """Kim biriktirilganini ko'rsatish"""
        if instance.employee:
            return format_html(
                f'<span style="color:#333;">👤 {instance.employee.last_name} {instance.employee.first_name}</span>')
        if instance.department:
            return format_html(f'<span style="color:#333; font-weight:bold;">🏢 {instance.department.name}</span>')
        return format_html('<span style="color:#999;">❌ Biriktirilmagan</span>')

    mapping_info.short_description = "Biriktirilgan"

    def employee_count_display(self, instance):
        count = instance.get_employee_count()
        if count > 0:
            return format_html(
                f'<span style="background:#dcfce7; color:#166534; padding:2px 8px; border-radius:10px; font-weight:bold;">{count} nafar</span>')
        return "-"

    employee_count_display.short_description = "Xodimlar"

    # --- URLS va VIEWS (Diagrammani chizish va API uchun) ---

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('visual-chart/', self.admin_site.admin_view(self.visual_chart_view), name='simplestructure_visual'),
            path('api-node-details/<int:node_id>/', self.admin_site.admin_view(self.node_details_api),
                 name='simplestructure_api'),
        ]
        return my_urls + urls

    def visual_chart_view(self, request):
        # Vizual ko'rinish sahifasi
        context = dict(
            self.admin_site.each_context(request),
            nodes=SimpleStructure.objects.all(),
            title="Tashkiliy Tuzilma (Vizual)"
        )
        return render(request, 'admin/kadrlar/simplestructure/chart_view.html', context)

    def node_details_api(self, request, node_id):
        # Modal oynasi uchun JSON qaytaruvchi API
        node = get_object_or_404(SimpleStructure, id=node_id)
        employees = node.get_employees()

        data = []
        for emp in employees:
            # Xodimning o'z lavozimlarini olamiz
            pos_list = ", ".join([p.name for p in emp.positions.all()])

            # Rasmni tekshirish
            if emp.photo:
                photo_url = emp.photo.url
            else:
                photo_url = "/static/img/default-user.png"

            data.append({
                'id': emp.id,  # <--- BU ID BO'LISHI SHART
                'full_name': f"{emp.last_name} {emp.first_name}",
                'position': pos_list,
                'photo': photo_url,
                'degree': emp.get_scientific_degree_display(),
            })

        return JsonResponse({
            'node_name': node.name,
            'employees': data
        })