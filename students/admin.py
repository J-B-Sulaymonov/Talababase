from itertools import groupby
from django.db.models.functions import Coalesce, TruncMonth, Lower
import json
from openpyxl.utils import get_column_letter
from collections import defaultdict
import re
import openpyxl
from import_export.widgets import DateWidget, ForeignKeyWidget, NumberWidget, Widget
from django.core.exceptions import ObjectDoesNotExist
from datetime import datetime, date, timedelta
from django.utils.safestring import mark_safe
from django import forms
from django.core.exceptions import ValidationError
from import_export.admin import ImportExportModelAdmin, ImportExportMixin
from decimal import Decimal, ROUND_HALF_UP
from django.db.models import Count, Prefetch, ExpressionWrapper
from django.db.models import (
    Sum, F, OuterRef, Subquery, Exists, Q, Value, Case, When, DecimalField
)
from django.http import HttpResponse, QueryDict, JsonResponse

from io import BytesIO
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.utils.html import format_html
from django.utils.translation import gettext_lazy as _
from import_export import resources, fields
from .models import (
    Country, Region, District,
    Specialty, Group, Student,
    Contract, Payment, Order, OrderType,
    AcademicYear, SubjectDebt, PerevodRate, Subject, Hisobot, StudentHistory,SubjectRate
)
from django.urls import path, reverse
from django.shortcuts import render
from django.contrib import admin

from django import forms

# 1. Maxsus pul maydoni (Probellarni tozalab qabul qiladi)
class MoneyField(forms.DecimalField):
    def to_python(self, value):
        if value in self.empty_values:
            return None
        if isinstance(value, str):
            # 1. Oddiy probelni olib tashlash
            value = value.replace(' ', '')
            # 2. MUHIM: JS formatlashdan kelgan maxsus probelni (\xa0) olib tashlash
            value = value.replace('\xa0', '')
            # 3. Vergulni olib tashlash
            value = value.replace(',', '')
        return super().to_python(value)

    def prepare_value(self, value):
        # Agar qiymat bazadan kelsa, uni shundayligicha qaytaramiz.
        # Formatlashni JS bajaradi.
        return super().prepare_value(value)

# 2. Contract (Shartnoma) uchun Form
class ContractForm(forms.ModelForm):
    # Shartnoma summasi
    amount = MoneyField(
        label="Shartnoma summasi",
        widget=forms.TextInput(attrs={'class': 'money-input', 'style': 'font-weight: bold; font-size: 1.1em; width: 150px;'})
    )
    # Grant summasi (Siz so'ragan joy)
    grant_amount = MoneyField(
        label="Grant summasi",
        required=False,
        widget=forms.TextInput(attrs={'class': 'money-input', 'style': 'width: 150px;'})
    )

    class Meta:
        model = Contract
        fields = '__all__'

class SubjectRateForm(forms.ModelForm):
    amount = MoneyField(
        label="Kontrakt narxi",
        widget=forms.TextInput(attrs={'class': 'money-input', 'style': 'font-weight: bold; font-size: 1.1em; width: 200px;'})
    )

    class Meta:
        model = SubjectRate
        fields = '__all__'

class PerevodRateForm(forms.ModelForm):
    # amount maydoniga "money-input" klassini beramiz
    amount = MoneyField(
        label="1 kredit narxi",
        widget=forms.TextInput(attrs={'class': 'money-input', 'style': 'font-weight: bold; font-size: 1.1em; width: 200px;'})
    )

    class Meta:
        model = PerevodRate
        fields = '__all__'

# 3. Payment (To'lov) uchun Form
class PaymentForm(forms.ModelForm):
    amount = MoneyField(
        label="To'lov summasi",
        widget=forms.TextInput(attrs={'class': 'money-input', 'style': 'font-weight: bold; color: green;'})
    )

    class Meta:
        model = Payment
        fields = '__all__'

# 4. Fan qarzlari uchun Form
class SubjectDebtForm(forms.ModelForm):
    amount = MoneyField(label="Qarzdorlik", required=False, widget=forms.TextInput(attrs={'class': 'money-input'}))
    amount_summ = MoneyField(label="To'lov", required=False, widget=forms.TextInput(attrs={'class': 'money-input'}))

    class Meta:
        model = SubjectDebt
        fields = '__all__'


def students_general_view(request):
    models_links = [
        {"title": "Yo'nalishlar", "subtitle": "Specialty", "url": reverse('admin:students_specialty_changelist'), "icon": "fas fa-graduation-cap"},
        {"title": "Guruhlar",      "subtitle": "Group",     "url": reverse('admin:students_group_changelist'),     "icon": "fas fa-users"},
        {"title": "Davlatlar",     "subtitle": "Country",   "url": reverse('admin:students_country_changelist'),   "icon": "fas fa-globe"},
        {"title": "Viloyatlar",    "subtitle": "Region",    "url": reverse('admin:students_region_changelist'),    "icon": "fas fa-map"},
        {"title": "Tumanlar",      "subtitle": "District",  "url": reverse('admin:students_district_changelist'),  "icon": "fas fa-map-marker-alt"},
        {"title": "Buyruq turlari","subtitle": "OrderType", "url": reverse('admin:students_ordertype_changelist'), "icon": "fas fa-file-signature"},
        {"title": "Buyruqlar",     "subtitle": "Order",     "url": reverse('admin:students_order_changelist'),     "icon": "fas fa-clipboard-list"},
        {"title": "Shartnomalar",  "subtitle": "Contract",  "url": reverse('admin:students_contract_changelist'),  "icon": "fas fa-file-contract"},
        {"title": "To'lovlar",     "subtitle": "Payment",   "url": reverse('admin:students_payment_changelist'),   "icon": "fas fa-money-bill-wave"},
        {"title": "O'quv yillari", "subtitle": "AcademicYear","url": reverse('admin:students_academicyear_changelist'),"icon": "fas fa-calendar-alt"},
        {"title": "Fanlar",        "subtitle": "Subject",   "url": reverse('admin:students_subject_changelist'),   "icon": "fas fa-book"},
        {"title": "Fan qarzlari",  "subtitle": "SubjectDebt","url": reverse('admin:students_subjectdebt_changelist'),"icon": "fas fa-exclamation-triangle"},
        {"title": "Perevod stavkasi","subtitle":"PerevodRate","url": reverse('admin:students_perevodrate_changelist'),"icon":"fas fa-credit-card"},
        {"title": "Fan stavkasi (DU)", "subtitle": "SubjectRate", "url": reverse('admin:students_subjectrate_changelist'), "icon": "fas fa-tags"},
    ]

    # MUHIM: admin.site.each_context bilan global admin kontekstini olish
    context = admin.site.each_context(request)
    # keyin o'zimiz uchun kerakli qiymatlarni qo'shamiz
    context.update({
        'title': "Umumiy bloklar",
        'models_links': models_links,
    })

    return render(request, 'admin/students/general.html', context)

original_get_urls = admin.site.get_urls

def get_urls():
    custom_urls = [
        path('students/general/', admin.site.admin_view(students_general_view), name='students_general'),

    ]
    return custom_urls + original_get_urls()

admin.site.get_urls = get_urls

# =============================================================================
# 🌍 JOYLAShUV MODELLARI
# (O'zgarishsiz qoldirildi)
# =============================================================================
class CountryResource(resources.ModelResource):
    class Meta:
        model = Country


@admin.register(Country)
class CountryAdmin(ImportExportModelAdmin):
    resource_class = CountryResource
    list_display = ('name', 'id')
    search_fields = ('name',)


class RegionResource(resources.ModelResource):
    class Meta:
        model = Region


@admin.register(Region)
class RegionAdmin(ImportExportModelAdmin):
    resource_class = RegionResource
    list_display = ('name', 'country', 'id')
    list_filter = ('country',)
    search_fields = ('name',)
    autocomplete_fields = ['country']


class DistrictResource(resources.ModelResource):
    class Meta:
        model = District


@admin.register(District)
class DistrictAdmin(ImportExportModelAdmin):
    resource_class = DistrictResource
    list_display = ('name', 'region', 'id')
    list_filter = ('region__country', 'region')
    search_fields = ('name',)
    autocomplete_fields = ['region']


# =============================================================================
# 🎓 TA'LIMGA OID MODELLAR
# (O'zgarishsiz qoldirildi)
# =============================================================================
class SpecialtyResource(resources.ModelResource):
    class Meta:
        model = Specialty


@admin.register(Specialty)
class SpecialtyAdmin(ImportExportModelAdmin):
    resource_class = SpecialtyResource
    list_display = ('name', 'code', 'id')
    search_fields = ('name', 'code')


class GroupResource(resources.ModelResource):
    specialty = fields.Field(
        column_name="Yo'nalishi",
        attribute='specialty__name'
    )

    # YANGI USTUNLAR
    group_course = fields.Field(
        column_name="Kursi",
        readonly=True
    )

    education_form = fields.Field(
        column_name="Ta'lim shakli",
        readonly=True
    )

    total_students = fields.Field(
        column_name="Jami talabalar soni",
        readonly=True
    )

    filtered_students_count = fields.Field(
        column_name="Filtr bo'yicha talabalar soni",
        readonly=True
    )

    class Meta:
        model = Group
        # Excelda chiqadigan barcha ustunlar ro'yxati
        fields = ('id', 'name', 'specialty', 'group_course', 'education_form', 'total_students',
                  'filtered_students_count')
        export_order = ('id', 'name', 'specialty', 'group_course', 'education_form', 'total_students',
                        'filtered_students_count')

    # MA'LUMOTLARNI OLISH LOGIKASI
    def dehydrate_group_course(self, group):
        """Guruhdagi birinchi talaba kursini oladi"""
        student = group.student_set.first()
        return f"{student.course_year}-kurs" if student else "-"

    def dehydrate_education_form(self, group):
        """Guruhdagi birinchi talaba ta'lim shaklini oladi"""
        student = group.student_set.first()
        return student.get_education_form_display() if student else "-"

    def dehydrate_total_students(self, group):
        return group.student_set.count()

    def dehydrate_filtered_students_count(self, group):
        return getattr(group, 'student_count', 0)


# =============================================================================
# 1. TALABALARNI GURUH ICHIDA KO'RSATISH UCHUN INLINE
# =============================================================================
class StudentInline(admin.TabularInline):
    model = Student
    fields = ('full_name', 'student_hemis_id', 'phone_number', 'status', 'contract_amount_display')
    readonly_fields = ('full_name', 'student_hemis_id', 'phone_number', 'status', 'contract_amount_display')
    can_delete = False
    extra = 0
    show_change_link = True  # Talabaning shaxsiy sahifasiga o'tish tugmasi
    verbose_name = "Guruh talabasi"
    verbose_name_plural = "Guruh talabalari"

    def contract_amount_display(self, obj):
        # Talabaning joriy yilgi shartnoma summasini ko'rsatish (ixtiyoriy)
        contract = obj.contract_set.last()  # Yoki active yil bo'yicha filter
        return f"{contract.amount:,.0f}" if contract else "-"

    contract_amount_display.short_description = "Shartnoma"


# =============================================================================
# 2. GURUH ADMINI (YANGILANGAN)
# =============================================================================
@admin.register(Group)
class GroupAdmin(ImportExportModelAdmin):
    resource_class = GroupResource
    list_display = ('name', 'specialty', 'get_student_count', 'view_students_link', 'id')

    list_filter = (
        'specialty',
        ('student__status', admin.ChoicesFieldListFilter),
        ('student__course_year', admin.ChoicesFieldListFilter),
        ('student__education_form', admin.ChoicesFieldListFilter),
    )
    search_fields = ('name',)

    def apply_student_count_annotation(self, queryset, request):
        """Filtrlar asosida dinamik hisoblash (Annotate)"""
        status_filter = request.GET.get('student__status__exact')
        course_filter = request.GET.get('student__course_year__exact')
        form_filter = request.GET.get('student__education_form__exact')

        count_filter = Q()
        if status_filter:
            count_filter &= Q(student__status=status_filter)
        if course_filter:
            count_filter &= Q(student__course_year=course_filter)
        if form_filter:
            count_filter &= Q(student__education_form=form_filter)

        return queryset.annotate(
            student_count=Count('student', filter=count_filter)
        )

    def get_queryset(self, request):
        """Admin paneli uchun"""
        qs = super().get_queryset(request)
        return self.apply_student_count_annotation(qs, request)

    def get_export_queryset(self, request):
        """Excel eksporti uchun"""
        qs = super().get_export_queryset(request)
        return self.apply_student_count_annotation(qs, request)

    @admin.display(description='Talabalar soni', ordering='student_count')
    def get_student_count(self, obj):
        return obj.student_count

    @admin.display(description="Ro'yxat")
    def view_students_link(self, obj):
        url = reverse('admin:students_student_changelist')
        return format_html(
            '<a class="button" style="background-color: #264b5d; color: white; padding: 5px 10px; border-radius: 4px;" '
            'href="{}?group__id__exact={}">Talabalarni ko\'rish</a>',
            url, obj.id
        )


# =============================================================================
# 🧩 TALABA ADMIN YORDAMCHI KLASSLAR
# (O'zgarishsiz qoldirildi)
# =============================================================================
class DatalistTextInput(forms.TextInput):
    def __init__(self, datalist, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._datalist = sorted(list(datalist))

    def render(self, name, value, attrs=None, renderer=None):
        if attrs is None:
            attrs = {}
        attrs['list'] = f'{name}_datalist'
        input_html = super().render(name, value, attrs)
        datalist_html = f'<datalist id="{attrs["list"]}">'
        for item in self._datalist:
            if item:
                datalist_html += f'<option value="{item}">'
        datalist_html += '</datalist>'
        return mark_safe(f'{input_html}{datalist_html}')


# =============================================================================
# 🧾 BUYRUQLAR INLINE FORM
# (O'zgarishsiz qoldirildi)
# =============================================================================
class OrderInlineForm(forms.ModelForm):
    class Meta:
        model = Order
        fields = '__all__'

    def clean(self):
        cleaned_data = super().clean()
        order_type = cleaned_data.get('order_type')

        if order_type and order_type.id == 2:
            tsch_reason = cleaned_data.get('tsch_reason')
            application_date = cleaned_data.get('application_date')
            document_taken_date = cleaned_data.get('document_taken_date')

            missing_fields = []
            if not tsch_reason:
                missing_fields.append("TSCH sababi")
            if not application_date:
                missing_fields.append("Ariza sanasi")
            if not document_taken_date:
                missing_fields.append("Hujjat olib ketilgan sanasi")

            if missing_fields:
                raise ValidationError(
                    f"Quyidagi maydonlarni to‘ldirish majburiy: {', '.join(missing_fields)}"
                )

        return cleaned_data


# =============================================================================
# 📑 INLINE MODELLAR
# (O'zgarishsiz qoldirildi)
# =============================================================================

class StudentHistoryInline(admin.TabularInline):
    model = StudentHistory
    extra = 1  # 1 ta bo'sh qator doim ko'rinib turadi (qo'shish oson bo'lishi uchun)

    fields = ('academic_year', 'group', 'course_year',  'education_form')


    ordering = ('-academic_year__name',)
    can_delete = True
    verbose_name = "O'quv yili tarixi"
    verbose_name_plural = "O'quv yillari bo'yicha tarixi"


class ContractInline(admin.StackedInline):
    model = Contract
    form = ContractForm
    extra = 0
    fields = (
        'academic_year', 'contract_type', 'contract_number', 'contract_date',
        'amount',
        'grant_type', 'grant_date', 'grant_percent',
        'grant_amount',
    )
    class Media:
        js = (
            'admin/js/contract_grant_auto_summ.js',
            )

    def save_model(self, request, obj, form, change):
        obj.save()


class OrderInline(admin.StackedInline):
    model = Order
    form = OrderInlineForm
    extra = 0
    fields = (
        'order_type', 'order_number', 'order_date',
        'application_date','tsch_by_whom',  'tsch_reason', 'notes', 'document_taken_date',
    )

    class Media:
        js = ('admin/js/order_inline_dynamic_fields.js',)




class SubjectDebtInline(admin.StackedInline):
    model = SubjectDebt
    form = SubjectDebtForm
    extra = 0
    readonly_fields = ['amount']
    raw_id_fields = ['subject']   # contractni ham qoldirish mumkin, lekin biz uni filterlaymiz
    fields = (
        'subject', 'academic_year', 'semester',
        'year_credit', 'credit', 'debt_type', 'amount',
        'contract', 'amount_summ', 'payment_date', 'status'
    )
    class Media:
        js = ('admin/js/disable_text_input_and_open_lookup.js',)

    def get_formset(self, request, obj=None, **kwargs):
        """
        obj = parent Student obyektidir (agar mavjud bo'lsa).
        Biz form klassining base_fields orqali 'contract' maydonining querysetini
        faqat shu studentga tegishlilarga limitlaymiz.
        """
        FormSet = super().get_formset(request, obj, **kwargs)

        # Agar obj mavjud bo'lsa (yani edit rejimida) -> kontraktlarni filterlaymiz
        try:
            if obj is not None:
                # form class ning maydonlarini o'zgartiramiz
                if 'contract' in FormSet.form.base_fields:
                    FormSet.form.base_fields['contract'].queryset = Contract.objects.filter(student=obj)
            else:
                # Yangi student yaratishda kontrakt bo'lmaydi -> bo'sh queryset
                if 'contract' in FormSet.form.base_fields:
                    FormSet.form.base_fields['contract'].queryset = Contract.objects.none()
        except Exception:
            # fallback: hech qanday o'zgartirish qilmasdan qaytaramiz
            pass

        return FormSet




# =============================================================================
# 🎓 TALABA ADMIN (IMPORT/EXPORT)
# (O'zgarishsiz qoldirildi)
# =============================================================================
class StudentResource(resources.ModelResource):
    # 1. Mavjud ForeignKey maydonlar (Import/Export uchun)
    group = fields.Field(
        column_name='group',
        attribute='group',
        widget=ForeignKeyWidget(Group, 'name')
    )

    citizenship = fields.Field(
        column_name='citizenship',
        attribute='citizenship',
        widget=ForeignKeyWidget(Country, 'id')
    )

    region = fields.Field(
        column_name='region',
        attribute='region',
        widget=ForeignKeyWidget(Region, 'id')
    )

    district = fields.Field(
        column_name='district',
        attribute='district',
        widget=ForeignKeyWidget(District, 'id')
    )

    previous_education_country = fields.Field(
        column_name='previous_education_country',
        attribute='previous_education_country',
        widget=ForeignKeyWidget(Country, 'name')
    )

    previous_education_region = fields.Field(
        column_name='previous_education_region',
        attribute='previous_education_region',
        widget=ForeignKeyWidget(Region, 'name')
    )

    # -------------------------------------------------------------------------
    # YANGI QO'SHILGAN MAYDONLAR (Qabul buyrug'i ma'lumotlari)
    # -------------------------------------------------------------------------
    qabul_order_number = fields.Field(
        column_name='Qabul buyruq raqami',
        readonly=True
    )

    qabul_order_date = fields.Field(
        column_name='Qabul buyruq sanasi',
        readonly=True
    )

    class Meta:
        model = Student
        import_id_fields = ['student_hemis_id']
        exclude = ('id', 'created_at', 'updated_at')
        skip_unchanged = True
        report_skipped = True
        use_bulk = False
        # Agar ustunlar ketma-ketligini o'zingiz xohlagandek qilmoqchi bo'lsangiz,
        # export_order ni ishlatishingiz mumkin. Hozircha default qoldiramiz.

    # -------------------------------------------------------------------------
    # YANGI LOGIKA: Buyruq ma'lumotlarini olish (Dehydrate)
    # -------------------------------------------------------------------------
    def dehydrate_qabul_order_number(self, student):
        # order_type_id=1 (Qabul) bo'lgan eng oxirgi buyruqni olamiz
        order = student.orders.filter(order_type_id=1).order_by('-order_date').first()
        return order.order_number if order else ""

    def dehydrate_qabul_order_date(self, student):
        order = student.orders.filter(order_type_id=1).order_by('-order_date').first()
        if order and order.order_date:
            return order.order_date.strftime('%d.%m.%Y')
        return ""

    # -------------------------------------------------------------------------
    # ESKI LOGIKA: Import jarayoni uchun (O'zgarishsiz)
    # -------------------------------------------------------------------------
    def get_instance(self, instance_loader, row):
        hemis_id = row.get('student_hemis_id')
        if hemis_id:
            try:
                return Student.objects.get(student_hemis_id=str(hemis_id).strip())
            except Student.DoesNotExist:
                return None
        return None

    def before_import_row(self, row, **kwargs):
        map_dictionaries = {
            'education_type': {
                'To‘lov-shartnoma': 'contract',
                'Davlat granti': 'grant',
            },
            'gender': {
                'Erkak': 'erkak',
                'Ayol': 'ayol',
            },
            'education_form': {
                'Kunduzgi': 'kunduzgi',
                'Sirtqi': 'sirtqi',
                'Kechki': 'kechki',
            },
            'status': {
                "O'qiydi": 'active',
                "Akademik ta'tilda": 'academic',
                "Chetlashtirilgan": 'expelled',
                "Bitirgan": 'graduated',
                'active': 'active',
            }
        }

        for field_name, mapping in map_dictionaries.items():
            if field_name in row and row[field_name] in mapping:
                row[field_name] = mapping[row[field_name]]

        text_fields_to_protect = [
            'address',
            'passport_issued_by',
            'phone_number',
            'passport_series_number',
            'personal_pin',
            'full_name'
        ]

        for field_name in text_fields_to_protect:
            if field_name in row and (row[field_name] == '' or row[field_name] is None):
                if field_name == 'phone_number':
                    row[field_name] = '000000000'
                else:
                    row[field_name] = 'Kiritilmagan'

        numeric_fields = [
            'entry_score',
            'previous_graduation_year',
            'citizenship',
            'region',
            'district',
        ]

        for field_name in numeric_fields:
            if field_name in row and (row[field_name] == '' or row[field_name] is None):
                row[field_name] = None

        defaulted_numeric_fields = [
            'course_year',
            'current_semester',
        ]

        for field_name in defaulted_numeric_fields:
            if field_name in row and (row[field_name] == '' or row[field_name] is None):
                row[field_name] = None

        foreign_key_fields = [
            'group',
            'previous_education_country',
            'previous_education_region'
        ]

        for field_name in foreign_key_fields:
            if field_name in row and (row[field_name] == '' or row[field_name] is None):
                row[field_name] = None

        date_fields = [
            'date_of_birth',
            'passport_issue_date',
            'passport_expiry_date'
        ]

        for field_name in date_fields:
            if field_name in row and (row[field_name] == '' or row[field_name] is None):
                row[field_name] = None




# =============================================================================
# 🔍 YANGI FILTERLAR
# =============================================================================

class AcademicYearFilter(admin.SimpleListFilter):
    title = _("O‘quv yili (Hisob-kitob uchun)")
    parameter_name = 'academic_year_calc'

    def lookups(self, request, model_admin):
        years = AcademicYear.objects.all().order_by('-name')
        # ID larni string qilib qaytaramiz
        return [(str(y.id), y.name) for y in years]

    def queryset(self, request, queryset):
        # Bu yerda hech narsa qilmaymiz, chunki logika StudentAdmin.get_queryset da
        return queryset

class StatusFilter(admin.SimpleListFilter):
    title = "Status"
    parameter_name = 'status'

    def lookups(self, request, model_admin):
        return (
            ('all', "Barcha talabalar"),
            ('active', "O'qiydi"),
            ('academic', "Akademik ta'tilda"),
            ('expelled', "Chetlashtirilgan"),
            ('graduated', "Bitirgan"),
        )

    def choices(self, cl):
        # BU YERDA SEHR QILINADI: Standart "Hammasi" ni olib tashlaymiz
        for lookup, title in self.lookup_choices:
            # Agar URLda status bo'lmasa -> active tanlangan bo'lsin
            selected = (self.value() == str(lookup)) or (self.value() is None and lookup == 'active')
            yield {
                'selected': selected,
                'query_string': cl.get_query_string({self.parameter_name: lookup}, []),
                'display': title,
            }

    def queryset(self, request, queryset):
        # Mantiq: Hech narsa tanlanmasa -> faqat 'active'
        if self.value() is None or self.value() == 'active':
            return queryset.filter(status='active')
        elif self.value() == 'all':
            return queryset # Hammasi kerak bo'lsa
        else:
            return queryset.filter(status=self.value())


# =========================================================
# 🛠 FILTER KLASSLAR (TO'LIQ RO'YXAT)
# =========================================================

# 1. TO'LOV FOIZI FILTERI (QAYTARILDI)
class PaymentPercentFilter(admin.SimpleListFilter):
    title = "To'lov foizi"
    parameter_name = 'payment_percent'

    def lookups(self, request, model_admin):
        return (
            ('0', '0% (To\'lamagan)'),
            ('1-24', '1% - 24%'),
            ('25', '25%'),
            ('26-49', '26% - 49%'),
            ('50', '50%'),
            ('51-74', '51% - 74%'),
            ('75', '75%'),
            ('76-99', '76% - 99%'),
            ('100', '100% (To\'liq)'),
            ('over', 'Ortiqcha (>100%)'),
        )

    def queryset(self, request, queryset):
        value = self.value()
        if not value:
            return queryset

        selected_values = value.split(',')
        combined_query = Q()

        for val in selected_values:
            # --- ORALIQLARDA BO'SHLIQ QOLMASLIGI UCHUN "LT" (<) ISHLATAMIZ ---

            if val == '0':
                # 1% dan kam (masalan 0.5% ham kiradi)
                combined_query |= Q(payment_percent__lt=1)

            elif val == '25':
                # 25.0 dan boshlab 26 gacha
                combined_query |= Q(payment_percent__gte=25, payment_percent__lt=26)

            elif val == '50':
                combined_query |= Q(payment_percent__gte=50, payment_percent__lt=51)

            elif val == '75':
                combined_query |= Q(payment_percent__gte=75, payment_percent__lt=76)

            elif val == '100':
                combined_query |= Q(payment_percent__gte=100, payment_percent__lt=101)

            elif val == 'over':
                combined_query |= Q(payment_percent__gte=101)

            elif '-' in val:
                # Masalan "1-24".
                # Biz buni "1 dan katta yoki teng" VA "25 dan KICHIK" deb olishimiz kerak.
                # Shunda 24.99% ham kirib ketadi.
                try:
                    parts = val.split('-')
                    min_val = int(parts[0])
                    max_val = int(parts[1])

                    # DIQQAT: max_val ga +1 qo'shib, 'lt' (kichik) ishlatamiz
                    # Masalan: 1-24 bo'lsa -> 1 dan 25 gacha (25 kirmaydi)
                    combined_query |= Q(payment_percent__gte=min_val, payment_percent__lt=max_val + 1)
                except ValueError:
                    pass

        return queryset.filter(combined_query)


# 2. GURUH FILTERI (YANGI)
class GroupFilter(admin.SimpleListFilter):
    title = "Guruh"
    parameter_name = 'group_filter'

    def lookups(self, request, model_admin):
        return [(g.id, g.name) for g in Group.objects.all()]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(group__id__in=self.value().split(','))
        return queryset


# 3. TA'LIM SHAKLI FILTERI (YANGI)
class EducationFormFilter(admin.SimpleListFilter):
    title = "Ta'lim shakli"
    parameter_name = 'education_form_filter'

    def lookups(self, request, model_admin):
        return Student.EducationFormChoices.choices

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(education_form__in=self.value().split(','))
        return queryset


# 4. KURS FILTERI (YANGI)
class CourseFilter(admin.SimpleListFilter):
    title = "Kurs"
    parameter_name = 'course_filter'

    def lookups(self, request, model_admin):
        return [(i, f"{i}-kurs") for i in range(1, 6)]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(course_year__in=self.value().split(','))
        return queryset

@admin.register(Student)
class StudentAdmin(admin.ModelAdmin):
    change_list_template = "admin/students/student/change_list.html"

    # ---------------------------------------------------------
    # 1. SHABLON VA DISPLAY SOZLAMALARI
    # ---------------------------------------------------------

    list_display = (
        'student_hemis_id',
        'full_name',
        # 'phone_number', # Kerak bo'lsa yoqing
        'education_form',
        'get_course_year',
        'get_group_name',

        # --- KONTRAKT USTUNLARI ---
        'get_contract_amount',
        'get_total_payment',
        'get_paid_percent',
        'get_payment_debt',

        # --- FAN QARZDORLIGI USTUNLARI ---
        'get_subject_debt_amount',  # Jami hisoblangan qarz
        'get_subject_debt_paid',  # To'langan qismi
        'get_subject_debt_diff',  # Qoldiq qarz
        'get_open_debt_count',
        'view_student_link',
    )

    list_filter = (
        PaymentPercentFilter,
        GroupFilter,
        EducationFormFilter,
        CourseFilter,
        AcademicYearFilter,
        StatusFilter,
    )

    search_fields = (
        'full_name',
        'student_hemis_id',
        'passport_series_number',
        'personal_pin',
        'group__name',  # Guruh nomi bo'yicha ham qidirish (juda qulay)
        'phone_number',  # Telefon raqam bo'yicha ham
    )


    raw_id_fields=['group','previous_education_country', 'citizenship',
        'region', 'district', 'previous_education_region']

    inlines = [ContractInline, OrderInline, SubjectDebtInline, StudentHistoryInline]

    fieldsets = (
        ('Asosiy ma\'lumotlar', {
            'fields': ('full_name', 'student_hemis_id', 'education_type', 'group', 'phone_number', 'phone_number_2')
        }),
        ('Pasport ma\'lumotlari', {
            'classes': ('collapse',),
            'fields': (
                'passport_series_number', 'personal_pin', 'passport_issued_by',
                'passport_issue_date', 'passport_expiry_date', 'date_of_birth',
                'birth_place', 'gender', 'nationality', 'citizenship', 'region', 'district', 'address'
            )
        }),
        ("O'qishga oid ma'lumotlar", {
            'fields': ('status', 'education_form', 'course_year', 'current_semester', 'entry_score', 'document',
                       'document_info')
        }),
        ("Oldingi ta'lim", {
            'classes': ('collapse',),
            'fields': (
                'previous_education_country', 'previous_education_region', 'previous_institution',
                'document_type', 'document_number', 'previous_graduation_year',
                'certificate_info', 'transferred_from_university'
            )
        }),
    )

    list_select_related = ('group__specialty', 'region', 'district')
    list_per_page = 50  # Sahifada kamroq bo'lsa tezroq ishlaydi
    show_full_result_count = False
    actions = None

    class Media:
        css = {
            'all': ('admin/css/custom_scroll.css',)
        }
        js = (
            'admin/js/jquery.init.js',
            'admin/js/money_input.js',
        )

    # ---------------------------------------------------------
    # 2. QUERYSET LOGIKASI (YIL VA HISOB-KITOB)
    # ---------------------------------------------------------
    def get_queryset(self, request):
        queryset = super().get_queryset(request)

        # 1. Tanlangan o'quv yilini aniqlash
        selected_year_id = request.GET.get('academic_year_calc')
        if selected_year_id:
            current_year = AcademicYear.objects.filter(id=selected_year_id).first()
        else:
            current_year = self.get_current_year()

        if not current_year:
            return queryset

        # --- A) KONTRAKT SUMMASI (GROSS - YALPI) ---
        # Bu yerda hech narsa ayirmaymiz, shunchaki shartnoma summasi
        contract_subquery = Contract.objects.filter(
            student=OuterRef('pk'),
            academic_year=current_year,
            contract_type='contract'
        ).values('student').annotate(
            total_amount=Sum('amount')
        ).values('total_amount')

        # --- B) GRANT SUMMASI ---
        grant_subquery = Contract.objects.filter(
            student=OuterRef('pk'),
            academic_year=current_year,
            contract_type='contract'
        ).values('student').annotate(
            total_grant=Sum('grant_amount')
        ).values('total_grant')

        # --- C) TO'LOV SUMMASI ---
        payment_subquery = Payment.objects.filter(
            contract__student=OuterRef('pk'),
            contract__academic_year=current_year,
            contract__contract_type='contract'
        ).values('contract__student').annotate(
            total=Sum('amount')
        ).values('total')

        # --- D) FAN QARZLARI ---
        subject_debt_total_subquery = SubjectDebt.objects.filter(
            student=OuterRef('pk')
        ).values('student').annotate(
            total=Sum('amount')
        ).values('total')

        subject_debt_paid_subquery = SubjectDebt.objects.filter(
            student=OuterRef('pk')
        ).values('student').annotate(
            total=Sum('amount_summ')
        ).values('total')

        open_debt_count_subquery = SubjectDebt.objects.filter(
            student=OuterRef('pk'),
            status='yopilmadi'
        ).values('student').annotate(
            cnt=Count('id')
        ).values('cnt')

        # 1-BOSQICH: MA'LUMOTLARNI OLIB KELISH
        queryset = queryset.annotate(
            gross_contract_amount=Coalesce(Subquery(contract_subquery), Value(Decimal(0))),  # <--- JAMI KONTRAKT UCHUN
            current_grant_amount=Coalesce(Subquery(grant_subquery), Value(Decimal(0))),  # Grant
            total_paid_amount=Coalesce(Subquery(payment_subquery[:1]), Value(Decimal(0))),

            subject_debt_total=Coalesce(Subquery(subject_debt_total_subquery), Value(Decimal(0))),
            subject_debt_paid=Coalesce(Subquery(subject_debt_paid_subquery), Value(Decimal(0))),
            open_debt_count=Coalesce(Subquery(open_debt_count_subquery), Value(0)),
        )

        # 2-BOSQICH: REAL KONTRAKTNI HISOBLASH (Jadval ustunlari uchun)
        # Jadvalda "Kontrakt" ustuni Real (Net) summani ko'rsatishi kerak
        queryset = queryset.annotate(
            current_contract_amount=ExpressionWrapper(
                F('gross_contract_amount') - F('current_grant_amount'),
                output_field=DecimalField()
            )
        )

        # 3-BOSQICH: QARZ VA FOIZLAR (Real kontraktga asoslanadi)
        queryset = queryset.annotate(
            payment_percent=Case(
                When(current_contract_amount__lte=0, then=Value(Decimal(0))),
                default=(F('total_paid_amount') * 100) / F('current_contract_amount'),
                output_field=DecimalField()
            ),
            # Qarzdorlik = Real Kontrakt - To'lov
            payment_diff=F('current_contract_amount') - F('total_paid_amount'),

            subject_debt_diff=F('subject_debt_total') - F('subject_debt_paid'),
        )

        return queryset

    # ---------------------------------------------------------
    # 3. STATISTIKA PANELINI CHIQARISH (YANGILANDI)
    # ---------------------------------------------------------
    def changelist_view(self, request, extra_context=None):
        if request.method == 'GET':
            q = request.GET.copy()
            if 'academic_year_calc' not in q:
                active_year = AcademicYear.objects.filter(is_active=True).first()
                if active_year:
                    q['academic_year_calc'] = str(active_year.id)
                    request.GET = q
                    request.META['QUERY_STRING'] = q.urlencode()

        groups_data = list(Group.objects.values('id', 'name').order_by('name'))
        extra_context = extra_context or {}
        extra_context['groups_json'] = json.dumps(groups_data)

        response = super().changelist_view(request, extra_context)

        if not hasattr(response, 'context_data') or 'cl' not in response.context_data:
            return response

        cl = response.context_data['cl']
        qs = cl.queryset

        # 4. AGGREGATSIYA (HISOBLASH)
        metrics = qs.aggregate(
            # A) JAMI KONTRAKT (Grant ayrilmagan - YALPI)
            jami_gross_shartnoma=Sum('gross_contract_amount'),

            # B) REAL KONTRAKT (Grant ayrilgan - SOF)
            jami_real_shartnoma=Sum('current_contract_amount'),

            jami_grant=Sum('current_grant_amount'),
            jami_tolov=Sum('total_paid_amount'),

            # Haqiqiy umumiy qarz (Real Shartnoma - To'lov)
            haqiqiy_qarz=Sum(
                Case(
                    When(
                        current_contract_amount__gt=F('total_paid_amount'),
                        then=F('current_contract_amount') - F('total_paid_amount')
                    ),
                    default=Value(Decimal('0')),
                    output_field=DecimalField()
                )
            ),
        )

        val_gross = metrics['jami_gross_shartnoma'] or 0
        val_real = metrics['jami_real_shartnoma'] or 0
        val_grant = metrics['jami_grant'] or 0
        val_paid = metrics['jami_tolov'] or 0
        val_debt = metrics['haqiqiy_qarz'] or 0

        response.context_data['footer_stats'] = {
            'contract': val_gross,  # <--- 1-BOX: Grant ayrilmagan summa (Jami Kontrakt)
            'real_contract': val_real,  # <--- 2-BOX: Grant ayrilgan summa (Real Kontrakt)
            'grant': val_grant,  # 3-BOX: Grant
            'paid': val_paid,  # 4-BOX: To'langan
            'debt': val_debt,  # 5-BOX: Qarz (Realga nisbatan)
        }

        return response

    # ---------------------------------------------------------
    # 4. DISPLAY METODLARI (CHIROYLI DIZAYN BILAN)
    # ---------------------------------------------------------

    def get_current_year(self):
        return AcademicYear.objects.filter(is_active=True).first()

    @admin.display(description="Kurs", ordering='course_year')
    def get_course_year(self, obj):
        return obj.course_year if obj.group else "-"

    @admin.display(description="Guruh", ordering='group__name')
    def get_group_name(self, obj):
        if obj.group:
            # Guruh nomini chiroyli kulrang fonda chiqarish
            return format_html(
                '<span class="group-badge">{}</span>',
                obj.group.name
            )
        return "-"

    # --- KONTRAKT DISPLAY ---
    @admin.display(description="kontrakt", ordering='current_contract_amount')
    def get_contract_amount(self, obj):
        if hasattr(obj, 'current_contract_amount') and obj.current_contract_amount > 0:
            return f"{obj.current_contract_amount:,.0f}".replace(",", " ")
        return "-"

    @admin.display(description="To‘lov", ordering='total_paid_amount')
    def get_total_payment(self, obj):
        if hasattr(obj, 'total_paid_amount') and obj.total_paid_amount > 0:
            return f"{obj.total_paid_amount:,.0f}".replace(",", " ")
        return "-"

    @admin.display(description="Foizi", ordering='payment_percent')  # <-- O'ZGARTIRILDI
    def get_paid_percent(self, obj):
        if not hasattr(obj, 'current_contract_amount') or obj.current_contract_amount == 0:
            return format_html('<span class="status-badge badge-secondary">-</span>')

        percent = obj.payment_percent.quantize(Decimal('1.'), rounding=ROUND_HALF_UP)

        # Rang logikasi
        if percent >= 100:
            css_class = 'badge-success'
            icon = '✓'
        elif percent >= 50:
            css_class = 'badge-info'
            icon = ''
        elif percent > 0:
            css_class = 'badge-warning'
            icon = ''
        else:
            css_class = 'badge-danger'
            icon = '!'

        return format_html(
            '<div class="status-badge {}">{} {}%</div>',
            css_class, icon, percent
        )

    @admin.display(description="Qarzi", ordering='payment_diff')
    def get_payment_debt(self, obj):
        if not hasattr(obj, 'current_contract_amount') or obj.current_contract_amount == 0:
            return "-"

        debt = obj.payment_diff

        if debt <= 0:
            # Qarzi yo'q bo'lsa
            return format_html('<span style="color: #20c997; font-weight: bold; font-size: 16px;">✓</span>')

        formatted = f"{debt:,.0f}".replace(",", " ")
        # Qarzi bor bo'lsa, qizil badge
        return format_html(
            '<span class="status-badge badge-danger">{}</span>',
            formatted
        )

    # --- FAN QARZDORLIGI DISPLAY (YANGI) ---
    @admin.display(description="K-M shartnoma", ordering='subject_debt_total')
    def get_subject_debt_amount(self, obj):
        if hasattr(obj, 'subject_debt_total') and obj.subject_debt_total > 0:
            return f"{obj.subject_debt_total:,.0f}".replace(",", " ")
        return "-"

    @admin.display(description="K-M To'lov", ordering='subject_debt_paid')
    def get_subject_debt_paid(self, obj):
        if hasattr(obj, 'subject_debt_paid') and obj.subject_debt_paid > 0:
            return f"{obj.subject_debt_paid:,.0f}".replace(",", " ")
        return "-"

    @admin.display(description="K-M Qarz", ordering='subject_debt_diff')
    def get_subject_debt_diff(self, obj):
        if not hasattr(obj, 'subject_debt_total') or obj.subject_debt_total == 0:
            return "-"

        debt = obj.subject_debt_diff

        if debt <= 0:
            return format_html('<span style="color: #20c997; font-weight: bold;">Yopilgan</span>')

        formatted = f"{debt:,.0f}".replace(",", " ")
        # Fan qarzi uchun alohida ogohlantirish stili
        return format_html(
            '<span class="status-badge badge-danger" style="border: 1px dashed #fa5252;">{}</span>',
            formatted
        )

    @admin.display(description="Qarz fanlar", ordering='open_debt_count')
    def get_open_debt_count(self, obj):
        # Annotatsiyadan kelgan qiymatni olamiz
        count = getattr(obj, 'open_debt_count', 0)

        if count == 0:
            # Agar qarzdorlik soni 0 bo'lsa, yashil belgi yoki chiziqcha
            return format_html('<span style="color: #20c997; font-weight: bold;">-</span>')

        # Agar qarzi bor bo'lsa, qizil badgeda sonini chiqaramiz
        return format_html(
            '<span class="status-badge badge-danger" style="font-size: 13px; padding: 4px 8px;">{} ta</span>',
            count
        )

    # ---------------------------------------------------------
    # 5. QO'SHIMCHA URL VA VIEWLAR
    # ---------------------------------------------------------

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                '<path:object_id>/detail/',
                self.admin_site.admin_view(self.student_detail_view),
                name='students_student_detail'
            ),
            path('export-excel/', self.admin_site.admin_view(self.export_excel_view), name='students_student_export'),
        ]
        return custom_urls + urls

    def export_excel_view(self, request):
        """
        Talabalar ro'yxatini Excelga export qilish.
        Default holatda eski ustunlar chiqadi.
        Contract raqami va sanasi faqat tanlansa chiqadi.
        """
        # 1. Filtrlangan ma'lumotlarni olish
        try:
            cl = self.get_changelist_instance(request)
            queryset = cl.get_queryset(request)
        except AttributeError:
            queryset = self.filter_queryset(self.get_queryset(request))

        # 2. Formadan tanlangan ustunlarni olish
        selected_fields = request.POST.getlist('selected_fields')

        # --- O'ZGARISH SHU YERDA: DEFAULT RO'YXAT ---
        # Agar hech narsa tanlanmagan bo'lsa, standart ustunlar chiqadi
        # LEKIN contract_number va contract_date BU YERDA YO'Q
        if not selected_fields:
            selected_fields = [
                'full_name',
                'student_hemis_id',
                'group',
                'education_form',
                'total_paid_amount',
                'payment_diff',
                'qabul_order_number',
                'qabul_order_date'
            ]
        # ---------------------------------------------

        # 3. Excel fayl yaratish
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Talabalar Export"

        # Dizayn
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        money_format = '#,##0'

        # 4. Header nomlari
        field_titles = {
            'student_hemis_id': 'ID (Hemis)',
            'full_name': 'F.I.SH.',
            'group': 'Guruh',
            'course_year': 'Kurs',
            'education_form': "Ta'lim shakli",
            'education_type': "Ta'lim turi",
            'payment_type': "To'lov turi",

            # --- YANGI MAYDONLAR ---
            'contract_number': 'Shartnoma raqami',
            'contract_date': 'Shartnoma sanasi',
            # -----------------------

            'current_contract_amount': 'Hisoblangan kontrakt',
            'total_paid_amount': "To'langan summa",
            'payment_diff': 'Qarzdorlik',
            'payment_percent': 'Foiz (%)',

            'passport_serial': 'Pasport Seriya',
            'passport_number': 'Pasport Raqami',
            'pinfl': 'JSHSHIR',
            'gender': 'Jinsi',
            'birth_date': "Tug'ilgan sana",
            'region': 'Viloyat',
            'district': 'Tuman',
            'address': 'Manzil',
            'phone_number': 'Telefon',

            'previous_institution': 'Oldingi muassasa',
            'document_type': 'Hujjat turi',
            'document_number': 'Hujjat raqami',
            'transferred_from_university': "Ko'chirgan OTM",

            'qabul_order_number': 'Qabul buyruq №',
            'qabul_order_date': 'Qabul buyruq sanasi',
        }

        # Headerlarni chizish
        for col_num, field in enumerate(selected_fields, 1):
            column_letter = get_column_letter(col_num)
            cell = ws.cell(row=1, column=col_num)
            cell.value = field_titles.get(field, field)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

            if field == 'full_name':
                ws.column_dimensions[column_letter].width = 35
            elif field in ['contract_number', 'contract_date']:
                ws.column_dimensions[column_letter].width = 20
            else:
                ws.column_dimensions[column_letter].width = 15

        # 5. Ma'lumotlarni yozish
        money_fields = ['current_contract_amount', 'total_paid_amount', 'payment_diff',
                        'subject_debt_total', 'subject_debt_paid', 'subject_debt_diff']

        row_num = 2
        active_year = AcademicYear.objects.filter(is_active=True).first()

        for obj in queryset:

            # --- SHARTNOMANI OLISH LOGIKASI ---
            active_contract = None
            need_contract = ('contract_number' in selected_fields or 'contract_date' in selected_fields)

            if need_contract and active_year:
                active_contract = obj.contract_set.filter(
                    academic_year=active_year,
                    contract_type='contract'
                ).first()

            # Qabul buyrug'ini olish (eski logikangiz bo'yicha)
            qabul_order = None
            if 'qabul_order_number' in selected_fields or 'qabul_order_date' in selected_fields:
                qabul_order = obj.order_set.filter(
                    order_type__name__icontains='qabul',
                    is_deleted=False
                ).order_by('-order_date').first()
            # ----------------------------------

            for col_num, field in enumerate(selected_fields, 1):
                cell = ws.cell(row=row_num, column=col_num)
                val = None

                # --- 1. Shartnoma ma'lumotlari ---
                if field == 'contract_number':
                    val = active_contract.contract_number if active_contract else ""
                    cell.alignment = center_align

                elif field == 'contract_date':
                    if active_contract and active_contract.contract_date:
                        val = active_contract.contract_date.strftime('%d.%m.%Y')
                    else:
                        val = ""
                    cell.alignment = center_align

                # --- 2. Qabul buyrug'i ---
                elif field == 'qabul_order_number':
                    val = qabul_order.order_number if qabul_order else ""
                    cell.alignment = center_align

                elif field == 'qabul_order_date':
                    if qabul_order and qabul_order.order_date:
                        val = qabul_order.order_date.strftime('%d.%m.%Y')
                    else:
                        val = ""
                    cell.alignment = center_align

                # --- 3. Obyektning o'z maydonlari ---
                elif hasattr(obj, field):
                    val = getattr(obj, field)
                    if isinstance(val, (datetime, date)):
                        val = val.strftime('%d.%m.%Y')
                    elif callable(val):
                        val = val()

                # --- 4. Bog'langan maydonlar ---
                elif field == 'group':
                    val = str(obj.group.name) if obj.group else ""
                elif field == 'course_year':
                    val = str(obj.course_year) if obj.course_year else ""
                elif field == 'region':
                    val = obj.region.name if obj.region else ""
                elif field == 'district':
                    val = obj.district.name if obj.district else ""

                # --- Qiymatni yozish ---
                if val is None:
                    val = ""

                if field in money_fields:
                    try:
                        cell.value = float(val) if val else 0
                        cell.number_format = money_format
                    except (ValueError, TypeError):
                        cell.value = val
                else:
                    cell.value = str(val)

                cell.border = thin_border

                if field not in ['full_name', 'address'] and field not in money_fields:
                    cell.alignment = center_align
                elif field == 'full_name':
                    cell.alignment = left_align

            row_num += 1

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        response['Content-Disposition'] = f'attachment; filename=Students_Export_{timestamp}.xlsx'
        wb.save(response)
        return response

    # D) ChangeList yordamchisi (Filter ishlashi uchun kerak)
    def get_changelist_instance(self, request):
        list_display = self.get_list_display(request)
        list_display_links = self.get_list_display_links(request, list_display)
        list_filter = self.get_list_filter(request)
        search_fields = self.get_search_fields(request)
        list_select_related = self.get_list_select_related(request)

        try:
            actions = self.get_actions(request)
            if actions:
                list_display = ['action_checkbox'] + list(list_display)
        except (AttributeError, KeyError):
            pass

        ChangeListClass = self.get_changelist(request)

        return ChangeListClass(
            request,
            self.model,
            list_display,
            list_display_links,
            list_filter,
            self.date_hierarchy,
            search_fields,
            list_select_related,
            self.list_per_page,
            self.list_max_show_all,
            self.list_editable,
            self,
            self.sortable_by,
            self.search_help_text,
        )

    @admin.display(description="Ko'rish")
    def view_student_link(self, obj):
        url = reverse('admin:students_student_detail', args=[obj.pk])
        # Ko'zchani chiroyli qilish
        return format_html(
            '<a href="{}" title="To\'liq ma\'lumot" class="view-link" style="color: #adb5bd; font-size: 1.2rem; transition: 0.2s;">'
            '<i style="color:blue" class="fas fa-eye"></i>'
            '</a>',
            url
        )

    def student_detail_view(self, request, object_id):
        student = self.get_object(request, object_id)
        if student is None:
            return self._get_obj_does_not_exist_redirect(request, self.model._meta, object_id)

        contracts = student.contract_set.all().order_by('-academic_year__name')
        payments = Payment.objects.filter(contract__student=student).order_by('-payment_date')
        orders = student.orders.all().order_by('-order_date')

        # --- O'ZGARTIRILGAN QISM (ANNOTATE QO'SHILDI) ---
        debts = student.subjectdebt_set.annotate(
            paid_amount=Coalesce('amount_summ', Value(Decimal('0'))),
            remaining_debt=F('amount') - Coalesce('amount_summ', Value(Decimal('0')))
        ).order_by('-academic_year__name', 'subject__name')
        # ------------------------------------------------

        context = {
            # ... (eski context qismi o'zgarishsiz)
            'title': f"{student.full_name} ma'lumotlari",
            'student': student,
            'contracts': contracts,
            'payments': payments,
            'orders': orders,
            'debts': debts,
            'opts': self.model._meta,
            'has_change_permission': self.has_change_permission(request, student),
        }
        return render(request, 'admin/students/student/student_detail.html', context)

    def formfield_for_dbfield(self, db_field, request, **kwargs):
        if db_field.name == 'birth_place':
            existing_places = Student.objects.exclude(birth_place__isnull=True).exclude(
                birth_place__exact='').values_list('birth_place', flat=True).distinct()
            kwargs['widget'] = DatalistTextInput(datalist=existing_places)
        elif db_field.name == 'nationality':
            existing_nationalities = Student.objects.exclude(nationality__isnull=True).exclude(
                nationality__exact='').values_list('nationality', flat=True).distinct()
            kwargs['widget'] = DatalistTextInput(datalist=existing_nationalities)
        return super().formfield_for_dbfield(db_field, request, **kwargs)


# =============================================================================
# 💰 MOLIYA MODELLARI
# (O'zgarishsiz qoldirildi)
# =============================================================================
# admin.py faylida PaymentResource klassini topib, shunday yangilang:

class PaymentResource(resources.ModelResource):
    # 1. Shartnoma ID sini bog'lash (Eski kod)
    contract = fields.Field(
        column_name='contract_id',
        attribute='contract',
        widget=ForeignKeyWidget(Contract, field='id')
    )

    # 2. Asosiy maydonlar (Eski kod)
    amount = fields.Field(column_name='amount', attribute='amount')
    payment_date = fields.Field(column_name='payment_date', attribute='payment_date')

    # --- YANGI QO'SHILGAN MAYDONLAR (EXPORT UCHUN) ---

    # 3. O'quv yili
    academic_year = fields.Field(
        column_name="O'quv yili",
        attribute='contract__academic_year__name',
        readonly=True
    )

    # 4. Kursi
    student_course = fields.Field(
        column_name="Kurs",
        attribute='contract__student__course_year',
        readonly=True
    )

    # 5. Status (Display - chiroyli ko'rinishda olish uchun dehydrate ishlatamiz)
    student_status = fields.Field(
        column_name="Status",
        readonly=True
    )

    # 6. Yo'nalish (Guruh orqali olinadi)
    specialty = fields.Field(
        column_name="Yo'nalish",
        readonly=True
    )

    class Meta:
        model = Payment
        # Export qilinadigan barcha ustunlar ro'yxati
        fields = (
            'id',
            'contract',
            'academic_year',  # <--- Qo'shildi
            'specialty',  # <--- Qo'shildi
            'student_course',  # <--- Qo'shildi
            'student_status',  # <--- Qo'shildi
            'amount',
            'payment_date',
            'description'
        )

        # Exceldagi ustunlar ketma-ketligi
        export_order = (
            'id',
            'contract',
            'academic_year',
            'specialty',
            'student_course',
            'student_status',
            'amount',
            'payment_date',
            'description'
        )

    # --- YANGI MAYDONLARNI HISOBLASH METODLARI ---

    def dehydrate_student_status(self, payment):
        """Talabaning statusini (active, expelled emas, "O'qiydi" deb) chiqarish"""
        if payment.contract and payment.contract.student:
            return payment.contract.student.get_status_display()
        return ""

    def dehydrate_specialty(self, payment):
        """Talabaning yo'nalishini guruh orqali topish"""
        try:
            return payment.contract.student.group.specialty.name
        except AttributeError:
            return ""

    def skip_row(self, instance, original, row, import_validation_errors=None):
        """
        Bo'sh qatorlarni yoki ma'lumoti chala qatorlarni tashlab ketish.
        """
        if not row.get('amount'):
            return True
        if not row.get('contract_id'):
            return True
        return super().skip_row(instance, original, row, import_validation_errors)


@admin.register(Payment)
class PaymentAdmin(ImportExportModelAdmin):
    resource_class = PaymentResource
    form = PaymentForm
    list_display = ('contract', 'payment_date', 'amount')
    list_filter = ('payment_date',)
    search_fields = ('contract__student__full_name', 'contract__contract_number','id','payment_date')
    autocomplete_fields = ['contract']

    class Media:
        js = (
            'admin/js/jquery.init.js',
            'admin/js/money_input.js',
            'admin/js/payment_contract_info.js',  # <--- YANGI JS FAYL
        )
        css = {
            'all': ('admin/css/payment_info.css',)  # Ixtiyoriy: chiroyli ko'rinish uchun
        }

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('get-contract-info/', self.admin_site.admin_view(self.get_contract_info_view),
                 name='payment_get_contract_info'),
        ]
        return custom_urls + urls



    def get_contract_info_view(self, request):
        contract_id = request.GET.get('contract_id')
        if not contract_id:
            return JsonResponse({'error': 'ID topilmadi'}, status=400)

        try:
            contract = Contract.objects.get(id=contract_id)

            # 1. Hisob-kitoblar
            total_amount = contract.amount
            paid_amount = Payment.objects.filter(contract=contract).aggregate(sum=Sum('amount'))['sum'] or 0
            debt = total_amount - paid_amount


            return JsonResponse({
                'contract_amount': float(total_amount),
                'paid_amount': float(paid_amount),
                'debt': float(debt),
                'student_name': contract.student.full_name
            })
        except Contract.DoesNotExist:
            print(f"❌ Terminal: {contract_id} IDli shartnoma topilmadi!")
            return JsonResponse({'error': 'Shartnoma topilmadi'}, status=404)


class CustomForeignKeyWidget(ForeignKeyWidget):
    def clean(self, value, row=None, **kwargs):
        if not value:
            return None
        val = str(value).strip()
        try:
            return super().clean(val, row, **kwargs)
        except ObjectDoesNotExist:
            raise ValueError(f"Bazada '{val}' IDli {self.model._meta.verbose_name} topilmadi!")


class GrantTypeWidget(Widget):
    def clean(self, value, row=None, *args, **kwargs):
        """
        Exceldagi matnni Contract.GrantTypeChoices kodiga o'giradi.
        """
        if not value:
            return Contract.GrantTypeChoices.NONE

        # 1. Matnni tozalash (probellarni olib tashlash va kichik harfga o'tkazish)
        val = str(value).strip().lower()

        # Qo'shtirnoqlarni standartlashtirish (Word/Excelda har xil bo'lishi mumkin)
        val = val.replace("‘", "'").replace("’", "'").replace("`", "'").replace('"', '')

        # 2. Bazadagi variantlarni tekshirish
        for code, label in Contract.GrantTypeChoices.choices:
            # Labelni ham tozalab olamiz
            clean_label = label.lower().replace("‘", "'").replace("’", "'").replace("`", "'").replace('"', '')
            clean_code = code.lower()

            # A) Agar kodning o'zi yozilgan bo'lsa (masalan: "CR")
            if clean_code == val:
                return code

            # B) Agar to'liq nomi yozilgan bo'lsa
            if clean_label == val:
                return code

            # C) Qisman moslik (Eng muhim qismi):
            # Agar kod Exceldagi matn ichida bo'lsa (va kod 'none' bo'lmasa)
            # Masalan: "Rag'batlantirish-CR" ichida "cr" bor.
            if clean_code != 'none' and clean_code in val:
                return code

            # D) Agar nomining asosiy qismi mos kelsa
            if len(val) > 3 and val in clean_label:
                return code

        # Agar hech narsa topilmasa
        return Contract.GrantTypeChoices.NONE


# =========================================================
# 2. RESOURCE (Import/Export sozlamalari)
# =========================================================
class ContractResource(resources.ModelResource):
    # --- 1. ID va Bog'lanishlar ---
    id = fields.Field(column_name='id', attribute='id')

    student = fields.Field(
        column_name='student',
        attribute='student',
        widget=CustomForeignKeyWidget(Student, 'student_hemis_id')
    )

    academic_year = fields.Field(
        column_name='academic_year',
        attribute='academic_year',
        widget=CustomForeignKeyWidget(AcademicYear, 'name')
    )

    # --- 2. Asosiy summalar ---
    amount = fields.Field(
        column_name='amount',
        attribute='amount',
        widget=NumberWidget()
    )

    contract_date = fields.Field(
        column_name='contract_date',
        attribute='contract_date',
        widget=DateWidget(format='%d.%m.%Y')
    )

    # --- 3. Grant turlari (Import uchun) ---
    grant_type = fields.Field(
        column_name='Grant turi',
        attribute='grant_type',
        widget=GrantTypeWidget()
    )

    # --- 4. YANGI QO'SHILGAN MAYDONLAR (EXPORT UCHUN) ---

    # Kurs
    student_course = fields.Field(
        column_name='Kurs',
        attribute='student__course_year',
        readonly=True
    )

    # Ta'lim shakli (Chiroyli nomini oladi)
    student_education_form = fields.Field(
        column_name="Ta'lim shakli",
        attribute='student__get_education_form_display',
        readonly=True
    )

    # Yo'nalish (Dehydrate orqali olinadi)
    student_specialty = fields.Field(
        column_name="Yo'nalish",
        readonly=True
    )

    # Grant turi (Display - Chiroyli nomi)
    grant_type_display = fields.Field(
        column_name='Grant turi (Display)',
        attribute='get_grant_type_display',
        readonly=True
    )

    grant_date = fields.Field(
        column_name='Grant sanasi',
        attribute='grant_date',
        widget=DateWidget(format='%d.%m.%Y')
    )

    grant_percent = fields.Field(
        column_name='Grant foizi',
        attribute='grant_percent',
        widget=NumberWidget()
    )

    grant_amount = fields.Field(
        column_name='Grant summasi',
        attribute='grant_amount',
        widget=NumberWidget()
    )

    class Meta:
        model = Contract

        # --- MUHIM O'ZGARISH SHU YERDA ---
        # "fields" ro'yxatiga barcha yangi fieldlarni qo'shish SHART!
        # Aks holda "UserWarning: cannot identify field" xatosi chiqadi.
        fields = (
            'id',
            'student',
            'student_specialty',  # <--- Qo'shildi
            'student_education_form',  # <--- Qo'shildi
            'student_course',  # <--- Qo'shildi
            'academic_year',
            'contract_number',
            'contract_date',
            'amount',
            'grant_type',
            'grant_type_display',  # <--- Qo'shildi
            'grant_date',
            'grant_percent',
            'grant_amount',
        )

        # Export qilish ketma-ketligi
        export_order = (
            'id',
            'student',
            'student_specialty',
            'student_education_form',
            'student_course',
            'academic_year',
            'contract_number',
            'contract_date',
            'amount',
            'grant_type',
            'grant_type_display',
            'grant_date',
            'grant_percent',
            'grant_amount'
        )

        import_id_fields = ('id',)
        raise_errors = False
        skip_unchanged = True
        report_skipped = True

    # Yo'nalish nomini olish uchun metod (Xatolik chiqmasligi uchun)
    def dehydrate_student_specialty(self, contract):
        try:
            if contract.student and contract.student.group and contract.student.group.specialty:
                return contract.student.group.specialty.name
        except Exception:
            return ""
        return ""

    def get_instance(self, instance_loader, row):
        row_id = row.get('id')
        if row_id:
            try:
                return Contract.objects.get(id=row_id)
            except Contract.DoesNotExist:
                pass
        return None


# =========================================================
# 3. ADMIN REGISTRATION
# =========================================================
@admin.register(Contract)
class ContractAdmin(ImportExportModelAdmin):
    resource_class = ContractResource
    form = ContractForm  # Agar sizda ContractForm bo'lsa

    list_display = (
        'student', 'id', 'academic_year', 'contract_number', 'contract_date', 'amount',
        'grant_type', 'grant_date', 'grant_percent',
    )
    list_filter = ('academic_year', 'grant_type')

    search_fields = (

        'contract_number',
        'student__full_name',
        'student__student_hemis_id',
        'student__passport_series_number',
        'student__personal_pin',
    )

    raw_id_fields = ['student']
    list_per_page = 500

    class Media:
        js = (
            'admin/js/jquery.init.js',
            'admin/js/money_input.js',
            'admin/js/contract_grant_standalone.js',
        )


# =============================================================================
# 📚 QO‘SHIMCHA MODELLAR
# (O'zgarishsiz qoldirildi)
# =============================================================================
class OrderTypeResource(resources.ModelResource):
    class Meta:
        model = OrderType


@admin.register(OrderType)
class OrderTypeAdmin(ImportExportModelAdmin):
    resource_class = OrderTypeResource
    list_display = ('name', 'id')
    search_fields = ('name',)


class MultiFormatDateWidget(DateWidget):
    def __init__(self, input_formats=None, render_format='%Y-%m-%d'):
        super().__init__(format=render_format)
        self.input_formats = input_formats or [
            '%d.%m.%Y',
            '%Y-%m-%d',
            '%Y-%m-%d %H:%M:%S',
        ]

    def clean(self, value, row=None, **kwargs):
        if value in (None, '', 'None'):
            return None
        v = str(value).strip()
        for fmt in self.input_formats:
            try:
                return datetime.strptime(v, fmt).date()
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(v.replace('Z', '')).date()
        except Exception:
            raise ValueError(f"Sana formati noto‘g‘ri: {v}")


# admin.py faylida OrderResource klassini to'liq shunday o'zgartiring:

class OrderResource(resources.ModelResource):
    # 1. Talabani HEMIS ID orqali topish
    student = fields.Field(
        column_name='student',
        attribute='student',
        widget=ForeignKeyWidget(Student, 'student_hemis_id')
    )

    # 2. Buyruq turini ID orqali topish (masalan, 1 - Qabul buyrug'i)
    order_type = fields.Field(
        column_name='order_type',
        attribute='order_type',
        widget=ForeignKeyWidget(OrderType, 'id')
    )

    # 3. Sanani to'g'ri o'qish (16.10.2025 formati uchun)
    order_date = fields.Field(
        column_name='order_date',
        attribute='order_date',
        widget=MultiFormatDateWidget()
    )

    # 4. Boshqa sanalar (agar Excelda bo'lsa, bo'lmasa shart emas)
    application_date = fields.Field(
        column_name='application_date',
        attribute='application_date',
        widget=MultiFormatDateWidget()
    )
    document_taken_date = fields.Field(
        column_name='document_taken_date',
        attribute='document_taken_date',
        widget=MultiFormatDateWidget()
    )

    class Meta:
        model = Order

        # --- ENG MUHIM QISM SHU YERDA ---
        # Tizim qatorni yangilash yoki yaratishni shu ikki maydon orqali aniqlaydi.
        # Ma'nosi: "Agar shu TALABAda shu TURDAGI buyruq (masalan, 1-turi) allaqachon bo'lsa -> UPDATE qil".
        # Agar bu ikkisi topilmasa -> CREATE qil.
        import_id_fields = ('student', 'order_type')

        fields = (
            'student',
            'order_type',
            'order_number',
            'order_date',
            'application_date',
            'document_taken_date',
            'notes',
            'tsch_reason'
        )

        skip_unchanged = True
        report_skipped = True


@admin.register(Order)
class OrderAdmin(ImportExportMixin, admin.ModelAdmin):
    resource_class = OrderResource
    list_display = ('order_date', 'order_type', 'order_number', 'student')
    list_filter = ('order_type', 'order_date')
    date_hierarchy = 'order_date'
    ordering = ('-order_date',)
    list_select_related = ('student', 'order_type')
    autocomplete_fields = ('student', 'order_type')
    search_fields = (
        'order_number',
        'student__full_name',
        'student__student_hemis_id',
    )



class AcademicYearResource(resources.ModelResource):
    class Meta:
        model = AcademicYear


@admin.register(AcademicYear)
class AcademicYearAdmin(ImportExportModelAdmin):
    resource_class = AcademicYearResource
    list_display = ('name', 'is_active')
    list_filter = ('is_active',)
    search_fields = ('name',)


class SubjectResource(resources.ModelResource):
    class Meta:
        model = Subject


@admin.register(Subject)
class SubjectAdmin(ImportExportModelAdmin):
    resource_class = SubjectResource
    list_display = ('name',)
    search_fields = ('name',)



class SubjectRateResource(resources.ModelResource):
    class Meta:
        model = SubjectRate
        # education_form ni import/exportga qo'shamiz
        fields = ('id', 'year', 'specialty', 'education_form', 'amount')

@admin.register(SubjectRate)
class SubjectRateAdmin(ImportExportModelAdmin):
    resource_class = SubjectRateResource
    form = SubjectRateForm
    list_display = ('year', 'specialty', 'education_form', 'get_amount_display')
    list_filter = ('year', 'education_form', 'specialty')            # Filterga ham qo'shildi
    search_fields = ('specialty__name', 'year__name')
    autocomplete_fields = ['year', 'specialty']
    @admin.display(description="Kontrakt narxi", ordering='amount')
    def get_amount_display(self, obj):
        # Agar summa bor bo'lsa, uni formatlaymiz
        if obj.amount:
            # 24000000 -> 24,000,000 -> 24 000 000
            return f"{obj.amount:,.0f}".replace(",", " ")
        return "0"

    class Media:
        js = (
            'admin/js/jquery.init.js',
            'admin/js/money_input.js',
        )

class PerevodRateResource(resources.ModelResource):
    class Meta:
        model = PerevodRate


@admin.register(PerevodRate)
class PerevodRateAdmin(ImportExportModelAdmin):
    resource_class = PerevodRateResource
    form = PerevodRateForm
    list_display = ('year','get_amount_display')
    search_fields = ('year__name',)
    autocomplete_fields = ['year']

    @admin.display(description="Kontrakt narxi", ordering='amount')
    def get_amount_display(self, obj):
        # Agar summa bor bo'lsa, uni formatlaymiz
        if obj.amount:
            # 24000000 -> 24,000,000 -> 24 000 000
            return f"{obj.amount:,.0f}".replace(",", " ")
        return "0"

    class Media:
        js = (
            'admin/js/jquery.init.js',
            'admin/js/money_input.js',
        )



def safe_str(v):
    if v is None:
        return ""
    try:
        return str(v)
    except Exception:
        return ""

def to_float_zero(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, Decimal):
        return float(v)
    try:
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).replace(" ", "").replace(",", "")
        return float(s) if s != "" else 0.0
    except Exception:
        return 0.0


class SubjectDebtEducationFormFilter(admin.SimpleListFilter):
    title = "Ta'lim shakli"
    parameter_name = 'student_education_form'

    def lookups(self, request, model_admin):
        # Student modelidagi variantlarni olib kelamiz
        return Student.EducationFormChoices.choices

    def queryset(self, request, queryset):
        if self.value():
            # SubjectDebt -> Student -> education_form bo'yicha filterlash
            return queryset.filter(student__education_form=self.value())
        return queryset

class SubjectDebtCourseFilter(admin.SimpleListFilter):
    title = "Kurs"
    parameter_name = 'student_course'

    def lookups(self, request, model_admin):
        # 1 dan 5 gacha kurslarni chiqarish
        return [(i, f"{i}-kurs") for i in range(1, 6)]

    def queryset(self, request, queryset):
        if self.value():
            # SubjectDebt -> Student -> course_year bo'yicha filterlash
            return queryset.filter(student__course_year=self.value())
        return queryset


class SubjectDebtStudentStatusFilter(admin.SimpleListFilter):
    title = "Talaba statusi"
    parameter_name = 'student_status_custom'

    def lookups(self, request, model_admin):
        return (
            ('all', "Barcha talabalar"),      # Hammasini ko'rish uchun variant
            ('active', "O'qiydi"),            # Default variant
            ('academic', "Akademik ta'tilda"),
            ('expelled', "Chetlashtirilgan"),
            ('graduated', "Bitirgan"),
        )

    def choices(self, cl):
        # Bu metod admin panelda filterni chizib beradi.
        # Agar URL da hech narsa tanlanmagan bo'lsa (self.value() is None),
        # 'active' ni tanlangan (selected) qilib ko'rsatamiz.
        for lookup, title in self.lookup_choices:
            selected = (self.value() == str(lookup)) or (self.value() is None and lookup == 'active')
            yield {
                'selected': selected,
                'query_string': cl.get_query_string({self.parameter_name: lookup}, []),
                'display': title,
            }

    def queryset(self, request, queryset):
        # Agar URL parametri bo'sh bo'lsa -> Default: 'active'
        if self.value() is None or self.value() == 'active':
            return queryset.filter(student__status='active')
        elif self.value() == 'all':
            return queryset  # Hech qanday filtrsiz hammasini qaytaradi
        else:
            # Boshqa variantlar (expelled, academic, graduated)
            return queryset.filter(student__status=self.value())


@admin.register(SubjectDebt)
class SubjectDebtAdmin(admin.ModelAdmin):
    # Shablon manzili
    change_list_template = "admin/students/student/subjectdebt_change_list.html"
    form = SubjectDebtForm

    list_display = (
        'get_student_name',
        'get_subject_name',
        'semester',
        'credit',
        'get_debt_type_display_custom',
        'get_amount_display',
        'get_paid_display',
        'get_diff_display',
        'get_status_display',
    )

    # --- FILTERLAR RO'YXATI ---
    list_filter = (
        SubjectDebtStudentStatusFilter,  # <--- YANGI DEFAULT FILTER (Eng birinchida)
        'debt_type',
        'academic_year',
        'status',  # Fan qarzdorligi statusi (Yopildi/Yopilmadi)
        'semester',
        SubjectDebtCourseFilter,  # Kurs bo'yicha (custom)
        SubjectDebtEducationFormFilter  # Ta'lim shakli bo'yicha (custom)
    )

    search_fields = ('student__full_name', 'subject__name', 'student__student_hemis_id')
    autocomplete_fields = ['student', 'subject', 'academic_year', 'contract']
    list_per_page = 50
    readonly_fields = ['amount']

    class Media:
        js = (
            'admin/js/jquery.init.js',
            'admin/js/disable_text_input_and_open_lookup.js',
            'admin/js/money_input.js',
        )

    # --- STATISTIKA (FOOTER QISMI) ---
    def changelist_view(self, request, extra_context=None):
        response = super().changelist_view(request, extra_context)
        if not hasattr(response, 'context_data') or 'cl' not in response.context_data:
            return response

        cl = response.context_data['cl']
        qs = cl.queryset  # Bu yerda filterlangan queryset keladi (status bo'yicha ham)

        total_count = qs.count()
        metrics = qs.aggregate(
            jami_qarz=Sum('amount'),
            jami_tolov=Sum('amount_summ')
        )

        total_debt = metrics['jami_qarz'] or 0
        total_paid = metrics['jami_tolov'] or 0
        remaining = total_debt - total_paid

        response.context_data['footer_stats'] = {
            'total_count': total_count,
            'total_debt': total_debt,
            'total_paid': total_paid,
            'remaining': remaining
        }
        return response

    # --- DISPLAY METODLARI ---
    @admin.display(description="Talaba", ordering='student__full_name')
    def get_student_name(self, obj):
        group_name = obj.student.group.name if obj.student.group else "Guruhsiz"
        # Talaba statusiga qarab rangni o'zgartirish (ixtiyoriy vizual qo'shimcha)
        status_color = "#2c3e50"
        if obj.student.status == 'expelled':
            status_color = "#c0392b"  # Qizil
        elif obj.student.status == 'graduated':
            status_color = "#27ae60"  # Yashil

        return format_html(
            '<div style="line-height: 1.2;">'
            '<span style="font-weight:bold; color:{}; font-size:14px;">{}</span><br>'
            '<span class="group-badge">{}</span>'
            '</div>',
            status_color, obj.student.full_name, group_name
        )

    @admin.display(description="Fan", ordering='subject__name')
    def get_subject_name(self, obj):
        return format_html('<span style="font-weight:500; color:#34495e;">{}</span>', obj.subject.name)

    @admin.display(description="Turi", ordering='debt_type')
    def get_debt_type_display_custom(self, obj):
        if obj.debt_type == 'du':
            return format_html('<span class="status-badge" style="background:#e3f2fd; color:#0d47a1;">DU</span>')
        return format_html('<span class="status-badge" style="background:#fff3e0; color:#e65100;">Perevod</span>')

    @admin.display(description="Hisoblangan", ordering='amount')
    def get_amount_display(self, obj):
        return f"{obj.amount:,.0f}".replace(",", " ") if obj.amount else "0"

    @admin.display(description="To'langan", ordering='amount_summ')
    def get_paid_display(self, obj):
        val = obj.amount_summ or 0
        color = "#2c3e50" if val > 0 else "#b2bec3"
        return format_html('<span style="color:{};">{}</span>', color, f"{val:,.0f}".replace(",", " "))

    @admin.display(description="Qoldiq")
    def get_diff_display(self, obj):
        diff = (obj.amount or 0) - (obj.amount_summ or 0)
        if diff <= 0:
            return format_html('<span style="color: #0ca678; font-weight: bold;">✓</span>')
        return format_html('<span class="status-badge badge-danger">{}</span>', f"{diff:,.0f}".replace(",", " "))

    @admin.display(description="Holati", ordering='status')
    def get_status_display(self, obj):
        if obj.status == 'yopildi':
            return format_html('<span class="status-badge badge-success">Yopildi</span>')
        return format_html('<span class="status-badge badge-warning">! Yopilmadi</span>')

    # --- EXPORT EXCEL QISMI ---
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('export-excel/', self.admin_site.admin_view(self.export_excel), name='subjectdebt_export_excel'),
        ]
        return custom_urls + urls

    def export_excel(self, request):
        # 1. Filtrlarni qo'llash (Admin paneldagi joriy filtrlar asosida)
        qs_from_form = request.POST.get('changelist_qs')
        if qs_from_form:
            if qs_from_form.startswith('?'):
                qs_from_form = qs_from_form[1:]
            request.GET = QueryDict(qs_from_form)

        try:
            # get_changelist_instance filtrlarni (shu jumladan bizning yangi status filterini) qo'llaydi
            cl = self.get_changelist_instance(request)
            queryset = cl.get_queryset(request)
        except Exception:
            queryset = self.get_queryset(request)

        # 2. Tartiblash va optimizatsiya
        queryset = queryset.select_related('student', 'student__group', 'subject') \
            .order_by('student__full_name', 'student__id')

        # 3. Tanlangan ustunlar
        selected_fields = request.POST.getlist('fields')

        # Config: (Sarlavha, ValueFunc, MergeQilishKerakmi?)
        field_config = {
            'student__full_name': ("F.I.Sh.", lambda o: o.student.full_name, True),
            'student__phone_number': ("Telefon", lambda o: o.student.phone_number, True),
            'student__group__name': ("Guruh", lambda o: o.student.group.name if o.student.group else "", True),
            'student__education_form': ("Ta'lim shakli", lambda o: o.student.get_education_form_display(), True),
            'student__course_year': ("Kurs", lambda o: str(o.student.course_year), True),
            'semester': ("Semestr", lambda o: str(o.semester), True),

            # Fanlar (O'ng taraf - Merge qilinmaydi)
            'subject__name': ("Fan nomi", lambda o: o.subject.name, False),
            'credit': ("Kredit", lambda o: o.credit, False),
            'amount': ("Hisoblangan summa", lambda o: o.amount, False),
            'amount_summ': ("To'langan summa", lambda o: o.amount_summ, False),
            'debt_type': ("Qarz turi", lambda o: o.get_debt_type_display(), False),
            'status': ("Status", lambda o: o.get_status_display(), False),
        }

        order_list = [
            'student__full_name', 'student__phone_number', 'student__group__name',
            'student__education_form', 'student__course_year', 'semester',
            'subject__name', 'credit', 'amount', 'amount_summ',
            'debt_type', 'status'
        ]

        final_fields = [f for f in order_list if f in selected_fields] if selected_fields else order_list

        # 4. Excel yaratish
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Fan Qarzlari"

        # Dizayn
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill("solid", fgColor="2C3E50")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))

        headers = [field_config[f][0] for f in final_fields]
        headers.extend(["Jami hisoblangan", "Jami qarz"])

        ws.append(headers)
        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        ws.row_dimensions[1].height = 30

        # 5. GURUHLASH LOGIKASI
        data_rows = list(queryset)
        grouped_data = groupby(data_rows, key=lambda x: x.student.id)

        current_row = 2

        for student_id, debts in grouped_data:
            debts_list = list(debts)

            # Jami hisoblar
            total_calc = sum(d.amount or 0 for d in debts_list)
            total_paid = sum(d.amount_summ or 0 for d in debts_list)
            total_debt = total_calc - total_paid

            start_row = current_row

            for i, obj in enumerate(debts_list):
                row_data = []
                for f in final_fields:
                    val = field_config[f][1](obj)
                    row_data.append(val if val is not None else "")

                # Jami ustunlar (Faqat 1-qatorga qiymat yozamiz)
                if i == 0:
                    row_data.append(total_calc)
                    row_data.append(total_debt)
                else:
                    row_data.append("")
                    row_data.append("")

                # Yozish
                for c_idx, val in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=c_idx)
                    cell.value = val
                    cell.border = thin_border
                    cell.alignment = center_align
                    if isinstance(val, (int, float)):
                        cell.number_format = '#,##0'

                current_row += 1

            # MERGE LOGIKASI
            if len(debts_list) > 1:
                end_row = current_row - 1

                # Chap ustunlar merge
                for col_idx, field_key in enumerate(final_fields, 1):
                    if field_config[field_key][2]:  # is_left == True (Merge qilinadigan ustun)
                        ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)

                # Jami ustunlar merge
                total_col = len(headers) - 1
                debt_col = len(headers)
                ws.merge_cells(start_row=start_row, start_column=total_col, end_row=end_row, end_column=total_col)
                ws.merge_cells(start_row=start_row, start_column=debt_col, end_row=end_row, end_column=debt_col)

        # Avto kenglik
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_len: max_len = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Qarzdorlik_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        response = HttpResponse(output.getvalue(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


@admin.register(Hisobot)
class HisobotAdmin(admin.ModelAdmin):
    search_fields = []
    list_filter = []

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('contingent/', self.admin_site.admin_view(self.contingent_view), name='hisobot_contingent'),
            path('contingent/export/', self.admin_site.admin_view(self.export_contingent_excel), name='hisobot_contingent_export'),

            path('kurs-swod/', self.admin_site.admin_view(self.kurs_swod_view), name='hisobot_kurs_swod'),
            path('kurs-swod/export/', self.admin_site.admin_view(self.export_kurs_swod_excel), name='hisobot_kurs_swod_export'),

            path('subject-debt-swod/', self.admin_site.admin_view(self.subject_debt_swod_view), name='hisobot_subject_debt_swod'),
            path('subject-debt-swod/export/', self.admin_site.admin_view(self.export_subject_debt_swod_excel), name='hisobot_subject_debt_swod_export'),

            path('tsch-analiz/', self.admin_site.admin_view(self.tsch_analiz_view), name='hisobot_tsch_analiz'),
            path('tsch-analiz/export/', self.admin_site.admin_view(self.export_tsch_analiz_excel),
                 name='hisobot_tsch_analiz_export'),

            path('internal-grant/', self.admin_site.admin_view(self.internal_grant_view), name='hisobot_internal_grant'),
            path('internal-grant/export/', self.admin_site.admin_view(self.export_internal_grant_excel), name='hisobot_internal_grant_export'),
        ]
        return my_urls + urls

    def changelist_view(self, request, extra_context=None):
        """
        Hisobot menyusi bosilganda chiqadigan ASOSIY sahifa (4 ta knopkali)
        """
        # 1. MUHIM: Admin saytining global kontekstini olamiz (bunda menyu, user va boshqalar bor)
        context = admin.site.each_context(request)

        # 2. O'zimizning maxsus ma'lumotlarni qo'shamiz
        context.update({
            'title': "Hisobotlar markazi",
            # Bu yerda endi 4 ta sub-menyu linklarini yuboramiz
            'menu_items': [
                {
                    'title': 'Contingent',
                    'url': 'contingent/',
                    'icon': 'fas fa-users',
                    'desc': 'Talabalar kontingenti bo‘yicha hisobot'
                },
                {
                    'title': 'Kurs Swod',
                    'url': 'kurs-swod/',
                    'icon': 'fas fa-list-alt',
                    'desc': 'Kurslar kesimida yig‘ma jild'
                },
                {
                    'title': 'Fan Qarzi Swod',  # --- YANGI QO'SHILGAN QISM ---
                    'url': 'subject-debt-swod/',
                    'icon': 'fas fa-exclamation-circle',
                    'desc': 'Fan qarzdorligi bo\'yicha yig\'ma tahlil'
                },
                {
                    'title': 'TSCH Analiz',
                    'url': 'tsch-analiz/',
                    'icon': 'fas fa-chart-pie',
                    'desc': 'Tahliliy hisobot va monitoring'
                },
                {
                    'title': 'Ichki Grant',
                    'url': 'internal-grant/',
                    'icon': 'fas fa-hand-holding-usd',
                    'desc': 'Universitet granti va chegirmalari'
                },
            ]
        })

        # 3. Agar extra_context kelgan bo'lsa, uni ham qo'shamiz
        if extra_context:
            context.update(extra_context)

        return render(request, "admin/hisobot_main.html", context)


    def export_contingent_excel(self, request):
        # 1. FILTRLARNI QABUL QILISH (contingent_view bilan bir xil)
        if request.GET:
            selected_statuses = request.GET.getlist('status')
            selected_forms = request.GET.getlist('form')
            selected_date = request.GET.get('date')
        else:
            selected_statuses = ['active']
            selected_forms = []
            selected_date = ''

        if not selected_statuses and not request.GET:
            selected_statuses = ['active']

        # 2. QUERYSET (contingent_view bilan bir xil mantiq)
        qs = Student.objects.select_related('group', 'group__specialty')

        if selected_forms:
            qs = qs.filter(education_form__in=selected_forms)

        if selected_date:
            qs = qs.filter(created_at__date__lte=selected_date)
            last_order_date_qs = Order.objects.filter(student=OuterRef('pk')).order_by('-order_date').values(
                'order_date')[:1]
            qs = qs.annotate(real_exit_date=Subquery(last_order_date_qs))

            status_conditions = Q()
            date_check = Q(real_exit_date__lte=selected_date) | Q(real_exit_date__isnull=True)

            if 'active' in selected_statuses:
                condition_still_active = Q(status='active')
                condition_was_active = Q(status__in=['expelled', 'graduated', 'academic'],
                                         real_exit_date__gt=selected_date)
                status_conditions |= (condition_still_active | condition_was_active)
            if 'expelled' in selected_statuses:
                status_conditions |= Q(status='expelled') & date_check
            if 'graduated' in selected_statuses:
                status_conditions |= Q(status='graduated') & date_check
            if 'academic' in selected_statuses:
                status_conditions |= Q(status='academic') & date_check

            if status_conditions:
                qs = qs.filter(status_conditions)
        else:
            if selected_statuses:
                qs = qs.filter(status__in=selected_statuses)

        # 3. MA'LUMOTLARNI YIG'ISH
        data_map = {}
        total_counts = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 'total': 0}
        form_totals = {
            'kunduzgi': {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 'total': 0},
            'sirtqi': {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 'total': 0},
            'kechki': {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 'total': 0},
        }

        grouped_data = qs.values('group__specialty__name', 'education_form', 'course_year').annotate(
            count=Count('id')).order_by('group__specialty__name')

        for item in grouped_data:
            spec_name = item['group__specialty__name'] or "Noma'lum yo'nalish"
            form = item['education_form']
            course = item['course_year']
            count = item['count']

            if spec_name not in data_map: data_map[spec_name] = {}
            if form not in data_map[spec_name]: data_map[spec_name][form] = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0,
                                                                             'row_total': 0}

            if course in [1, 2, 3, 4, 5]:
                data_map[spec_name][form][course] += count
                data_map[spec_name][form]['row_total'] += count
                total_counts[course] += count
                total_counts['total'] += count
                if form in form_totals:
                    form_totals[form][course] += count
                    form_totals[form]['total'] += count

        # 4. EXCEL YARATISH
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kontingent"
        ws.sheet_view.showGridLines = False

        # --- STYLES ---
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="2C3E50")

        # Badge ranglari (HTMLdagi kabi)
        fill_kunduzgi = PatternFill("solid", fgColor="E8F5E9")  # Yashilroq
        font_kunduzgi = Font(color="2E7D32", bold=True)

        fill_sirtqi = PatternFill("solid", fgColor="E3F2FD")  # Ko'kroq
        font_sirtqi = Font(color="1565C0", bold=True)

        fill_kechki = PatternFill("solid", fgColor="FFF3E0")  # Zarg'aldoq
        font_kechki = Font(color="EF6C00", bold=True)

        footer_fill = PatternFill("solid", fgColor="ECF0F1")
        footer_font = Font(bold=True, color="2C3E50")

        grand_total_fill = PatternFill("solid", fgColor="2C3E50")
        grand_total_font = Font(bold=True, color="FFFFFF", size=12)

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # --- SARLAVHA ---
        current_date_str = selected_date if selected_date else datetime.now().strftime("%d.%m.%Y")
        ws.merge_cells('A1:I1')
        title_cell = ws['A1']
        title_cell.value = f"Talabalar kontingenti hisoboti ({current_date_str})"
        title_cell.font = Font(bold=True, size=14, color="2C3E50")
        title_cell.alignment = align_center

        # --- JADVAL BOSHI ---
        headers = ["№", "Ta'lim yo'nalishi", "Ta'lim shakli", "1-kurs", "2-kurs", "3-kurs", "4-kurs", "5-kurs", "Jami"]
        row_num = 3

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border

            # Kengliklar
            col_letter = cell.column_letter
            if col_idx == 2:
                ws.column_dimensions[col_letter].width = 40
            elif col_idx == 3:
                ws.column_dimensions[col_letter].width = 15
            elif col_idx == 1:
                ws.column_dimensions[col_letter].width = 5
            else:
                ws.column_dimensions[col_letter].width = 10

        # --- DATA ROWS ---
        row_num += 1
        counter = 1
        form_priority = {'kunduzgi': 1, 'sirtqi': 2, 'kechki': 3}

        for spec_name, forms in data_map.items():
            sorted_forms = sorted(forms.items(), key=lambda item: form_priority.get(item[0], 10))
            start_row = row_num

            for form_name, counts in sorted_forms:
                # 1. Index va Nomi (keyinroq merge qilinadi)
                ws.cell(row=row_num, column=1).value = counter
                ws.cell(row=row_num, column=2).value = spec_name

                # 2. Ta'lim shakli (Rangli badge effekti)
                form_cell = ws.cell(row=row_num, column=3)
                form_cell.value = form_name.capitalize()
                form_cell.alignment = align_center

                if form_name == 'kunduzgi':
                    form_cell.fill = fill_kunduzgi
                    form_cell.font = font_kunduzgi
                elif form_name == 'sirtqi':
                    form_cell.fill = fill_sirtqi
                    form_cell.font = font_sirtqi
                elif form_name == 'kechki':
                    form_cell.fill = fill_kechki
                    form_cell.font = font_kechki

                # 3. Raqamlar
                courses = [1, 2, 3, 4, 5, 'row_total']
                for i, c_key in enumerate(courses, 4):
                    val = counts[c_key]
                    c_cell = ws.cell(row=row_num, column=i)
                    c_cell.value = val if val > 0 else ""
                    c_cell.alignment = align_center

                    # Jami ustuni
                    if c_key == 'row_total':
                        c_cell.font = Font(bold=True)
                        c_cell.fill = PatternFill("solid", fgColor="F8F9FA")

                # Borderni barcha kataklarga qo'yish
                for c in range(1, 10):
                    ws.cell(row=row_num, column=c).border = thin_border
                    if c == 1: ws.cell(row=row_num, column=c).alignment = align_center
                    if c == 2: ws.cell(row=row_num, column=c).alignment = align_left

                row_num += 1

            # Merge qilish (Spec va Index)
            if len(sorted_forms) > 1:
                end_row = row_num - 1
                ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)

                # Vertical center qilish merge qilingan kataklarni
                ws.cell(row=start_row, column=1).alignment = align_center
                ws.cell(row=start_row, column=2).alignment = align_left

            counter += 1

        # --- FOOTER (Jami qatorlari) ---
        def write_footer_row(label, data_dict, fill_style=footer_fill, font_style=footer_font):
            nonlocal row_num
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            label_cell = ws.cell(row=row_num, column=1)
            label_cell.value = label
            label_cell.alignment = Alignment(horizontal="right", vertical="center")

            # Style apply loop
            for c in range(1, 10):
                cell = ws.cell(row=row_num, column=c)
                cell.fill = fill_style
                cell.font = font_style
                cell.border = thin_border

            # Form nomi
            if "Kunduzgi" in label:
                f_cell = ws.cell(row=row_num, column=3, value="Kunduzgi")
                f_cell.font = font_kunduzgi
                f_cell.fill = fill_kunduzgi
            elif "Sirtqi" in label:
                f_cell = ws.cell(row=row_num, column=3, value="Sirtqi")
                f_cell.font = font_sirtqi
                f_cell.fill = fill_sirtqi
            elif "Kechki" in label:
                f_cell = ws.cell(row=row_num, column=3, value="Kechki")
                f_cell.font = font_kechki
                f_cell.fill = fill_kechki

            # Raqamlar
            for i, c_key in enumerate([1, 2, 3, 4, 5, 'total'], 4):
                val = data_dict[c_key]
                ws.cell(row=row_num, column=i).value = val
                ws.cell(row=row_num, column=i).alignment = align_center

            row_num += 1

        write_footer_row("Jami:", form_totals['kunduzgi'])
        write_footer_row("Jami:", form_totals['sirtqi'])

        # Kechki bor bo'lsa
        if form_totals['kechki']['total'] > 0:
            write_footer_row("Jami:", form_totals['kechki'])

        # Grand Total
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
        gt_label = ws.cell(row=row_num, column=1, value="UMUMIY JAMI:")
        gt_label.alignment = Alignment(horizontal="right", vertical="center")

        for c in range(1, 10):
            cell = ws.cell(row=row_num, column=c)
            cell.fill = grand_total_fill
            cell.font = grand_total_font
            cell.border = thin_border
            # Values
            if c >= 4:
                idx = c - 3
                if c == 9:
                    key = 'total'
                else:
                    key = idx
                cell.value = total_counts[key]
                cell.alignment = align_center

        # --- RESPONSE ---
        filename = f"Kontingent_{current_date_str}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    def contingent_view(self, request):
        # --- 1. FILTRLARNI QABUL QILISH ---
        if request.GET:
            selected_statuses = request.GET.getlist('status')
            selected_forms = request.GET.getlist('form')
            selected_date = request.GET.get('date')
        else:
            # Default: Hech narsa tanlanmasa faqat Active
            selected_statuses = ['active']
            selected_forms = []
            selected_date = ''

        # Agar status tanlanmagan bo'lsa, default active qilamiz (logika ishlashi uchun)
        if not selected_statuses and not request.GET:
            selected_statuses = ['active']

        # --- 2. QUERYSET TUZISH ---
        qs = Student.objects.select_related('group', 'group__specialty')

        # Filter: Ta'lim shakli (o'zgarmas bo'lgani uchun oddiy filter)
        if selected_forms:
            qs = qs.filter(education_form__in=selected_forms)

        # --- MUHIM QISM: SANA VA STATUS MANTIGI ---
        if selected_date:
            # 1. Avval shu sanagacha bazaga kiritilganlarni olamiz
            qs = qs.filter(created_at__date__lte=selected_date)

            # 2. Har bir talabaning ENG SO'NGGI buyruq sanasini aniqlaymiz.
            # Biz taxmin qilamizki, status o'zgarishiga sabab bo'lgan narsa - bu oxirgi buyruq.
            last_order_date_qs = Order.objects.filter(
                student=OuterRef('pk')
            ).order_by('-order_date').values('order_date')[:1]

            qs = qs.annotate(real_exit_date=Subquery(last_order_date_qs))

            # 3. Murakkab Status Filteri (Tarixiy tiklash)
            status_conditions = Q()

            # A) Agar foydalanuvchi "ACTIVE" (O'qiydi) ni so'rasa:
            if 'active' in selected_statuses:
                # Mantiq:
                # 1. Hozir ham statusi 'active' bo'lsa -> Olamiz.
                # 2. Hozir statusi 'expelled'/'graduated'/'academic' bo'lsa,
                #    LEKIN buyruq sanasi tanlangan sanadan KEYIN bo'lsa -> Demak o'sha kuni u hali Active edi -> Olamiz.

                condition_still_active = Q(status='active')

                # Hozir no-aktiv, lekin o'sha paytda aktiv bo'lganlar (Buyrug'i keyin chiqqan)
                condition_was_active = Q(
                    status__in=['expelled', 'graduated', 'academic'],
                    real_exit_date__gt=selected_date
                )

                # Ikkalasini birlashtiramiz
                status_conditions |= (condition_still_active | condition_was_active)

            # B) Agar foydalanuvchi "EXPELLED" (Chetlashtirilgan) ni so'rasa:
            if 'expelled' in selected_statuses:
                # Mantiq: Hozir statusi 'expelled' VA buyruq sanasi o'sha sanadan oldin yoki teng bo'lsa.
                # Agar bugun chetlashtirsam, kechagi sanada bu shart bajarilmaydi (to'g'ri ishlaydi).
                status_conditions |= Q(status='expelled', real_exit_date__lte=selected_date)

            # C) Boshqa statuslar (Bitirgan va h.k.)
            if 'graduated' in selected_statuses:
                status_conditions |= Q(status='graduated', real_exit_date__lte=selected_date)

            if 'academic' in selected_statuses:
                status_conditions |= Q(status='academic', real_exit_date__lte=selected_date)

            # D) "Hammasi" (Jami) tanlansa yoki hech narsa tanlanmasa
            # Yuqoridagi shartlar ishlatiladi. Agar status_conditions bo'sh bo'lmasa, filterlaymiz.
            if status_conditions:
                qs = qs.filter(status_conditions)

        else:
            # SANA YO'Q BO'LSA - ODDIY STATUS FILTER (Hozirgi holat)
            if selected_statuses:
                qs = qs.filter(status__in=selected_statuses)

        # --- 3. MA'LUMOTLARNI YIG'ISH (Eski kod bilan bir xil) ---
        data_map = {}
        total_counts = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 'total': 0}
        form_totals = {
            'kunduzgi': {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 'total': 0},
            'sirtqi': {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 'total': 0},
            'kechki': {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 'total': 0},
        }

        # Optimallashtirish: annotate qilingan queryni ishlatamiz
        grouped_data = qs.values(
            'group__specialty__name',
            'education_form',
            'course_year'
        ).annotate(count=Count('id')).order_by('group__specialty__name')

        for item in grouped_data:
            spec_name = item['group__specialty__name'] or "Noma'lum yo'nalish"
            form = item['education_form']
            course = item['course_year']
            count = item['count']

            if spec_name not in data_map:
                data_map[spec_name] = {}
            if form not in data_map[spec_name]:
                data_map[spec_name][form] = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 'row_total': 0}

            if course in [1, 2, 3, 4, 5]:
                data_map[spec_name][form][course] += count
                data_map[spec_name][form]['row_total'] += count
                total_counts[course] += count
                total_counts['total'] += count
                if form in form_totals:
                    form_totals[form][course] += count
                    form_totals[form]['total'] += count

        # --- 4. VIEW TAYYORLASH ---
        report_data = []
        counter = 1
        form_priority = {'kunduzgi': 1, 'sirtqi': 2, 'kechki': 3}

        for spec_name, forms in data_map.items():
            rows = []
            sorted_forms = sorted(forms.items(), key=lambda item: form_priority.get(item[0], 10))
            for form_name, counts in sorted_forms:
                rows.append({
                    'form': form_name,
                    'c1': counts[1], 'c2': counts[2], 'c3': counts[3],
                    'c4': counts[4], 'c5': counts[5], 'total': counts['row_total']
                })
            report_data.append({'index': counter, 'spec_name': spec_name, 'rows': rows, 'rowspan': len(rows)})
            counter += 1

        context = admin.site.each_context(request)
        context.update({
            'title': "Talabalar kontingenti",
            'report_data': report_data,
            'total_counts': total_counts,
            'form_totals': form_totals,
            'current_date': selected_date if selected_date else datetime.now().strftime("%Y-%m-%d"),
            'filter_options': {
                'statuses': Student.StatusChoices.choices,
                'forms': Student.EducationFormChoices.choices,
            },
            'selected_filters': {
                'status': selected_statuses,
                'form': selected_forms,
                'date': selected_date
            }
        })
        return render(request, "admin/reports/contingent.html", context)

    def _get_kurs_swod_data(self, request):
        # --- 1. FILTRLAR ---
        all_years_qs = AcademicYear.objects.all().order_by('-name')
        selected_year = request.GET.get('year')
        target_years = []

        if selected_year:
            target_years = [selected_year]
        else:
            active_year = AcademicYear.objects.filter(is_active=True).first()
            if active_year:
                selected_year = str(active_year.id)
                target_years = [str(active_year.id)]

        limit_date = request.GET.get('limit_date')

        if request.GET:
            selected_forms = request.GET.getlist('form')
            selected_courses = request.GET.getlist('course')
            selected_statuses = request.GET.getlist('status') if 'status' in request.GET else []
        else:
            selected_forms = []
            selected_courses = []
            selected_statuses = ['active']

        # --- 2. QUERYSET ---
        qs = Student.objects.select_related('group', 'group__specialty')

        if selected_statuses: qs = qs.filter(status__in=selected_statuses)
        if selected_forms: qs = qs.filter(education_form__in=selected_forms)
        if selected_courses:
            try:
                cl = [int(c) for c in selected_courses if str(c).isdigit()]
                if cl: qs = qs.filter(course_year__in=cl)
            except ValueError:
                pass

        # --- 3. ANNOTATSIYA (O'ZGARTIRILDI) ---
        # Grant summasini shartnoma summasidan ayirib tashlaymiz
        contract_subquery = Contract.objects.filter(
            student=OuterRef('pk'),
            academic_year__id__in=target_years,
            contract_type='contract'
        ).values('student').annotate(
            real_total=Sum(
                F('amount') - Coalesce(F('grant_amount'), Value(Decimal('0')))
            )
        ).values('real_total')

        payment_filters = Q(
            contract__student=OuterRef('pk'),
            contract__academic_year__id__in=target_years,
            contract__contract_type='contract'
        )
        if limit_date: payment_filters &= Q(payment_date__lte=limit_date)

        payment_subquery = Payment.objects.filter(payment_filters).values('contract__student').annotate(
            total=Sum('amount')
        ).values('total')

        qs = qs.annotate(
            contract_sum=Coalesce(Subquery(contract_subquery), Value(Decimal(0))),
            paid_sum=Coalesce(Subquery(payment_subquery), Value(Decimal(0)))
        )

        # --- 4. HISOBLASH ---
        data_map = {}
        grand_total = {'count': 0, 'p0': 0, 'p1_24': 0, 'p25': 0, 'p26_49': 0, 'p50': 0, 'p51_74': 0, 'p75': 0,
                       'p76_99': 0, 'p100': 0, 'contract': 0, 'paid': 0, 'debt': 0}

        patok_configs = [
            {'id': 1, 'date': '1-oktabr', 'percent': 0.25, 'label': '25%'},
            {'id': 2, 'date': '1-yanvar', 'percent': 0.50, 'label': '50%'},
            {'id': 3, 'date': '1-mart', 'percent': 0.75, 'label': '75%'},
            {'id': 4, 'date': '1-may', 'percent': 1.00, 'label': '100%'},
        ]
        patok_stats = [
            {'id': p['id'], 'date': p['date'], 'label': p['label'], 'target_percent': int(p['percent'] * 100),
             'total_contract': 0, 'target_amount': 0, 'paid_amount': 0, 'debt_amount': 0} for p in patok_configs]

        for student in qs:
            spec_name = student.group.specialty.name if student.group and student.group.specialty else "Noma'lum"
            form = student.education_form
            course = student.course_year

            c_sum = int(student.contract_sum) # Bu yerda endi (Contract - Grant) keladi
            p_sum = int(student.paid_sum)
            debt = max(0, c_sum - p_sum)

            # Patok Logikasi
            prev_tgt, prev_paid = 0, 0
            for i, p_conf in enumerate(patok_configs):
                cur_tgt = int(c_sum * p_conf['percent'])
                this_tgt = cur_tgt - prev_tgt

                cur_paid = min(p_sum, cur_tgt)
                this_paid = cur_paid - prev_paid
                this_debt = max(0, this_tgt - this_paid)

                patok_stats[i]['total_contract'] += c_sum
                patok_stats[i]['target_amount'] += this_tgt
                patok_stats[i]['paid_amount'] += this_paid
                patok_stats[i]['debt_amount'] += this_debt

                prev_tgt, prev_paid = cur_tgt, cur_paid

            # Percent
            percent = round((p_sum / c_sum * 100), 1) if c_sum > 0 else 0
            if c_sum <= 0 and p_sum == 0:
                # Agar kontrakt 0 bo'lsa (grant 100% bo'lsa), talaba 100% to'lagan hisoblanadi
                 percent = 100 # yoki 0, vaziyatga qarab. Lekin qarz yo'q.

            # Map to structure
            if spec_name not in data_map: data_map[spec_name] = {}
            if form not in data_map[spec_name]: data_map[spec_name][form] = {}
            if course not in data_map[spec_name][form]:
                data_map[spec_name][form][course] = {'count': 0, 'p0': 0, 'p1_24': 0, 'p25': 0, 'p26_49': 0, 'p50': 0,
                                                     'p51_74': 0, 'p75': 0, 'p76_99': 0, 'p100': 0, 'contract': 0,
                                                     'paid': 0, 'debt': 0}

            item = data_map[spec_name][form][course]
            for k, v in [('count', 1), ('contract', c_sum), ('paid', p_sum), ('debt', debt)]:
                item[k] += v
                grand_total[k] += v

            p_key = 'p0'
            if percent <= 0:
                p_key = 'p0'
            elif 0 < percent < 25:
                p_key = 'p1_24'
            elif 25 <= percent < 26:
                p_key = 'p25'
            elif 26 <= percent < 50:
                p_key = 'p26_49'
            elif 50 <= percent < 51:
                p_key = 'p50'
            elif 51 <= percent < 75:
                p_key = 'p51_74'
            elif 75 <= percent < 76:
                p_key = 'p75'
            elif 76 <= percent < 100:
                p_key = 'p76_99'
            else:
                p_key = 'p100'

            item[p_key] += 1
            grand_total[p_key] += 1

        # Final Report List
        report_data = []
        counter = 1
        form_priority = {'kunduzgi': 1, 'sirtqi': 2, 'kechki': 3}

        for spec_name, forms in sorted(data_map.items()):
            rows = []
            for form_name, courses in sorted(forms.items(), key=lambda x: form_priority.get(x[0], 10)):
                sub = {'count': 0, 'p0': 0, 'p1_24': 0, 'p25': 0, 'p26_49': 0, 'p50': 0, 'p51_74': 0, 'p75': 0,
                       'p76_99': 0, 'p100': 0, 'contract': 0, 'paid': 0, 'debt': 0}
                for course_num, vals in sorted(courses.items()):
                    rows.append({'form': form_name, 'course': f"{course_num}-kurs", 'vals': vals, 'is_total': False})
                    for k in sub: sub[k] += vals[k]
                rows.append({'form': form_name, 'course': 'Jami', 'vals': sub, 'is_total': True})

            report_data.append({'index': counter, 'spec_name': spec_name, 'rows': rows, 'rowspan': len(rows)})
            counter += 1

        return {
            'report_data': report_data,
            'grand_total': grand_total,
            'patok_stats': patok_stats,
            'selected_year': selected_year,
            'limit_date': limit_date,
            'years': all_years_qs,
            'forms': Student.EducationFormChoices.choices,
            'courses': [str(i) for i in range(1, 6)],
            'statuses': Student.StatusChoices.choices,
            'selected_forms': selected_forms,
            'selected_courses': selected_courses,
            'selected_statuses': selected_statuses
        }

    def kurs_swod_view(self, request):
        data = self._get_kurs_swod_data(request)
        grand_total = data['grand_total']

        # --- FOIZLARNI HISOBLASH (YANGILANGAN) ---
        grand_total_percent = {}
        total_count = grand_total['count']
        total_contract = grand_total['contract']

        # 1. Talabalar soni bo'yicha foizlar
        keys_count = ['p0', 'p1_24', 'p25', 'p26_49', 'p50', 'p51_74', 'p75', 'p76_99', 'p100']
        for key in keys_count:
            if total_count > 0:
                percent = (grand_total[key] / total_count) * 100
                grand_total_percent[key] = f"{percent:.1f}%"
            else:
                grand_total_percent[key] = "0.0%"

        # 2. Summalar bo'yicha foizlar (MANTIQ O'ZGARTIRILDI)
        if total_contract > 0:
            # A) To'lov foizini hisoblaymiz va 1 xonagacha yaxlitlaymiz
            paid_raw = (grand_total['paid'] / total_contract) * 100
            paid_str = f"{paid_raw:.1f}"
            paid_float = float(paid_str)

            # B) Qarz foizini hisoblaymiz
            # Agar to'lov 100% dan oshmagan bo'lsa, qarzni [100 - to'lov] qilib olamiz.
            # Bu vizual xatolikni (100.1%) yo'qotadi.
            if paid_float <= 100:
                debt_float = 100.0 - paid_float
                debt_str = f"{debt_float:.1f}"
            else:
                # Agar ortiqcha to'lov bo'lsa (masalan 105%), qarzni o'z holicha hisoblaymiz
                debt_raw = (grand_total['debt'] / total_contract) * 100
                debt_str = f"{debt_raw:.1f}"

            grand_total_percent['paid'] = f"{paid_str}%"
            grand_total_percent['debt'] = f"{debt_str}%"
        else:
            grand_total_percent['paid'] = "0.0%"
            grand_total_percent['debt'] = "0.0%"

        context = admin.site.each_context(request)
        context.update({
            'title': "Kontrakt Swod Hisoboti",
            'report_data': data['report_data'],
            'grand_total': data['grand_total'],
            'grand_total_percent': grand_total_percent,
            'patok_stats': data['patok_stats'],
            'filter_options': {
                'years': data['years'],
                'forms': data['forms'],
                'courses': [(i, f"{i}-kurs") for i in range(1, 6)],
                'statuses': data['statuses']
            },
            'selected_filters': {
                'year': data['selected_year'],
                'form': data['selected_forms'],
                'course': data['selected_courses'],
                'status': data['selected_statuses'],
                'limit_date': data['limit_date']
            },
            'current_date': datetime.now().strftime("%d.%m.%Y")
        })
        return render(request, "admin/reports/kurs_swod.html", context)

    def export_kurs_swod_excel(self, request):
        # 1. Ma'lumotlarni olish
        data = self._get_kurs_swod_data(request)
        report_data = data['report_data']
        grand_total = data['grand_total']
        patok_stats = data['patok_stats']

        # 2. Excel yaratish
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kontrakt Swod"
        ws.sheet_view.showGridLines = False

        # --- STYLES (DIZAYN) ---
        header_fill = PatternFill('solid', fgColor='2C3E50')
        header_font = Font(bold=True, color='FFFFFF', name='Calibri', size=10)

        # Headerdagi maxsus ranglar
        fill_green_head = PatternFill('solid', fgColor='1E8449')
        fill_blue_head = PatternFill('solid', fgColor='2874A6')
        fill_red_head = PatternFill('solid', fgColor='C0392B')

        # Katakchalar fon ranglari (Jami qatorlar uchun)
        fill_green_cell = PatternFill('solid', fgColor='D4EFDF')
        fill_blue_cell = PatternFill('solid', fgColor='D6EAF8')
        fill_red_cell = PatternFill('solid', fgColor='FADBD8')

        # Yangi foiz qatori uchun ranglar
        fill_pct_label = PatternFill('solid', fgColor='D6EAF8')  # Och ko'k fon label uchun
        fill_pct_cell = PatternFill('solid', fgColor='EBF5FB')  # Juda och ko'k fon raqamlar uchun

        fill_total_row = PatternFill('solid', fgColor='ECF0F1')
        font_total = Font(bold=True, name='Calibri', size=10)

        thin_border = Border(left=Side(style='thin', color='AAAAAA'),
                             right=Side(style='thin', color='AAAAAA'),
                             top=Side(style='thin', color='AAAAAA'),
                             bottom=Side(style='thin', color='AAAAAA'))

        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        align_right = Alignment(horizontal='right', vertical='center')

        money_fmt = '#,##0'

        # =========================================================
        # 1. PATOKLAR (YONMA-YON)
        # =========================================================
        patok_colors = ['F39C12', '3498DB', '9B59B6', '27AE60']

        # Asosiy sarlavha
        ws.merge_cells('B1:Q1')
        ws['B1'] = "To'lov Rejasi (Patoklar bo'yicha tahlil)"
        ws['B1'].font = Font(bold=True, size=14, color='2C3E50')
        ws['B1'].alignment = align_center

        # Patoklar boshlanish joyi
        start_row = 3
        base_col = 2  # B ustunidan boshlanadi
        block_width = 4  # Har bir patok 4 ta ustun egallaydi

        for idx, stat in enumerate(patok_stats):
            col_start = base_col + (idx * block_width)
            col_end = col_start + block_width - 1
            current_r = start_row

            # A) Header
            cell = ws.cell(row=current_r, column=col_start)
            ws.merge_cells(start_row=current_r, start_column=col_start, end_row=current_r, end_column=col_end)
            cell.value = f"{stat['id']}-Patok ({stat['label']}) | {stat['date']}"
            cell.fill = PatternFill('solid', fgColor=patok_colors[idx])
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = align_center
            cell.border = thin_border
            current_r += 1

            # B) Body
            metrics = [
                (f"Reja ({stat['target_percent']}%):", stat['target_amount']),
                ("Qoplangan qismi:", stat['paid_amount']),
                ("Muddati o'tgan qarz:", stat['debt_amount'])
            ]

            for label, val in metrics:
                l_cell = ws.cell(row=current_r, column=col_start)
                ws.merge_cells(start_row=current_r, start_column=col_start, end_row=current_r, end_column=col_start + 1)
                l_cell.value = label
                l_cell.alignment = align_left
                l_cell.font = Font(color='555555', size=9)
                l_cell.border = thin_border

                v_cell = ws.cell(row=current_r, column=col_start + 2)
                ws.merge_cells(start_row=current_r, start_column=col_start + 2, end_row=current_r, end_column=col_end)
                v_cell.value = val
                v_cell.number_format = money_fmt
                v_cell.alignment = align_right
                v_cell.font = Font(bold=True, size=10)
                v_cell.border = thin_border

                if "Qoplangan" in label: v_cell.font = Font(bold=True, color='27AE60', size=10)
                if "qarz" in label: v_cell.font = Font(bold=True, color='E74C3C', size=10)
                current_r += 1

        # =========================================================
        # 2. ASOSIY JADVAL HEADERS
        # =========================================================
        table_start_row = 9

        # 1-Qator Headerlari
        headers = [
            ('A', '№', 1, 2),
            ('B', "Ta'lim yo'nalishi", 1, 2),
            ('C', "Shakl", 1, 2),
            ('D', "Kurs", 1, 2),
            ('E', "Talaba\nsoni", 1, 2),
            ('F', "To'lov foizlari (talabalar soni)", 9, 1),
            ('O', "Jami Shartnoma", 1, 2),
            ('P', "To'langan Summa", 1, 2),
            ('Q', "Qarzdorlik", 1, 2),
        ]

        for col_char, text, colspan, rowspan in headers:
            cell = ws[f"{col_char}{table_start_row}"]
            cell.value = text
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border

            if "Shartnoma" in text: cell.fill = fill_green_head
            if "To'langan" in text: cell.fill = fill_blue_head
            if "Qarzdorlik" in text: cell.fill = fill_red_head

            if colspan > 1 or rowspan > 1:
                col_idx = openpyxl.utils.column_index_from_string(col_char)
                ws.merge_cells(
                    start_row=table_start_row, start_column=col_idx,
                    end_row=table_start_row + rowspan - 1, end_column=col_idx + colspan - 1
                )

        # 2-Qator Headerlari (Foizlar)
        percents = ["0%", "1-24%", "25%", "26-49%", "50%", "51-74%", "75%", "76-99%", "100%"]
        start_col_idx = 6
        for p in percents:
            cell = ws.cell(row=table_start_row + 1, column=start_col_idx)
            cell.value = p
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border
            start_col_idx += 1

        # =========================================================
        # 3. JADVAL BODY
        # =========================================================
        current_row = table_start_row + 2

        for item in report_data:
            spec_start_row = current_row

            cell_idx = ws.cell(row=current_row, column=1, value=item['index'])
            cell_idx.alignment = align_center
            cell_idx.border = thin_border
            cell_idx.font = Font(bold=True)

            cell_name = ws.cell(row=current_row, column=2, value=item['spec_name'])
            cell_name.alignment = align_left
            cell_name.border = thin_border
            cell_name.font = Font(bold=True)

            for row_data in item['rows']:
                is_total = row_data.get('is_total', False)
                vals = row_data['vals']

                row_font = font_total if is_total else Font(name='Calibri', size=10)
                row_fill = fill_total_row if is_total else PatternFill(fill_type=None)

                cols = [
                    (3, row_data['form'] if not is_total else "Jami"),
                    (4, row_data['course'] if not is_total else "-"),
                    (5, vals['count']),
                    (6, vals['p0']), (7, vals['p1_24']), (8, vals['p25']),
                    (9, vals['p26_49']), (10, vals['p50']), (11, vals['p51_74']),
                    (12, vals['p75']), (13, vals['p76_99']), (14, vals['p100']),
                    (15, vals['contract']), (16, vals['paid']), (17, vals['debt'])
                ]

                for col_idx, val in cols:
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.value = val if val != 0 else ""

                    if col_idx in [15, 16, 17] and val != "":
                        cell.value = val
                        cell.number_format = money_fmt

                    cell.font = row_font
                    cell.fill = row_fill
                    cell.border = thin_border
                    cell.alignment = align_center

                    if is_total:
                        if col_idx == 15: cell.fill = fill_green_cell
                        if col_idx == 16: cell.fill = fill_blue_cell
                        if col_idx == 17: cell.fill = fill_red_cell

                current_row += 1

            if item['rowspan'] > 1:
                ws.merge_cells(start_row=spec_start_row, start_column=1, end_row=current_row - 1, end_column=1)
                ws.merge_cells(start_row=spec_start_row, start_column=2, end_row=current_row - 1, end_column=2)

        # =========================================================
        # 4. FOOTER - RAQAMLAR (UMUMIY JAMI)
        # =========================================================
        cell_label = ws.cell(row=current_row, column=4, value="UMUMIY JAMI:")
        cell_label.font = header_font
        cell_label.fill = header_fill
        cell_label.alignment = align_right

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        for c in range(1, 5):
            c_tmp = ws.cell(row=current_row, column=c)
            c_tmp.fill = header_fill
            c_tmp.border = thin_border

        cols_total = [
            (5, grand_total['count']),
            (6, grand_total['p0']), (7, grand_total['p1_24']), (8, grand_total['p25']),
            (9, grand_total['p26_49']), (10, grand_total['p50']), (11, grand_total['p51_74']),
            (12, grand_total['p75']), (13, grand_total['p76_99']), (14, grand_total['p100']),
            (15, grand_total['contract']), (16, grand_total['paid']), (17, grand_total['debt'])
        ]

        for col_idx, val in cols_total:
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = val
            if col_idx >= 15: cell.number_format = money_fmt
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border

        # =========================================================
        # 5. FOOTER - FOIZLAR (YANGI QO'SHILDI)
        # =========================================================
        current_row += 1

        cell_label_pct = ws.cell(row=current_row, column=4, value="FOIZ KO'RSATGICHI:")
        cell_label_pct.font = Font(bold=True, italic=True, color='2C3E50')
        cell_label_pct.fill = fill_pct_label
        cell_label_pct.alignment = align_right

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        for c in range(1, 5):
            c_tmp = ws.cell(row=current_row, column=c)
            c_tmp.fill = fill_pct_label
            c_tmp.border = thin_border

        total_cnt = grand_total['count']
        total_sum = grand_total['contract']

        # Oddiy foiz hisoblash funksiyasi
        def get_pct_val(val, total):
            if total > 0:
                return (val / total) * 100
            return 0.0

        # To'lov va Qarz foizlarini maxsus hisoblash
        paid_pct_val = get_pct_val(grand_total['paid'], total_sum)
        paid_pct_str = f"{paid_pct_val:.1f}"
        paid_pct_float = float(paid_pct_str)

        if paid_pct_float <= 100:
            debt_pct_str = f"{100.0 - paid_pct_float:.1f}"
        else:
            debt_pct_val = get_pct_val(grand_total['debt'], total_sum)
            debt_pct_str = f"{debt_pct_val:.1f}"

        cols_percent = [
            (5, "100%"),
            (6, f"{get_pct_val(grand_total['p0'], total_cnt):.1f}%"),
            (7, f"{get_pct_val(grand_total['p1_24'], total_cnt):.1f}%"),
            (8, f"{get_pct_val(grand_total['p25'], total_cnt):.1f}%"),
            (9, f"{get_pct_val(grand_total['p26_49'], total_cnt):.1f}%"),
            (10, f"{get_pct_val(grand_total['p50'], total_cnt):.1f}%"),
            (11, f"{get_pct_val(grand_total['p51_74'], total_cnt):.1f}%"),
            (12, f"{get_pct_val(grand_total['p75'], total_cnt):.1f}%"),
            (13, f"{get_pct_val(grand_total['p76_99'], total_cnt):.1f}%"),
            (14, f"{get_pct_val(grand_total['p100'], total_cnt):.1f}%"),
            (15, "100%"),
            (16, f"{paid_pct_str}%"),  # To'lov
            (17, f"{debt_pct_str}%")  # Qarz (To'g'irlangan)
        ]

        for col_idx, val in cols_percent:
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = val
            cell.font = Font(bold=True, italic=True)
            cell.fill = fill_pct_cell  # Juda och ko'k fon
            cell.alignment = align_center
            cell.border = thin_border

        # =========================================================
        # 6. USTUN KENGLIKLARI
        # =========================================================
        dims = {
            'A': 5, 'B': 40, 'C': 12, 'D': 10, 'E': 8,
            'F': 11, 'G': 11, 'H': 11, 'I': 11,
            'J': 11, 'K': 11, 'L': 11, 'M': 11,
            'N': 11, 'O': 18, 'P': 18, 'Q': 18
        }
        for col, width in dims.items():
            ws.column_dimensions[col].width = width

        # =========================================================
        # 7. RESPONSE
        # =========================================================
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response[
            'Content-Disposition'] = f'attachment; filename=Kontrakt_Swod_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
        wb.save(response)
        return response

    def _get_subject_debt_swod_data(self, request):
        """
        Ma'lumotlarni yig'ish: Yil (Multi), Sana, Fan holati va boshqa filtrlar bilan.
        """
        # --- 1. FILTRLARNI OLISH ---

        # A) O'QUV YILI (Multi-select)
        all_years_qs = AcademicYear.objects.all().order_by('-name')
        selected_years_str = request.GET.getlist('year')

        target_years = []
        # Agar URLda year parametri bo'lmasa -> BARCHASINI olamiz
        if not request.GET.get('year') and not request.GET.get('status'):
            # (Sahifa yangi ochilganda)
            target_years = [y.id for y in all_years_qs]
            selected_years_str = [str(y.id) for y in all_years_qs]
        else:
            # Agar filtr ishlatilgan bo'lsa
            if selected_years_str:
                target_years = [int(y) for y in selected_years_str if y.isdigit()]
            else:
                # Filtr bosilib, hech narsa tanlanmasa
                target_years = []

        # B) FAN HOLATI (Yopilgan/Yopilmagan)
        selected_subject_status = request.GET.get('subject_status', 'all')

        # C) SANA (Tarixiy holat)
        limit_date_str = request.GET.get('date')
        limit_date_obj = None
        if limit_date_str:
            try:
                limit_date_obj = datetime.strptime(limit_date_str, "%Y-%m-%d").date()
            except ValueError:
                pass

        # D) BOSHQA FILTRLAR
        if request.GET:
            selected_forms = request.GET.getlist('form')
            selected_courses = request.GET.getlist('course')
            selected_statuses = request.GET.getlist('status') if 'status' in request.GET else []
        else:
            selected_forms = []
            selected_courses = []
            selected_statuses = ['active']

        # --- 2. QUERYSET (Prefetch bilan) ---

        # SubjectDebt ni filtrlaymiz
        debt_qs = SubjectDebt.objects.filter(academic_year__id__in=target_years).select_related('subject')

        # Fan holati bo'yicha filtr
        if selected_subject_status == 'closed':
            debt_qs = debt_qs.filter(status='yopildi')
        elif selected_subject_status == 'open':
            debt_qs = debt_qs.filter(status='yopilmadi')

        # Prefetch
        debt_prefetch = Prefetch(
            'subjectdebt_set',
            queryset=debt_qs,
            to_attr='current_debts'
        )

        qs = Student.objects.select_related('group', 'group__specialty').prefetch_related(debt_prefetch)

        # Talabalarni filtrlash
        if selected_statuses: qs = qs.filter(status__in=selected_statuses)
        if selected_forms: qs = qs.filter(education_form__in=selected_forms)
        if selected_courses:
            try:
                cl = [int(c) for c in selected_courses if str(c).isdigit()]
                if cl: qs = qs.filter(course_year__in=cl)
            except ValueError:
                pass

        # --- 3. HISOBLASH VA GURUHLASH ---
        data_map = {}
        grand_total = {
            'count': 0, 'debtors_count': 0, 'debt_subject_count': 0,
            'contract': 0, 'paid': 0, 'debt': 0
        }

        for student in qs:
            spec_name = student.group.specialty.name if student.group and student.group.specialty else "Noma'lum"
            form = student.education_form
            course = student.course_year

            st_calc = 0
            st_paid = 0
            st_debt = 0
            student_subjects_details = []

            # Har bir talabaning qarzlarini aylanamiz
            if hasattr(student, 'current_debts'):
                for debt_obj in student.current_debts:
                    amount = debt_obj.amount or 0
                    real_paid = debt_obj.amount_summ or 0
                    p_date = debt_obj.payment_date

                    # Tarixiy mantiq (Vaqt bo'yicha orqaga qaytish)
                    current_paid = real_paid
                    if limit_date_obj:
                        # Agar to'lov sanasi limit_date dan keyin bo'lsa -> To'lanmagan deb hisoblaymiz
                        if p_date and p_date > limit_date_obj:
                            current_paid = 0
                        # Agar to'lov sanasi yo'q bo'lsa (lekin pul bor bo'lsa), ehtiyotkorlik bilan yondashamiz
                        # (Hozircha o'zgarishsiz qoldiramiz)

                    rem = max(0, amount - current_paid)

                    st_calc += amount
                    st_paid += current_paid
                    st_debt += rem

                    student_subjects_details.append({
                        'subject': debt_obj.subject.name,
                        'amount': amount,
                        'paid': current_paid,
                        'debt': rem,
                        'is_closed': debt_obj.status == 'yopildi'
                    })

            # Map strukturasini tuzish
            if spec_name not in data_map: data_map[spec_name] = {}
            if form not in data_map[spec_name]: data_map[spec_name][form] = {}
            if course not in data_map[spec_name][form]:
                data_map[spec_name][form][course] = {
                    'count': 0, 'debtors_count': 0, 'debt_subject_count': 0,
                    'debt_subject_names': set(),
                    'contract': 0, 'paid': 0, 'debt': 0,
                    'debtor_students_list': []
                }

            item = data_map[spec_name][form][course]

            # Summalarni qo'shish
            item['count'] += 1
            grand_total['count'] += 1
            item['contract'] += st_calc
            grand_total['contract'] += st_calc
            item['paid'] += st_paid
            grand_total['paid'] += st_paid
            item['debt'] += st_debt
            grand_total['debt'] += st_debt

            # Agar qarz bo'lsa yoki "Yopilmagan" filtri tanlanganda qarz bo'lmasa ham
            # ro'yxatga qo'shish kerakmi? SWOD moliya haqida.
            # Shuning uchun faqat moliyaviy qarzi borlarni (st_debt > 0) "Qarzdor" deb hisoblaymiz.
            if st_debt > 0:
                item['debtors_count'] += 1
                grand_total['debtors_count'] += 1

                item['debtor_students_list'].append({
                    'full_name': student.full_name,
                    'hemis_id': student.student_hemis_id,
                    'total_debt': st_debt,
                    'subjects': student_subjects_details
                })

                # Qarzdor fanlarni sanash (faqat moliyaviy qarzi bor fanlar)
                open_subjects = [s['subject'] for s in student_subjects_details if s['debt'] > 0]
                if open_subjects:
                    cnt = len(open_subjects)
                    item['debt_subject_count'] += cnt
                    grand_total['debt_subject_count'] += cnt
                    item['debt_subject_names'].update(open_subjects)

        # --- 4. DATA TAYYORLASH ---
        report_data = []
        counter = 1
        form_priority = {'kunduzgi': 1, 'sirtqi': 2, 'kechki': 3}

        for spec_name, forms in sorted(data_map.items()):
            rows = []
            for form_name, courses in sorted(forms.items(), key=lambda x: form_priority.get(x[0], 10)):
                sub = {
                    'count': 0, 'debtors_count': 0, 'debt_subject_count': 0,
                    'contract': 0, 'paid': 0, 'debt': 0,
                    'debt_subject_names': set(),
                    'debtor_students_list': []
                }
                for course_num, vals in sorted(courses.items()):
                    vals['row_id'] = f"row_{counter}_{form_name}_{course_num}"
                    vals['debt_subjects_str'] = ", ".join(sorted(vals['debt_subject_names']))

                    rows.append({'form': form_name, 'course': f"{course_num}-kurs", 'vals': vals, 'is_total': False})

                    # Subtotal
                    for k in ['count', 'debtors_count', 'debt_subject_count', 'contract', 'paid', 'debt']:
                        sub[k] += vals[k]
                    sub['debt_subject_names'].update(vals['debt_subject_names'])
                    sub['debtor_students_list'].extend(vals['debtor_students_list'])

                sub['row_id'] = f"row_{counter}_{form_name}_total"
                sub['debt_subjects_str'] = ", ".join(sorted(sub['debt_subject_names']))
                rows.append({'form': form_name, 'course': 'Jami', 'vals': sub, 'is_total': True})

            report_data.append({'index': counter, 'spec_name': spec_name, 'rows': rows, 'rowspan': len(rows)})
            counter += 1

        return {
            'report_data': report_data,
            'grand_total': grand_total,
            'selected_years': selected_years_str,
            'limit_date': limit_date_str,
            'selected_subject_status': selected_subject_status,
            'years': all_years_qs,
            'forms': Student.EducationFormChoices.choices,
            'courses': [str(i) for i in range(1, 6)],
            'statuses': Student.StatusChoices.choices,
            'selected_forms': selected_forms,
            'selected_courses': selected_courses,
            'selected_statuses': selected_statuses
        }

    def subject_debt_swod_view(self, request):
        data = self._get_subject_debt_swod_data(request)
        grand_total = data['grand_total']

        # Footer foizlari
        grand_total_percent = {}
        total_contract = float(grand_total['contract'])

        if total_contract > 0:
            total_paid = float(grand_total['paid'])
            total_debt = float(grand_total['debt'])

            paid_raw = (total_paid / total_contract) * 100
            grand_total_percent['paid'] = f"{paid_raw:.1f}%"

            if paid_raw <= 100:
                debt_float = 100.0 - paid_raw
                grand_total_percent['debt'] = f"{debt_float:.1f}%"
            else:
                debt_raw = (total_debt / total_contract) * 100
                grand_total_percent['debt'] = f"{debt_raw:.1f}%"
        else:
            grand_total_percent['paid'] = "0.0%"
            grand_total_percent['debt'] = "0.0%"

        context = admin.site.each_context(request)
        context.update({
            'title': "Fan Qarzlari Swod Hisoboti",
            'report_data': data['report_data'],
            'grand_total': data['grand_total'],
            'grand_total_percent': grand_total_percent,
            'filter_options': {
                'years': data['years'],
                'forms': data['forms'],
                'courses': [(i, f"{i}-kurs") for i in range(1, 6)],
                'statuses': data['statuses']
            },
            'selected_filters': {
                'year': data['selected_years'],
                'form': data['selected_forms'],
                'course': data['selected_courses'],
                'status': data['selected_statuses'],
                'date': data['limit_date'],
                'subject_status': data['selected_subject_status'],
            },
            'current_date': datetime.now().strftime("%d.%m.%Y")
        })
        return render(request, "admin/reports/subject_debt_swod.html", context)

    def export_subject_debt_swod_excel(self, request):
        data = self._get_subject_debt_swod_data(request)
        report_data = data['report_data']
        grand_total = data['grand_total']

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Fan Qarzlari"
        ws.sheet_view.showGridLines = False

        # Styles
        header_fill = PatternFill('solid', fgColor='2C3E50')
        header_font = Font(bold=True, color='FFFFFF', name='Calibri', size=10)

        fill_green_head = PatternFill('solid', fgColor='1E8449')
        fill_blue_head = PatternFill('solid', fgColor='2874A6')
        fill_red_head = PatternFill('solid', fgColor='C0392B')

        fill_green_cell = PatternFill('solid', fgColor='D4EFDF')
        fill_blue_cell = PatternFill('solid', fgColor='D6EAF8')
        fill_red_cell = PatternFill('solid', fgColor='FADBD8')

        fill_total_row = PatternFill('solid', fgColor='ECF0F1')
        font_total = Font(bold=True, name='Calibri', size=10)

        thin_border = Border(left=Side(style='thin', color='AAAAAA'),
                             right=Side(style='thin', color='AAAAAA'),
                             top=Side(style='thin', color='AAAAAA'),
                             bottom=Side(style='thin', color='AAAAAA'))
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        money_fmt = '#,##0'

        # HEADER
        headers = [
            ('A', '№', 1, 2),
            ('B', "Ta'lim yo'nalishi", 1, 2),
            ('C', "Shakl", 1, 2),
            ('D', "Kurs", 1, 2),
            ('E', "Jami\nTalaba", 1, 2),
            ('F', "Qarzdor\nTalaba", 1, 2),
            ('G', "Qarzdor\nFanlar", 1, 2),
            ('H', "Hisoblangan\nQarz", 1, 2),
            ('I', "To'langan\nSumma", 1, 2),
            ('J', "Qoldiq\nQarz", 1, 2),
        ]

        for col_char, text, colspan, rowspan in headers:
            cell = ws[f"{col_char}2"]
            cell.value = text
            cell.font = header_font;
            cell.fill = header_fill;
            cell.alignment = align_center;
            cell.border = thin_border

            if "Hisoblangan" in text: cell.fill = fill_green_head
            if "To'langan" in text: cell.fill = fill_blue_head
            if "Qoldiq" in text: cell.fill = fill_red_head

            col_idx = openpyxl.utils.column_index_from_string(col_char)
            if col_idx == 2:
                ws.column_dimensions['B'].width = 40
            elif col_idx >= 8:
                ws.column_dimensions[col_char].width = 18
            else:
                ws.column_dimensions[col_char].width = 12

        current_row = 4
        for item in report_data:
            start_row = current_row
            ws.cell(row=current_row, column=1, value=item['index']).alignment = align_center
            ws.cell(row=current_row, column=2, value=item['spec_name']).alignment = Alignment(horizontal='left',
                                                                                              vertical='center',
                                                                                              wrap_text=True)

            for row_data in item['rows']:
                vals = row_data['vals']
                is_total = row_data.get('is_total')
                font = font_total if is_total else Font(name='Calibri')
                bg = fill_total_row if is_total else PatternFill(fill_type=None)

                for c in range(1, 11):
                    cell = ws.cell(row=current_row, column=c)
                    cell.border = thin_border
                    cell.font = font
                    if is_total: cell.fill = bg
                    if c > 2: cell.alignment = align_center

                ws.cell(row=current_row, column=3, value=row_data['form'] if not is_total else "Jami")
                ws.cell(row=current_row, column=4, value=row_data['course'] if not is_total else "-")
                ws.cell(row=current_row, column=5, value=vals['count'])

                d_cell = ws.cell(row=current_row, column=6, value=vals['debtors_count'])
                if vals['debtors_count'] > 0: d_cell.font = Font(bold=True, color='C0392B')

                ws.cell(row=current_row, column=7, value=vals['debt_subject_count'])

                c_cell = ws.cell(row=current_row, column=8, value=vals['contract']);
                c_cell.number_format = money_fmt
                p_cell = ws.cell(row=current_row, column=9, value=vals['paid']);
                p_cell.number_format = money_fmt
                db_cell = ws.cell(row=current_row, column=10, value=vals['debt']);
                db_cell.number_format = money_fmt

                if is_total:
                    c_cell.fill = fill_green_cell
                    p_cell.fill = fill_blue_cell
                    db_cell.fill = fill_red_cell

                current_row += 1

            if item['rowspan'] > 1:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=current_row - 1, end_column=1)
                ws.merge_cells(start_row=start_row, start_column=2, end_row=current_row - 1, end_column=2)
                ws.cell(row=start_row, column=1).alignment = align_center
                ws.cell(row=start_row, column=2).alignment = Alignment(horizontal='left', vertical='center',
                                                                       wrap_text=True)

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        ws.cell(row=current_row, column=1, value="UMUMIY JAMI").alignment = Alignment(horizontal='right')

        for c in range(1, 5):
            ws.cell(row=current_row, column=c).fill = header_fill;
            ws.cell(row=current_row, column=c).font = header_font;
            ws.cell(row=current_row, column=c).border = thin_border

        totals = [grand_total['count'], grand_total['debtors_count'], grand_total['debt_subject_count'],
                  grand_total['contract'], grand_total['paid'], grand_total['debt']]
        for idx, val in enumerate(totals, 5):
            cell = ws.cell(row=current_row, column=idx, value=val)
            cell.font = header_font;
            cell.fill = header_fill;
            cell.alignment = align_center;
            cell.border = thin_border
            if idx >= 8: cell.number_format = money_fmt

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Fan_Qarzi_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
        wb.save(response)
        return response

    def tsch_analiz_view(self, request):
        # 1. Barcha mavjud o'quv yillari
        all_years = AcademicYear.objects.all().order_by('-name')
        active_year = AcademicYear.objects.filter(is_active=True).first()

        # 2. URL parametrlar
        selected_year_ids = request.GET.getlist('year')
        active_tab = request.GET.get('active_tab', 'TabKunduzgi')
        view_mode = request.GET.get('view_mode', 'separate')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        if not selected_year_ids:
            if active_year:
                selected_year_ids = [str(active_year.id)]
            elif all_years.exists():
                selected_year_ids = [str(all_years.first().id)]
            else:
                selected_year_ids = []

        selected_year_ids_int = [int(x) for x in selected_year_ids if str(x).isdigit()]

        forms_config = {
            'kunduzgi': {'courses': [1, 2, 3, 4], 'label': "Kunduzgi"},
            'sirtqi': {'courses': [1, 2, 3, 4, 5], 'label': "Sirtqi"},
        }

        reports_list = []

        # STATISTIKA UCHUN O'ZGARUVCHILAR
        global_reasons_map = defaultdict(int)
        du_reasons_map = defaultdict(int)
        student_reasons_map = defaultdict(int)
        global_initiator_map = defaultdict(int)
        global_total_lost_cnt = 0
        du_total_cnt = 0
        student_total_cnt = 0

        initiator_labels = dict(Order.TschChoices.choices)

        # --- AGAR UMUMIY BO'LSA, MA'LUMOTNI SHU YERDA YIG'AMIZ ---
        is_combined = (view_mode == 'general')
        combined_storage = {'kunduzgi': {}, 'sirtqi': {}} if is_combined else None

        for year_id in selected_year_ids:
            target_admission_year = AcademicYear.objects.filter(id=year_id).first()
            if not target_admission_year:
                continue

            # Joriy kursni hisoblash (Faqat alohida rejimda ishlatiladi)
            current_course_num = 0
            if not is_combined and active_year and target_admission_year:
                try:
                    act_y = int(active_year.name[:4])
                    adm_y = int(target_admission_year.name[:4])
                    current_course_num = act_y - adm_y + 1
                except:
                    pass

            # Yillarni tartiblash
            sorted_years = list(AcademicYear.objects.filter(name__gte=target_admission_year.name).order_by('name'))
            course_year_map = {}
            for i in range(1, 6):
                if (i - 1) < len(sorted_years):
                    course_year_map[i] = sorted_years[i - 1].id
                else:
                    course_year_map[i] = None

            # 1-kursga qabul qilinganlar (Kogorta)
            cohort_qs = StudentHistory.objects.filter(
                academic_year=target_admission_year,
                course_year=1
            ).select_related('student', 'group__specialty')

            cohort_ids = list(cohort_qs.values_list('student_id', flat=True))

            if cohort_ids:
                all_history = StudentHistory.objects.filter(student_id__in=cohort_ids)
                history_map = {(h.student_id, h.course_year): h for h in all_history}

                # Shartnomalar va To'lovlar
                all_contracts = Contract.objects.filter(student_id__in=cohort_ids).values('student_id',
                                                                                          'academic_year_id').annotate(
                    sum_amount=Sum('amount'))
                contract_map = {(c['student_id'], c['academic_year_id']): (c['sum_amount'] or 0) for c in all_contracts}

                all_payments = Payment.objects.filter(contract__student_id__in=cohort_ids).values(
                    'contract__student_id', 'contract__academic_year_id').annotate(sum_pay=Sum('amount'))
                payment_map = {(p['contract__student_id'], p['contract__academic_year_id']): (p['sum_pay'] or 0) for p
                               in all_payments}

                # Ketganlar buyruqlari
                expulsion_orders_qs = Order.objects.filter(student_id__in=cohort_ids).filter(
                    Q(order_type__name__icontains="chetlash") |
                    Q(order_type__name__icontains="chiqarish") |
                    Q(order_type__name__icontains="safidan") |
                    Q(order_type__name__icontains="expel")
                )

                # --- SANA BO'YICHA FILTERLASH ---
                if start_date:
                    expulsion_orders_qs = expulsion_orders_qs.filter(order_date__gte=start_date)
                if end_date:
                    expulsion_orders_qs = expulsion_orders_qs.filter(order_date__lte=end_date)
                # --------------------------------

                lost_map = defaultdict(list)
                for o in expulsion_orders_qs.values('student_id', 'order_date'):
                    lost_map[o['student_id']].append(o['order_date'])

                graduated_students = set(
                    Student.objects.filter(id__in=cohort_ids, status='graduated').values_list('id', flat=True))

                # STATISTIKA YIG'ISH (Bu qism doimiy ishlayveradi)
                orders_data = expulsion_orders_qs.values('tsch_reason', 'tsch_by_whom')
                for item in orders_data:
                    raw_reason = item['tsch_reason']
                    initiator_code = item['tsch_by_whom']
                    reason_name = str(raw_reason).strip().capitalize() if raw_reason and str(
                        raw_reason).strip() else "Sababi ko'rsatilmagan"

                    global_reasons_map[reason_name] += 1
                    global_total_lost_cnt += 1

                    initiator_label = initiator_labels.get(initiator_code, initiator_code) or "Ko'rsatilmagan"
                    global_initiator_map[initiator_label] += 1

                    if initiator_code == Order.TschChoices.DU:
                        du_reasons_map[reason_name] += 1
                        du_total_cnt += 1
                    elif initiator_code == Order.TschChoices.STUDENTS:
                        student_reasons_map[reason_name] += 1
                        student_total_cnt += 1

                # --- ASOSIY MA'LUMOT YIG'ISH ---
                current_storage = combined_storage if is_combined else {'kunduzgi': {}, 'sirtqi': {}}

                for h in cohort_qs:
                    if not h.group or not h.group.specialty: continue
                    form_key = h.education_form
                    if form_key not in current_storage: continue

                    spec_name = h.group.specialty.name
                    st_id = h.student_id

                    if spec_name not in current_storage[form_key]:
                        courses_list = forms_config[form_key]['courses']
                        current_storage[form_key][spec_name] = {
                            c: {'count': 0, 'amount': 0, 'lost': 0, 'lost_amount': 0, 'lost_ids': []}
                            for c in courses_list
                        }
                        current_storage[form_key][spec_name]['grad'] = {'graduated': 0, 'failed': 0}

                    previous_lost = False
                    target_courses = forms_config[form_key]['courses']

                    for course in target_courses:
                        if previous_lost: break
                        mapped_year_id = course_year_map.get(course)

                        # Mantiq: 1-kursda hamma bor. Keyingi kurslarda bazada bo'lsa bor.
                        is_present = False
                        if course == 1:
                            is_present = True
                        else:
                            hist = history_map.get((st_id, course))
                            contract_amt = 0
                            if mapped_year_id:
                                contract_amt = contract_map.get((st_id, mapped_year_id), 0)

                            if hist or contract_amt > 0:
                                is_present = True

                            # Fallback (Active yil uchun)
                            if not is_present and active_year and mapped_year_id == active_year.id:
                                if h.student.status == 'active' and h.student.course_year == course:
                                    is_present = True

                        if is_present:
                            cell = current_storage[form_key][spec_name][course]

                            c_amt = 0
                            if mapped_year_id:
                                c_amt = contract_map.get((st_id, mapped_year_id), 0)
                                cell['amount'] += c_amt

                            # Ketganlik tekshiruvi (faqat order bor bo'lsa)
                            is_lost_this_year = False
                            if mapped_year_id:
                                try:
                                    ac_obj = next((y for y in sorted_years if y.id == mapped_year_id), None)
                                    if ac_obj:
                                        y_start = int(ac_obj.name[:4])
                                        d_start = date(y_start, 9, 1)
                                        d_end = date(y_start + 1, 8, 31)
                                        if st_id in lost_map:
                                            # Bu yerda lost_map allaqachon sana bo'yicha filtrlangan
                                            for ld in lost_map[st_id]:
                                                if d_start <= ld <= d_end:
                                                    is_lost_this_year = True
                                                    p_amt = payment_map.get((st_id, mapped_year_id), 0)
                                                    actual_loss = max(0, c_amt - p_amt)

                                                    cell['lost'] += 1
                                                    cell['lost_amount'] += actual_loss
                                                    cell['lost_ids'].append(st_id)
                                                    break
                                except:
                                    pass

                            if is_lost_this_year:
                                previous_lost = True
                            else:
                                cell['count'] += 1

                    if st_id in graduated_students:
                        current_storage[form_key][spec_name]['grad']['graduated'] += 1

            # --- AGAR ALOHIDA REJIM BO'LSA, DARHOL JADVALNI YASAYMIZ ---
            if not is_combined:
                year_report = {
                    'year_name': target_admission_year.name,
                    'current_course': current_course_num,
                    'kunduzgi': self._process_storage_to_rows(current_storage['kunduzgi'], 'kunduzgi', forms_config),
                    'sirtqi': self._process_storage_to_rows(current_storage['sirtqi'], 'sirtqi', forms_config)
                }
                reports_list.append(year_report)

        # --- LOOP TUGADI. AGAR UMUMIY BO'LSA, ENDI JADVAL YASAYMIZ ---
        if is_combined:
            combined_report = {
                'year_name': "Tanlangan yillar (Jami)",
                'current_course': 0,
                'kunduzgi': self._process_storage_to_rows(combined_storage['kunduzgi'], 'kunduzgi', forms_config),
                'sirtqi': self._process_storage_to_rows(combined_storage['sirtqi'], 'sirtqi', forms_config)
            }
            if combined_report['kunduzgi']['rows'] or combined_report['sirtqi']['rows']:
                reports_list.append(combined_report)

        # Statistikalar
        def prepare_stat(data_map, total_count):
            stat_list = []
            for r_name, count in data_map.items():
                stat_list.append({
                    'name': r_name, 'count': count,
                    'percent': (count / total_count * 100) if total_count > 0 else 0
                })
            stat_list.sort(key=lambda x: x['count'], reverse=True)
            return stat_list

        reasons_stat = prepare_stat(global_reasons_map, global_total_lost_cnt)
        du_reasons_stat = prepare_stat(du_reasons_map, du_total_cnt)
        student_reasons_stat = prepare_stat(student_reasons_map, student_total_cnt)
        initiator_stat = prepare_stat(global_initiator_map, global_total_lost_cnt)

        context = admin.site.each_context(request)
        context.update({
            'title': "TSCH Analiz (Kogorta)",
            'all_years': all_years,
            'selected_year_ids': selected_year_ids_int,
            'reports_list': reports_list,
            'active_tab': active_tab,
            'view_mode': view_mode,
            'start_date': start_date,
            'end_date': end_date,
            'reasons_stat': reasons_stat,
            'du_reasons_stat': du_reasons_stat,
            'student_reasons_stat': student_reasons_stat,
            'initiator_stat': initiator_stat,
            'stats_totals': {'du': du_total_cnt, 'student': student_total_cnt, 'global': global_total_lost_cnt}
        })
        return render(request, 'admin/reports/tsch_analiz.html', context)

    # --- YORDAMCHI FUNKSIYA (Rowlarni yasash uchun) ---
    def _process_storage_to_rows(self, specs_data, f_key, forms_config):
        if not specs_data:
            return {'rows': [], 'grand_total': None, 'global_stats': None}

        rows = []
        idx = 1
        target_courses = forms_config[f_key]['courses']

        grand_total = {
            'courses': {c: {'count': 0, 'amount': 0, 'lost': 0, 'display_lost_amt': 0} for c in target_courses},
            'grad': {'graduated': 0}
        }

        for spec, data in sorted(specs_data.items()):
            row = {'index': idx, 'spec': spec, 'cells': [], 'grad': {}}

            for c in target_courses:
                current_data = data[c]

                # Mantiq: count - bu hozirda Active bo'lganlar (yoki shu kursni bitirganlar).
                # Lost - bu shu kurs davomida ketganlar.
                # Demak, kurs boshida: total_at_start = Active + Lost
                total_at_start = current_data['count'] + current_data['lost']
                lost_cnt = current_data['lost']

                percent = (lost_cnt / total_at_start * 100) if total_at_start > 0 else 0

                cell = {
                    'count': total_at_start,  # Talaba soni (boshlanishida)
                    'amount': current_data['amount'],
                    'show_loss': True,
                    'lost': lost_cnt,
                    'lost_amount': current_data['lost_amount'],
                    'percent': percent,
                    'lost_ids': current_data['lost_ids']
                }
                row['cells'].append(cell)

                grand_total['courses'][c]['count'] += cell['count']
                grand_total['courses'][c]['amount'] += cell['amount']
                grand_total['courses'][c]['lost'] += cell['lost']
                grand_total['courses'][c]['display_lost_amt'] += cell['lost_amount']

            grad_cnt = data['grad']['graduated']
            row['grad'] = {'graduated': grad_cnt}
            grand_total['grad']['graduated'] += grad_cnt

            rows.append(row)
            idx += 1

        grand_total_list = [grand_total['courses'][c] for c in target_courses]

        gl_contract = sum(grand_total['courses'][c]['amount'] for c in target_courses)
        gl_lost = sum(grand_total['courses'][c]['display_lost_amt'] for c in target_courses)

        return {
            'rows': rows,
            'grand_total': {'cells': grand_total_list, 'grad': grand_total['grad']},
            'global_stats': {
                'total_contract': gl_contract,
                'total_lost': gl_lost,
                'lost_percent': (gl_lost / gl_contract * 100) if gl_contract > 0 else 0
            }
        }

    def export_tsch_analiz_excel(self, request):
        selected_year_ids = request.GET.getlist('year')
        selected_parts = request.GET.getlist('parts')
        view_mode = request.GET.get('view_mode', 'separate')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        # Agar yil tanlanmagan bo'lsa, active yilni olamiz
        if not selected_year_ids:
            active_year = AcademicYear.objects.filter(is_active=True).first()
            if active_year:
                selected_year_ids = [str(active_year.id)]
            else:
                selected_year_ids = []

        # Excel yaratish
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TSCH Analiz"
        ws.sheet_view.showGridLines = False

        # --- STYLES (DIZAYN) ---
        fill_header = PatternFill("solid", fgColor="E9ECEF")
        fill_course = PatternFill("solid", fgColor="CED4DA")
        fill_grad = PatternFill("solid", fgColor="D4EDDA")
        fill_loss = PatternFill("solid", fgColor="FFEBEE")

        font_header = Font(bold=True, color="495057", size=10)
        font_grad_title = Font(bold=True, color="155724", size=10)
        font_loss = Font(bold=True, color="C62828")
        font_bold = Font(bold=True)
        font_white_bold = Font(bold=True, color="FFFFFF", size=14)

        fill_dark_blue = PatternFill("solid", fgColor="2C3E50")

        border = Border(left=Side(style='thin', color='DEE2E6'), right=Side(style='thin', color='DEE2E6'),
                        top=Side(style='thin', color='DEE2E6'), bottom=Side(style='thin', color='DEE2E6'))

        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        money_fmt = '#,##0'

        current_row = 1

        # Statistika uchun global o'zgaruvchilar
        global_reasons_map = defaultdict(int)
        global_initiator_map = defaultdict(int)
        global_total_lost_cnt = 0
        global_du_reasons = defaultdict(int)
        global_student_reasons = defaultdict(int)
        global_du_cnt = 0
        global_st_cnt = 0

        # --- UMUMIY REJIM UCHUN YIG'UVCHI O'ZGARUVCHI ---
        general_storage = {
            'kunduzgi': {},
            'sirtqi': {}
        }

        # Qaysi formalar kerakligini aniqlaymiz
        forms_to_export = []
        if 'kunduzgi' in selected_parts: forms_to_export.append(('kunduzgi', [1, 2, 3, 4]))
        if 'sirtqi' in selected_parts: forms_to_export.append(('sirtqi', [1, 2, 3, 4, 5]))

        # --- YILLAR BO'YICHA LOOP ---
        for year_id in selected_year_ids:
            target_year = AcademicYear.objects.filter(id=year_id).first()
            if not target_year: continue

            # Yordamchi ma'lumotlarni tayyorlash
            sorted_years = list(AcademicYear.objects.filter(name__gte=target_year.name).order_by('name'))
            course_year_map = {i: (sorted_years[i - 1].id if (i - 1) < len(sorted_years) else None) for i in
                               range(1, 6)}

            # Kogortani olish (1-kursga kirganlar)
            cohort_qs = StudentHistory.objects.filter(academic_year=target_year, course_year=1).select_related(
                'student', 'group__specialty')
            cohort_ids = list(cohort_qs.values_list('student_id', flat=True))

            if not cohort_ids:
                if view_mode == 'separate' and forms_to_export:
                    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
                    ws.cell(row=current_row, column=1, value=f"Qabul yili: {target_year.name} - Ma'lumot yo'q")
                    current_row += 2
                continue

            all_history = StudentHistory.objects.filter(student_id__in=cohort_ids)
            history_map = {(h.student_id, h.course_year): h for h in all_history}

            # Kontraktlar
            all_contracts = Contract.objects.filter(student_id__in=cohort_ids).values('student_id',
                                                                                      'academic_year_id').annotate(
                s=Sum('amount'))
            contract_map = {(c['student_id'], c['academic_year_id']): (c['s'] or 0) for c in all_contracts}

            # To'lovlar
            all_payments = Payment.objects.filter(contract__student_id__in=cohort_ids).values('contract__student_id',
                                                                                              'contract__academic_year_id').annotate(
                sum_pay=Sum('amount'))
            payment_map = {(p['contract__student_id'], p['contract__academic_year_id']): (p['sum_pay'] or 0) for p in
                           all_payments}

            # Buyruqlar (Ketganlar)
            expulsion_orders = Order.objects.filter(student_id__in=cohort_ids).filter(
                Q(order_type__name__icontains="chetlash") | Q(order_type__name__icontains="chiqarish") |
                Q(order_type__name__icontains="safidan") | Q(order_type__name__icontains="expel")
            )

            # --- SANA BO'YICHA FILTERLASH ---
            if start_date:
                expulsion_orders = expulsion_orders.filter(order_date__gte=start_date)
            if end_date:
                expulsion_orders = expulsion_orders.filter(order_date__lte=end_date)
            # --------------------------------

            # Statistika yig'ish (Ikkala rejimda ham ishlayveradi)
            for order in expulsion_orders:
                r = str(order.tsch_reason).strip().capitalize() if order.tsch_reason else "Sababi ko'rsatilmagan"
                w = order.tsch_by_whom
                global_reasons_map[r] += 1
                global_total_lost_cnt += 1
                label = dict(Order.TschChoices.choices).get(w, w) or "Ko'rsatilmagan"
                global_initiator_map[label] += 1
                if w == 'du':
                    global_du_reasons[r] += 1
                    global_du_cnt += 1
                elif w == 'student':
                    global_student_reasons[r] += 1
                    global_st_cnt += 1

            # Ketgan sanalar xaritasi
            lost_map = defaultdict(list)
            for o in expulsion_orders.values('student_id', 'order_date'):
                lost_map[o['student_id']].append(o['order_date'])

            graduated_students = set(
                Student.objects.filter(id__in=cohort_ids, status='graduated').values_list('id', flat=True))

            # --- MA'LUMOTLARNI HISOBLASH (FORMALAR KESIMIDA) ---
            for f_key, courses_list in forms_to_export:

                # Agar SEPARATE bo'lsa, Sarlavha yozamiz
                if view_mode == 'separate':
                    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=20)
                    title_cell = ws.cell(row=current_row, column=1,
                                         value=f"Qabul yili: {target_year.name} ({f_key.upper()})")
                    title_cell.font = font_white_bold
                    title_cell.fill = fill_dark_blue
                    title_cell.alignment = align_center
                    current_row += 2

                    current_row = self._draw_excel_headers(ws, current_row, courses_list, fill_course, fill_grad,
                                                           fill_header, font_header, font_grad_title, font_loss, border,
                                                           align_center)

                # HISOBLASH LOGIKASI
                temp_grouped = {}  # Bu yil uchun lokal

                for h in cohort_qs:
                    if h.education_form != f_key: continue
                    spec = h.group.specialty.name
                    st_id = h.student_id

                    if spec not in temp_grouped:
                        temp_grouped[spec] = {c: {'cnt': 0, 'amt': 0, 'lost_cnt': 0, 'lost_amt': 0} for c in
                                              courses_list}
                        temp_grouped[spec]['grad'] = 0

                    prev_lost = False
                    for c in courses_list:
                        if prev_lost: break
                        y_id = course_year_map.get(c)
                        exists = False
                        if c == 1:
                            exists = True
                        else:
                            if (st_id, c) in history_map:
                                exists = True
                            elif y_id and (st_id, y_id) in contract_map:
                                exists = True

                        if exists:
                            c_amt = contract_map.get((st_id, y_id), 0) if y_id else 0
                            is_lost_this_year = False

                            if y_id:
                                try:
                                    ac_obj = next((y for y in sorted_years if y.id == y_id), None)
                                    if ac_obj:
                                        ystart = int(ac_obj.name[:4])
                                        d1, d2 = date(ystart, 9, 1), date(ystart + 1, 8, 31)
                                        if st_id in lost_map:
                                            for ld in lost_map[st_id]:
                                                if d1 <= ld <= d2:
                                                    is_lost_this_year = True
                                                    p_amt = payment_map.get((st_id, y_id), 0)
                                                    actual_loss = max(0, c_amt - p_amt)

                                                    temp_grouped[spec][c]['lost_cnt'] += 1
                                                    temp_grouped[spec][c]['lost_amt'] += actual_loss
                                                    temp_grouped[spec][c]['amt'] += c_amt
                                                    prev_lost = True
                                                    break
                                except:
                                    pass

                            if not is_lost_this_year:
                                temp_grouped[spec][c]['cnt'] += 1  # Active count
                                temp_grouped[spec][c]['amt'] += c_amt

                    if st_id in graduated_students: temp_grouped[spec]['grad'] += 1

                # --- AGAR SEPARATE BO'LSA - DARHOL YOZAMIZ ---
                if view_mode == 'separate':
                    current_row = self._write_excel_rows(ws, current_row, temp_grouped, courses_list, border, font_loss,
                                                         font_grad_title, font_bold, money_fmt, align_center)
                    current_row += 2  # Yillar orasida joy tashlash

                # --- AGAR GENERAL BO'LSA - YIG'AMIZ ---
                else:
                    target_store = general_storage[f_key]
                    for spec, data in temp_grouped.items():
                        if spec not in target_store:
                            target_store[spec] = {c: {'cnt': 0, 'amt': 0, 'lost_cnt': 0, 'lost_amt': 0} for c in
                                                  courses_list}
                            target_store[spec]['grad'] = 0

                        target_store[spec]['grad'] += data['grad']
                        for c in courses_list:
                            s_cell = target_store[spec][c]
                            d_cell = data[c]
                            s_cell['cnt'] += d_cell['cnt']
                            s_cell['amt'] += d_cell['amt']
                            s_cell['lost_cnt'] += d_cell['lost_cnt']
                            s_cell['lost_amt'] += d_cell['lost_amt']

        # --- AGAR GENERAL REJIM BO'LSA - ENDI YOZAMIZ ---
        if view_mode == 'general':
            # Sarlavha (Bir marta)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=20)
            title_cell = ws.cell(row=current_row, column=1, value="Tanlangan yillar bo'yicha YIG'MA hisobot")
            title_cell.font = font_white_bold
            title_cell.fill = fill_dark_blue
            title_cell.alignment = align_center
            current_row += 2

            for f_key, courses_list in forms_to_export:
                data_map = general_storage[f_key]
                if not data_map: continue

                # Sarlavha (Kunduzgi/Sirtqi)
                ws.cell(row=current_row, column=1, value=f"{f_key.upper()} TA'LIM").font = Font(bold=True, size=12)
                current_row += 1

                # Headerlar
                current_row = self._draw_excel_headers(ws, current_row, courses_list, fill_course, fill_grad,
                                                       fill_header, font_header, font_grad_title, font_loss, border,
                                                       align_center)

                # Qatorlar
                current_row = self._write_excel_rows(ws, current_row, data_map, courses_list, border, font_loss,
                                                     font_grad_title, font_bold, money_fmt, align_center)
                current_row += 2

        # --- YIG'MA STATISTIKA (PASTKI QISM) ---
        if 'stats' in selected_parts and (global_total_lost_cnt > 0):
            if current_row == 1: current_row = 1
            ws.cell(row=current_row, column=1, value="YIG'MA STATISTIKA").font = Font(bold=True, size=14);
            current_row += 2

            def draw_stat_table(title, data_map, total, start_col, r_start):
                r = r_start
                ws.merge_cells(start_row=r, start_column=start_col, end_row=r, end_column=start_col + 2)
                cell = ws.cell(row=r, column=start_col, value=title);
                cell.font = font_bold;
                cell.alignment = align_center;
                cell.fill = PatternFill("solid", fgColor="E9ECEF");
                cell.border = border;
                r += 1
                headers = ["Sabab", "Soni", "Ulushi"]
                for i, h in enumerate(headers):
                    cell = ws.cell(row=r, column=start_col + i, value=h);
                    cell.font = font_bold;
                    cell.border = border;
                    cell.alignment = align_center
                r += 1
                sorted_items = sorted(data_map.items(), key=lambda x: x[1], reverse=True)
                for k, v in sorted_items:
                    ws.cell(row=r, column=start_col, value=k).border = border
                    ws.cell(row=r, column=start_col + 1, value=v).border = border
                    pct = (v / total * 100) if total > 0 else 0
                    ws.cell(row=r, column=start_col + 2, value=f"{pct:.1f}%").border = border;
                    r += 1
                ws.cell(row=r, column=start_col, value="JAMI").font = font_bold;
                ws.cell(row=r, column=start_col).border = border
                ws.cell(row=r, column=start_col + 1, value=total).font = font_bold;
                ws.cell(row=r, column=start_col + 1).border = border
                ws.cell(row=r, column=start_col + 2, value="100%").font = font_bold;
                ws.cell(row=r, column=start_col + 2).border = border

            draw_stat_table("1. DU tashabbusi", global_du_reasons, global_du_cnt, 2, current_row)
            draw_stat_table("2. Talaba tashabbusi", global_student_reasons, global_st_cnt, 6, current_row)
            current_row += max(len(global_du_reasons), len(global_student_reasons)) + 5
            draw_stat_table("3. Tashabbuskor bo'yicha", global_initiator_map, global_total_lost_cnt, 6, current_row)
            draw_stat_table("4. Barcha sabablar", global_reasons_map, global_total_lost_cnt, 2, current_row)

        # Ustun kengliklari
        ws.column_dimensions['A'].width = 5;
        ws.column_dimensions['B'].width = 30
        for i in range(3, 150): ws.column_dimensions[get_column_letter(i)].width = 13

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        fname = f'TSCH_Analiz_{view_mode}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={fname}'
        wb.save(response)
        return response

    # --- YORDAMCHI METOD: JADVAL HEADERLARINI CHIZISH ---
    def _draw_excel_headers(self, ws, current_row, courses_list, fill_course, fill_grad, fill_header, font_header,
                            font_grad_title, font_loss, border, align_center):
        # HEADER QATOR 1: KURSLAR
        col_ptr = 3
        for c_num in courses_list:
            ws.merge_cells(start_row=current_row, start_column=col_ptr, end_row=current_row, end_column=col_ptr + 4)
            cell = ws.cell(row=current_row, column=col_ptr, value=f"{c_num}-KURS")
            cell.fill = fill_course;
            cell.alignment = align_center;
            cell.border = border
            col_ptr += 5

        # Bitiruv Header
        ws.merge_cells(start_row=current_row, start_column=col_ptr, end_row=current_row + 1, end_column=col_ptr)
        c_grad = ws.cell(row=current_row, column=col_ptr, value="BITIRGANLAR\nSONI")
        c_grad.fill = fill_grad;
        c_grad.font = font_grad_title;
        c_grad.alignment = align_center;
        c_grad.border = border

        # Asosiy Headerlar
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 1, end_column=1)
        ws.cell(row=current_row, column=1, value="№").alignment = align_center;
        ws.cell(row=current_row, column=1).border = border
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row + 1, end_column=2)
        ws.cell(row=current_row, column=2, value="Ta'lim yo'nalishi").alignment = align_center;
        ws.cell(row=current_row, column=2).border = border

        current_row += 1

        # HEADER QATOR 2: USTUN NOMALARI
        headers = []
        for _ in courses_list:
            headers.extend(["Talaba soni", "Ketgan soni", "Foizi", "Shartnoma", "Yo'qotilgan summa"])

        for i, h in enumerate(headers, 3):
            cell = ws.cell(row=current_row, column=i, value=h)
            cell.fill = fill_header;
            cell.font = font_header;
            cell.alignment = align_center;
            cell.border = border
            if "Ketgan" in h or "Yo'qotilgan" in h or "Foiz" in h: cell.font = font_loss

        return current_row + 1

    # --- YORDAMCHI METOD: MA'LUMOTLARNI YOZISH ---
    def _write_excel_rows(self, ws, current_row, temp_grouped, courses_list, border, font_loss, font_grad_title,
                          font_bold, money_fmt, align_center):
        idx = 1
        grand_totals = defaultdict(int)

        for spec, data in sorted(temp_grouped.items()):
            ws.cell(row=current_row, column=1, value=idx).border = border
            ws.cell(row=current_row, column=2, value=spec).border = border
            col = 3

            for c in courses_list:
                d = data[c]
                total_students_in_year = d['cnt'] + d['lost_cnt']

                # 1. Talaba soni (Jami)
                ws.cell(row=current_row, column=col, value=total_students_in_year).border = border
                grand_totals[col] += total_students_in_year
                col += 1
                # 2. Ketgan soni
                cell = ws.cell(row=current_row, column=col, value=d['lost_cnt'])
                cell.border = border;
                cell.font = font_loss
                grand_totals[col] += d['lost_cnt']
                col += 1
                # 3. Foizi
                pct = (d['lost_cnt'] / total_students_in_year * 100) if total_students_in_year > 0 else 0
                cell = ws.cell(row=current_row, column=col, value=f"{pct:.1f}%")
                cell.border = border;
                cell.font = font_loss
                col += 1
                # 4. Shartnoma summasi
                cell = ws.cell(row=current_row, column=col, value=d['amt'])
                cell.border = border;
                cell.number_format = money_fmt
                grand_totals[col] += d['amt']
                col += 1
                # 5. Yo'qotilgan summa
                cell = ws.cell(row=current_row, column=col, value=d['lost_amt'])
                cell.border = border;
                cell.number_format = money_fmt;
                cell.font = font_loss
                grand_totals[col] += d['lost_amt']
                col += 1

            # Bitiruv ustuni
            cell = ws.cell(row=current_row, column=col, value=data['grad'])
            cell.border = border;
            cell.font = font_grad_title
            grand_totals[col] += data['grad']

            idx += 1
            current_row += 1

        # FOOTER (JAMI)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        ws.cell(row=current_row, column=1, value="JAMI:").alignment = Alignment(horizontal='right', vertical='center');
        ws.cell(row=current_row, column=1).font = font_bold

        col = 3
        for _ in courses_list:
            # Soni
            ws.cell(row=current_row, column=col, value=grand_totals[col]).border = border;
            ws.cell(row=current_row, column=col).font = font_bold;
            col += 1
            # Ketgan
            ws.cell(row=current_row, column=col, value=grand_totals[col]).border = border;
            ws.cell(row=current_row, column=col).font = font_bold;
            col += 1
            # Foiz
            total_st = grand_totals[col - 2]
            total_ls = grand_totals[col - 1]
            t_pct = (total_ls / total_st * 100) if total_st > 0 else 0
            ws.cell(row=current_row, column=col, value=f"{t_pct:.1f}%").border = border;
            ws.cell(row=current_row, column=col).font = font_loss;
            col += 1
            # Shartnoma
            c_amt = ws.cell(row=current_row, column=col, value=grand_totals[col])
            c_amt.border = border;
            c_amt.font = font_bold;
            c_amt.number_format = money_fmt;
            col += 1
            # Yo'q summa
            l_amt = ws.cell(row=current_row, column=col, value=grand_totals[col])
            l_amt.border = border;
            l_amt.font = font_loss;
            l_amt.number_format = money_fmt;
            col += 1

        # Grad Jami
        g_cell = ws.cell(row=current_row, column=col, value=grand_totals[col])
        g_cell.border = border;
        g_cell.font = font_grad_title

        return current_row + 1

    def internal_grant_view(self, request):
        # 1. BARCHA OPTIONLARNI OLISH
        all_years = AcademicYear.objects.all().order_by('-name')
        all_groups = Group.objects.select_related('specialty').all().order_by('name')
        grant_types = Contract.GrantTypeChoices.choices[1:]
        courses = [(1, '1-kurs'), (2, '2-kurs'), (3, '3-kurs'), (4, '4-kurs'), (5, '5-kurs')]

        # 2. URL PARAMETRLARNI OLISH
        selected_years = request.GET.getlist('year')
        selected_grant_types = request.GET.getlist('grant_type')
        selected_courses = request.GET.getlist('course')
        selected_groups = request.GET.getlist('group')
        selected_date = request.GET.get('date')  # <--- YANGI: Sana parametri

        # Default yil
        if not selected_years and not request.GET:
            active_year = AcademicYear.objects.filter(is_active=True).first()
            if active_year:
                selected_years = [str(active_year.id)]

        # 3. QUERYSET TAYYORLASH
        # Faqat granti bor shartnomalar
        qs = Contract.objects.exclude(grant_type='none').select_related(
            'student', 'student__group', 'student__group__specialty', 'academic_year'
        )

        # --- FILTRLAR ---
        if selected_years:
            qs = qs.filter(academic_year_id__in=selected_years)

        if selected_grant_types:
            qs = qs.filter(grant_type__in=selected_grant_types)

        if selected_courses:
            try:
                c_ints = [int(c) for c in selected_courses]
                qs = qs.filter(student__course_year__in=c_ints)
            except ValueError:
                pass

        if selected_groups:
            qs = qs.filter(student__group_id__in=selected_groups)

        # --- YANGI: SANA BO'YICHA FILTR ---
        # Agar sana tanlangan bo'lsa, faqat o'sha sanagacha berilgan grantlarni ko'rsatamiz.
        # Bizda grant_date bor, agar u bo'sh bo'lsa contract_date ga qaraymiz.
        if selected_date:
            qs = qs.filter(
                Q(grant_date__lte=selected_date) |
                Q(grant_date__isnull=True, contract_date__lte=selected_date)
            )

        # 4. STATISTIKA HISOBLASH
        total_students = qs.count()
        total_grant_sum = qs.aggregate(sum=Sum('grant_amount'))['sum'] or 0

        # A) Grant turlari bo'yicha
        by_type_raw = qs.values('grant_type').annotate(
            count=Count('id'),
            sum=Sum('grant_amount')
        ).order_by('-sum')

        by_type_stats = []
        grant_choices = dict(Contract.GrantTypeChoices.choices)

        for item in by_type_raw:
            g_code = item['grant_type']
            by_type_stats.append({
                'code': g_code,
                'name': grant_choices.get(g_code, g_code),
                'count': item['count'],
                'sum': item['sum'],
                'percent': (item['count'] / total_students * 100) if total_students > 0 else 0
            })

        # B) Batafsil ro'yxat va Yo'nalishlar
        detailed_list = []
        spec_stats = {}

        for contract in qs:
            student = contract.student
            spec_name = student.group.specialty.name if student.group and student.group.specialty else "Noma'lum"
            g_type_display = contract.get_grant_type_display()

            detailed_list.append({
                'full_name': student.full_name,
                'hemis_id': student.student_hemis_id,
                'group': student.group.name if student.group else "-",
                'spec': spec_name,
                'course': student.course_year,
                'grant_type': g_type_display,
                'percent': contract.grant_percent,
                'grant_amount': contract.grant_amount,
                'contract_amount': contract.amount
            })

            if spec_name not in spec_stats:
                spec_stats[spec_name] = {'count': 0, 'sum': 0}
            spec_stats[spec_name]['count'] += 1
            spec_stats[spec_name]['sum'] += (contract.grant_amount or 0)

        sorted_spec_stats = sorted(spec_stats.items(), key=lambda x: x[1]['sum'], reverse=True)

        # 5. CONTEXT
        context = admin.site.each_context(request)
        context.update({
            'title': "Ichki Grant va Chegirmalar Tahlili",
            'stats': {'total_students': total_students, 'total_sum': total_grant_sum},
            'by_type_stats': by_type_stats,
            'sorted_spec_stats': sorted_spec_stats,
            'detailed_list': detailed_list,

            # Filter options
            'years': all_years,
            'grant_types': grant_types,
            'all_groups': all_groups,
            'courses': courses,

            # Selected values
            'selected_years': [str(x) for x in selected_years],
            'selected_grant_types': selected_grant_types,
            'selected_courses': [str(x) for x in selected_courses],
            'selected_groups': [str(x) for x in selected_groups],
            'selected_date': selected_date,  # <--- Templatega qaytarish

            'current_date': datetime.now().strftime("%d.%m.%Y")
        })

        return render(request, "admin/reports/internal_grant.html", context)

    def export_internal_grant_excel(self, request):
        # Parametrlarni list sifatida olish
        selected_years = request.GET.getlist('year')
        selected_grant_types = request.GET.getlist('grant_type')
        selected_courses = request.GET.getlist('course')
        selected_groups = request.GET.getlist('group')

        qs = Contract.objects.exclude(grant_type='none').select_related(
            'student', 'student__group', 'student__group__specialty', 'academic_year'
        )

        if selected_years:
            qs = qs.filter(academic_year_id__in=selected_years)
        if selected_grant_types:
            qs = qs.filter(grant_type__in=selected_grant_types)
        if selected_courses:
            try:
                c_ints = [int(c) for c in selected_courses]
                qs = qs.filter(student__course_year__in=c_ints)
            except ValueError:
                pass
        if selected_groups:
            qs = qs.filter(student__group_id__in=selected_groups)

        # Excel yaratish
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ichki Grantlar"
        ws.sheet_view.showGridLines = False

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="2C3E50")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                        bottom=Side(style='thin'))
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        money_fmt = '#,##0'

        headers = ["№", "F.I.SH", "ID", "Guruh", "Yo'nalish", "Kurs", "Grant Turi", "Foiz", "Chegirma Summasi",
                   "Asl Kontrakt"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font;
            cell.fill = header_fill;
            cell.alignment = align_center;
            cell.border = border
            w = 15
            if "F.I.SH" in h: w = 35
            if "Yo'nalish" in h: w = 30
            if "Guruh" in h: w = 15
            if "№" in h: w = 5
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

        row = 2
        total_grant = 0

        for idx, contract in enumerate(qs, 1):
            st = contract.student
            spec = st.group.specialty.name if st.group and st.group.specialty else "-"

            data = [
                idx, st.full_name, st.student_hemis_id, st.group.name if st.group else "-", spec,
                st.course_year, contract.get_grant_type_display(),
                f"{contract.grant_percent}%" if contract.grant_percent else "-",
                contract.grant_amount, contract.amount
            ]
            if contract.grant_amount: total_grant += contract.grant_amount

            for col_idx, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.border = border
                if col_idx in [9, 10]: cell.number_format = money_fmt
                if col_idx not in [2, 4, 5]: cell.alignment = align_center

            row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        total_label = ws.cell(row=row, column=1, value="JAMI CHIQIM (GRANT):")
        total_label.font = Font(bold=True, size=12);
        total_label.alignment = Alignment(horizontal='right')
        total_val = ws.cell(row=row, column=9, value=total_grant)
        total_val.font = Font(bold=True, size=12, color="C0392B");
        total_val.number_format = money_fmt;
        total_val.alignment = align_center;
        total_val.border = border

        filename = f"Ichki_Grant_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response